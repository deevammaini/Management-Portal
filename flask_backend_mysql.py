#!/usr/bin/env python3
# Flask backend for Yellowstone Management Portal with MySQL

from flask import Flask, jsonify, request, send_file, session, send_from_directory, redirect, make_response
from flask_cors import CORS
from flask_socketio import SocketIO, emit, join_room, leave_room
import os
from datetime import datetime, timedelta
from urllib.parse import unquote
import threading
import time
import mysql.connector
from mysql.connector import Error
import bcrypt
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import uuid
import random
import string
import secrets
import base64
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from io import BytesIO
import zipfile
import uuid
import secrets
import csv
import pandas as pd
from werkzeug.utils import secure_filename
import json

# Import configuration
try:
    from config_production import SECRET_KEY, FLASK_ENV, DB_CONFIG, SMTP_SERVER, SMTP_PORT, SMTP_USERNAME, SMTP_PASSWORD
except ImportError:
    # Fallback configuration for local development
    SECRET_KEY = 'dev-secret-key-change-in-production'
    FLASK_ENV = 'development'
    DB_CONFIG = {
        'host': os.environ.get('DB_HOST', 'localhost'),
        'user': os.environ.get('DB_USER', 'root'),
        'password': os.environ.get('DB_PASSWORD', 'deevammaini'),
        'database': os.environ.get('DB_NAME', 'vendor_management'),
        'port': int(os.environ.get('DB_PORT', '3306')),
        'charset': 'utf8mb4',
        'collation': 'utf8mb4_unicode_ci'
    }
    SMTP_SERVER = 'smtp.gmail.com'
    SMTP_PORT = 465
    SMTP_USERNAME = 'connect@yellowstonexperiences.com'
    SMTP_PASSWORD = 'Connect@2025'
    ADMIN_EMAIL = 'connect@yellowstonexperiences.com'

app = Flask(__name__)
CORS(app, 
     supports_credentials=True,
     origins=[
         "http://localhost:3000",
         "http://localhost:3001",
         "http://localhost:8000",
         "http://127.0.0.1:3000",
         "http://127.0.0.1:3001",
         "http://127.0.0.1:8000",
         "http://192.168.1.1:3000",
         "http://192.168.1.1:3001",
         "http://192.168.1.1:8000"
     ])

# Initialize SocketIO for real-time updates (threading mode avoids eventlet/gevent requirements)
socketio = SocketIO(app,
     cors_allowed_origins=[
         "http://localhost:3000",
         "http://localhost:3001",
         "http://localhost:8000",
         "http://127.0.0.1:3000",
         "http://127.0.0.1:3001",
         "http://127.0.0.1:8000",
         "http://192.168.1.1:3000",
         "http://192.168.1.1:3001",
         "http://192.168.1.1:8000"
     ],
     async_mode='threading',
     logger=False,
     engineio_logger=False)

# Configure session
app.secret_key = SECRET_KEY
app.config['SESSION_COOKIE_SECURE'] = FLASK_ENV == 'production'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Configure upload folder 
UPLOAD_FOLDER = 'uploads'
PROFILE_PHOTOS_FOLDER = 'profile_photos'
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROFILE_PHOTOS_FOLDER'] = PROFILE_PHOTOS_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Create upload directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROFILE_PHOTOS_FOLDER, exist_ok=True)

# Current logged-in user details
CURRENT_USER = {
    "id": None,
    "name": None,
    "email": None,
    "designation": None,
    "department": None,
    "manager": None,
    "employee_code": None,
    "user_type": None
}

def generate_reference_number():
    """Generate a unique reference number for NDA forms"""
    year = datetime.now().strftime("%Y")
    random_part = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"NDA-{year}-{random_part}"

def generate_vendor_password():
    """Generate a unique password for vendor login"""
    password_length = 12
    characters = string.ascii_letters + string.digits + "!@#$%^&*"
    password = ''.join(secrets.choice(characters) for _ in range(password_length))
    return password

def load_smtp_settings_from_db():
    """Load SMTP settings from database and update global variables"""
    global SMTP_SERVER, SMTP_PORT, SMTP_USERNAME, SMTP_PASSWORD
    try:
        query = """
        SELECT setting_key, setting_value
        FROM system_settings
        WHERE setting_category = 'email' AND setting_key IN ('smtp_server', 'smtp_port', 'smtp_username', 'smtp_password')
        """
        settings = execute_query(query, fetch_all=True)
        
        for setting in settings:
            key = setting['setting_key']
            value = setting['setting_value']
            
            if key == 'smtp_server':
                SMTP_SERVER = value if value else 'smtp.gmail.com'
            elif key == 'smtp_port':
                SMTP_PORT = int(value) if value else 587
            elif key == 'smtp_username':
                SMTP_USERNAME = value if value else ''
            elif key == 'smtp_password':
                SMTP_PASSWORD = value if value else ''
        
        # print(f"✅ SMTP settings loaded from database: {SMTP_USERNAME}")
    except Exception as e:
        # print(f"⚠️ Could not load SMTP settings from database, using defaults: {e}")
        # Use existing hardcoded values as fallback
        pass

def get_smtp_settings():
    """Get current SMTP settings (returns database values or fallback to globals)"""
    try:
        query = """
        SELECT setting_key, setting_value
        FROM system_settings
        WHERE setting_category = 'email' AND setting_key IN ('smtp_server', 'smtp_port', 'smtp_username', 'smtp_password', 'from_name', 'reply_to')
        """
        settings = execute_query(query, fetch_all=True)
        
        result = {}
        for setting in settings:
            result[setting['setting_key']] = setting['setting_value']
        
        return {
            'smtp_server': result.get('smtp_server', SMTP_SERVER),
            'smtp_port': int(result.get('smtp_port', SMTP_PORT)),
            'smtp_username': result.get('smtp_username', SMTP_USERNAME),
            'smtp_password': result.get('smtp_password', SMTP_PASSWORD),
            'from_name': result.get('from_name', 'YellowStone XPs'),
            'reply_to': result.get('reply_to', '')
        }
    except Exception as e:
        # Return global values as fallback
        return {
            'smtp_server': SMTP_SERVER,
            'smtp_port': SMTP_PORT,
            'smtp_username': SMTP_USERNAME,
            'smtp_password': SMTP_PASSWORD,
            'from_name': 'YellowStone XPs',
            'reply_to': ''
        }

# Database connection helper
def get_db_connection():
    """Get MySQL database connection"""
    try:
        # Attempting to connect to MySQL database
        # Database connection details
        
        import mysql.connector
        connection = mysql.connector.connect(**DB_CONFIG)
        # Database connection successful
        return connection
    except Error as e:
        print(f"âŒ Error connecting to MySQL: {e}")
        print(f"âŒ Connection details: {DB_CONFIG}")
        return None

def execute_query(query, params=None, fetch_one=False, fetch_all=False):
    """Execute MySQL query and return results as dictionaries"""
    connection = get_db_connection()
    if not connection:
        # No database connection available
        if fetch_all:
            return []
        return None
    
    cursor = None
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute(query, params)
        
        if fetch_one:
            result = cursor.fetchone()
            return result
        elif fetch_all:
            results = cursor.fetchall()
            return results if results else []
        else:
            # For INSERT, UPDATE, DELETE queries
            results = cursor.fetchall()
            connection.commit()
            return results if results else []
    except Error as e:
        print(f"❌ MySQL query error: {e}")
        print(f"Query: {query}")
        print(f"Params: {params}")
        if connection:
            connection.rollback()
        if fetch_all:
            return []
        return None
    except Exception as e:
        print(f"❌ Unexpected error in execute_query: {e}")
        print(f"Query: {query}")
        if connection:
            connection.rollback()
        if fetch_all:
            return []
        return None
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

def _generate_avatar(name):
    """Generate avatar initials from name"""
    if not name:
        return 'U'
    
    clean_name = name.strip()
    name_parts = clean_name.split(' ')
    
    first_char = name_parts[0][0].upper() if name_parts[0] else 'U'
    
    second_char = ''
    if len(name_parts) > 1 and name_parts[1]:
        second_char = name_parts[1][0].upper()
    
    return first_char + second_char

# Helper functions
def allowed_image_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS

def get_tasks_data():
    """Get all tasks from MySQL"""
    try:
        query = "SELECT * FROM tasks"
        tasks = execute_query(query, fetch_all=True)
        return tasks or []
    except Exception as e:
        print(f"Error getting tasks: {e}")
        return []

def save_task(task_data):
    """Save task to MySQL"""
    try:
        query = """
        INSERT INTO tasks (title, description, status, priority, assigned_to, assigned_by, due_date, created_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        params = (
            task_data.get('title'),
            task_data.get('description'),
            task_data.get('status', 'Pending'),
            task_data.get('priority', 'Medium'),
            task_data.get('assigned_to'),
            task_data.get('assigned_by'),
            task_data.get('due_date'),
            datetime.now()
        )
        task_id = execute_query(query, params)
        if task_id:
            task_data['id'] = task_id
            return task_data
        return None
    except Exception as e:
        print(f"Error saving task: {e}")
        return None

def get_projects_data():
    """Get all projects from MySQL"""
    try:
        query = "SELECT * FROM projects"
        projects = execute_query(query, fetch_all=True)
        return projects or []
    except Exception as e:
        print(f"Error getting projects: {e}")
        return []

def save_project(project_data):
    """Save project to MySQL"""
    try:
        query = """
        INSERT INTO projects (name, description, status, start_date, end_date, created_at)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        params = (
            project_data.get('name'),
            project_data.get('description'),
            project_data.get('status', 'Active'),
            project_data.get('start_date'),
            project_data.get('end_date'),
            datetime.now()
        )
        project_id = execute_query(query, params)
        if project_id:
            project_data['id'] = project_id
            return project_data
        return None
    except Exception as e:
        print(f"Error saving project: {e}")
        return None

def get_workflows_data():
    """Get all workflows from MySQL"""
    try:
        query = "SELECT * FROM workflows"
        workflows = execute_query(query, fetch_all=True)
        return workflows or []
    except Exception as e:
        print(f"Error getting workflows: {e}")
        return []

def save_workflow(workflow_data):
    """Save workflow to MySQL"""
    try:
        query = """
        INSERT INTO workflows (name, description, status, priority, assigned_to, created_at)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        params = (
            workflow_data.get('name'),
            workflow_data.get('description'),
            workflow_data.get('status', 'Active'),
            workflow_data.get('priority', 'Medium'),
            workflow_data.get('assigned_to'),
            datetime.now()
        )
        workflow_id = execute_query(query, params)
        if workflow_id:
            workflow_data['id'] = workflow_id
            return workflow_data
        return None
    except Exception as e:
        print(f"Error saving workflow: {e}")
        return None

def get_tickets_data():
    """Get all tickets from MySQL"""
    try:
        query = "SELECT * FROM tickets"
        tickets = execute_query(query, fetch_all=True)
        return tickets or []
    except Exception as e:
        print(f"Error getting tickets: {e}")
        return []

def save_ticket(ticket_data):
    """Save ticket to MySQL"""
    try:
        query = """
        INSERT INTO tickets (title, description, status, priority, assigned_to, created_by, created_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        params = (
            ticket_data.get('title'),
            ticket_data.get('description'),
            ticket_data.get('status', 'Open'),
            ticket_data.get('priority', 'Medium'),
            ticket_data.get('assigned_to'),
            ticket_data.get('created_by'),
            datetime.now()
        )
        ticket_id = execute_query(query, params)
        if ticket_id:
            ticket_data['id'] = ticket_id
            return ticket_data
        return None
    except Exception as e:
        print(f"Error saving ticket: {e}")
        return None

def get_vendors_data():
    """Get all vendors from MySQL"""
    try:
        table_check = execute_query("SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_name = 'vendors') as exists", fetch_one=True)
        if not table_check or not table_check['exists']:
            print("Vendors table does not exist!")
            return []
        
        query = "SELECT * FROM vendors"
        print(f"Executing vendors query: {query}")
        vendors = execute_query(query, fetch_all=True)
        print(f"Vendors query result: {vendors}")
        if vendors is None:
            print("Vendors query returned None")
            return []
        print(f"Vendors query returned {len(vendors) if vendors else 0} records")
        return vendors
    except Exception as e:
        print(f"Error getting vendors: {e}")
        return []

def get_nda_requests_data():
    """Get all NDA requests from MySQL"""
    try:
        table_check = execute_query("SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_name = 'nda_requests') as exists", fetch_one=True)
        if not table_check or not table_check['exists']:
            print("NDA requests table does not exist!")
            return []
        
        query = "SELECT * FROM nda_requests"
        print(f"Executing NDA requests query: {query}")
        nda_requests = execute_query(query, fetch_all=True)
        print(f"NDA requests query result: {nda_requests}")
        if nda_requests is None:
            print("NDA requests query returned None")
            return []
        print(f"NDA requests query returned {len(nda_requests) if nda_requests else 0} records")
        return nda_requests
    except Exception as e:
        print(f"Error getting NDA requests: {e}")
        return []

def get_notifications_data():
    """Get all notifications from MySQL"""
    try:
        query = "SELECT * FROM notifications"
        notifications = execute_query(query, fetch_all=True)
        return notifications or []
    except Exception as e:
        print(f"Error getting notifications: {e}")
        return []

def get_users_data():
    """Get all users from MySQL"""
    try:
        query = "SELECT * FROM users"
        users = execute_query(query, fetch_all=True)
        return users or []
    except Exception as e:
        print(f"Error getting users: {e}")
        return []

# Authentication routes
@app.route('/api/auth/employee-login-by-id', methods=['POST'])
def employee_login_by_id():
    try:
        data = request.get_json()
        employee_id = data.get('employee_id')
        password = data.get('password')
        
        if not employee_id or not password:
            return jsonify({'success': False, 'message': 'Employee ID and password are required'}), 400
        
        query = """
        SELECT u.*, COALESCE(m.name, '') as manager_name
        FROM users u
        LEFT JOIN users m ON u.manager = m.id
        WHERE u.employee_id = %s AND u.user_type = 'employee'
        """
        user = execute_query(query, (employee_id,), fetch_one=True)
        
        if not user:
            return jsonify({'success': False, 'message': 'Invalid employee ID'}), 401
        
        if not user['password_hash'] or user['password_hash'] == '':
            return jsonify({'success': False, 'message': 'No password set. Please register first.'}), 401
        
        if not bcrypt.checkpw(password.encode('utf-8'), user['password_hash'].encode('utf-8')):
            return jsonify({'success': False, 'message': 'Invalid password'}), 401
        
        session['user_id'] = user['id']
        session['user_type'] = 'employee'
        
        CURRENT_USER.update({
            'id': user['id'],
            'name': user['name'],
            'email': user['email'],
            'designation': user.get('designation'),
            'department': user.get('department'),
            'manager': user.get('manager'),
            'employee_code': user['employee_id'],
            'user_type': 'employee'
        })
        
        return jsonify({
            'success': True,
            'message': 'Login successful',
            'user': {
                'id': user['id'],
                'name': user['name'],
                'email': user['email'],
                'employee_id': user['employee_id'],
                'designation': user.get('designation'),
                'department': user.get('department'),
                'manager': user.get('manager_name'),
                'user_type': 'employee'
            }
        })
        
    except Exception as e:
        print(f"Employee login error: {e}")
        return jsonify({'success': False, 'message': 'Login failed'}), 500

@app.route('/api/auth/current-user', methods=['GET'])
def get_current_user():
    try:
        user_id = session.get('user_id')
        user_type = session.get('user_type')
        
        if not user_id or not user_type:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401
        
        if user_type == 'admin':
            query = "SELECT * FROM admins WHERE id = %s AND is_active = 1"
            user = execute_query(query, (user_id,), fetch_one=True)
            
            if not user:
                return jsonify({'success': False, 'message': 'Admin not found'}), 404
                
            return jsonify({
                'success': True,
                'user': {
                    'id': user['id'],
                    'name': user['full_name'],
                    'email': user['email'],
                    'username': user['username'],
                    'role': user['role'],
                    'user_type': 'admin'
                }
            })
            
        elif user_type == 'employee':
            query = "SELECT * FROM users WHERE id = %s AND user_type = 'employee'"
            user = execute_query(query, (user_id,), fetch_one=True)
            
            if not user:
                return jsonify({'success': False, 'message': 'Employee not found'}), 404
            
            return jsonify({
                'success': True,
                'user': {
                    'id': user['id'],
                    'name': user['name'],
                    'email': user['email'],
                    'employee_id': user.get('employee_id'),
                    'designation': user.get('designation'),
                    'department': user.get('department'),
                    'manager': user.get('manager'),
                    'user_type': 'employee'
                }
            })
            
        elif user_type == 'vendor':
            query = "SELECT vl.*, v.company_name, v.contact_person FROM vendor_logins vl JOIN vendors v ON vl.vendor_id = v.id WHERE vl.vendor_id = %s AND vl.is_active = 1"
            user = execute_query(query, (user_id,), fetch_one=True)
            
            if not user:
                return jsonify({'success': False, 'message': 'Vendor not found'}), 404
            
            return jsonify({
                'success': True,
                'user': {
                    'id': user['vendor_id'],
                    'name': user['company_name'],
                    'email': user['email'],
                    'company': user['company_name'],
                    'contact_person': user['contact_person'],
                    'user_type': 'vendor'
                }
            })
        
        return jsonify({'success': False, 'message': 'Invalid user type'}), 400
        
    except Exception as e:
        print(f"Get current user error: {e}")
        return jsonify({'success': False, 'message': 'Failed to get user'}), 500

@app.route('/api/auth/temporary-login', methods=['POST'])
def temporary_login():
    """Handle temporary login for vendor registration"""
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({'success': False, 'message': 'Email and password are required'}), 400
        
        email_normalized = email.strip().lower()
        print(f"[temporary-login] Attempt for email={email_normalized}")
        
        # Query to get temporary login credentials
        query = """
        SELECT tl.id, tl.vendor_id, tl.email, tl.password_hash, tl.company_name as vendor_company, 
               tl.contact_person as vendor_contact, tl.reference_number,
               tl.expires_at, tl.created_at
        FROM temporary_logins tl
        WHERE LOWER(tl.email) = %s AND tl.expires_at > NOW()
        ORDER BY tl.created_at DESC
        LIMIT 1
        """
        temp_login = execute_query(query, (email_normalized,), fetch_one=True)
        
        if not temp_login:
            print(f"[temporary-login] No active temp login found for {email_normalized}")
            return jsonify({'success': False, 'message': 'Invalid credentials or expired'}), 401
        
        if not bcrypt.checkpw(password.encode('utf-8'), temp_login['password_hash'].encode('utf-8')):
            print(f"[temporary-login] Password mismatch for {email_normalized}")
            return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
        
        print(f"[temporary-login] Successful login for {email_normalized}")
        
        return jsonify({
            'success': True,
            'message': 'Login successful',
            'user': {
                'id': temp_login['vendor_id'],
                'email': temp_login['email'],
                'company': temp_login['vendor_company'],
                'contact_person': temp_login['vendor_contact'],
                'reference_number': temp_login['reference_number'],
                'user_type': 'temp_vendor',
                'is_temporary': True
            }
        })
        
    except Exception as e:
        print(f"Temporary login error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Login failed'}), 500

@app.route('/api/auth/vendor-login', methods=['POST'])
def vendor_login():
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({'success': False, 'message': 'Email and password are required'}), 400
        
        query = """
        SELECT vl.id, vl.vendor_id, vl.email, vl.password_hash, vl.company_name, vl.contact_person, 
               vl.phone, vl.address, vl.is_active, vl.created_at, vl.last_login,
               v.company_name as vendor_company, v.contact_person as vendor_contact, 
               v.portal_access, v.nda_status, v.reference_number, vr.status as registration_status
        FROM vendor_logins vl 
        JOIN vendors v ON vl.vendor_id = v.id 
        LEFT JOIN vendor_registrations vr ON vr.email = vl.email
        WHERE vl.email = %s AND vl.is_active = 1
        """
        vendor = execute_query(query, (email,), fetch_one=True)
        
        if not vendor:
            return jsonify({'success': False, 'message': 'Invalid email or password'}), 401
        
        if not bcrypt.checkpw(password.encode('utf-8'), vendor['password_hash'].encode('utf-8')):
            return jsonify({'success': False, 'message': 'Invalid email or password'}), 401
        
        update_query = "UPDATE vendor_logins SET last_login = NOW() WHERE id = %s"
        execute_query(update_query, (vendor['id'],))
        
        session['user_id'] = vendor['vendor_id']
        session['user_type'] = 'vendor'
        
        CURRENT_USER.update({
            'id': vendor['vendor_id'],
            'name': vendor['vendor_company'],
            'email': vendor['email'],
            'company': vendor['vendor_company'],
            'contact_person': vendor['vendor_contact'],
            'user_type': 'vendor'
        })
        
        has_full_access = (
            vendor['portal_access'] == 1 and 
            vendor['registration_status'] == 'approved'
        )
        
        return jsonify({
            'success': True,
            'message': 'Login successful',
            'user': {
                'id': vendor['vendor_id'],
                'name': vendor['vendor_company'],
                'email': vendor['email'],
                'company': vendor['vendor_company'],
                'contact_person': vendor['vendor_contact'],
                'user_type': 'vendor',
                'has_full_access': has_full_access,
                'portal_access': vendor['portal_access'],
                'registration_status': vendor['registration_status'],
                'nda_status': vendor['nda_status'],
                'reference_number': vendor['reference_number']
            }
        })
        
    except Exception as e:
        print(f"Vendor login error: {e}")
        return jsonify({'success': False, 'message': 'Login failed'}), 500

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    try:
        session.clear()
        CURRENT_USER.update({
            "id": None,
            "name": None,
            "email": None,
            "designation": None,
            "department": None,
            "manager": None,
            "employee_code": None,
            "user_type": None
        })
        return jsonify({'success': True, 'message': 'Logged out successfully'})
    except Exception as e:
        print(f"Logout error: {e}")
        return jsonify({'success': False, 'message': 'Logout failed'}), 500

@app.route('/api/auth/register-admin', methods=['POST'])
def register_admin():
    try:
        data = request.get_json()
        email = data.get('email')
        username = data.get('username')
        full_name = data.get('full_name')
        password = data.get('password')
        
        if not email or not username or not full_name or not password:
            return jsonify({'success': False, 'message': 'All fields are required'}), 400
        
        query = "SELECT id FROM admins WHERE email = %s OR username = %s"
        existing_admin = execute_query(query, (email, username), fetch_one=True)
        
        if existing_admin:
            return jsonify({'success': False, 'message': 'Admin with this email or username already exists'}), 400
        
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        query = """
        INSERT INTO admins (email, username, password_hash, full_name, role, is_active)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        admin_id = execute_query(query, (email, username, password_hash, full_name, 'admin', True))
        
        if admin_id:
            return jsonify({
                'success': True,
                'message': 'Admin account created successfully',
                'admin': {
                    'id': admin_id,
                    'email': email,
                    'username': username,
                    'full_name': full_name,
                    'role': 'admin'
                }
            })
        else:
            return jsonify({'success': False, 'message': 'Failed to create admin account'}), 500
            
    except Exception as e:
        print(f"Admin registration error: {e}")
        return jsonify({'success': False, 'message': 'Registration failed'}), 500

@app.route('/api/auth/register-employee', methods=['POST'])
def register_employee():
    try:
        data = request.get_json()
        email = data.get('email')
        name = data.get('name')
        employee_id = data.get('employee_id')
        designation = data.get('designation')
        department = data.get('department')
        manager = data.get('manager')
        password = data.get('password')
        
        if not email or not name or not employee_id or not password:
            return jsonify({'success': False, 'message': 'Email, name, employee ID, and password are required'}), 400
        
        query = "SELECT id, password_hash FROM users WHERE email = %s OR employee_id = %s"
        existing_user = execute_query(query, (email, employee_id), fetch_one=True)
        
        if existing_user:
            if not existing_user['password_hash']:
                password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                
                update_query = """
                UPDATE users 
                SET password_hash = %s, updated_at = NOW()
                WHERE id = %s
                """
                execute_query(update_query, (password_hash, existing_user['id']))
                
                return jsonify({'success': True, 'message': 'Employee account created successfully'})
            else:
                return jsonify({'success': False, 'message': 'An account already exists for this employee. Please use the login page or contact your administrator to reset your password.'}), 400
        
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        query = """
        INSERT INTO users (email, password_hash, name, employee_id, designation, department, manager, user_type)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        user_id = execute_query(query, (email, password_hash, name, employee_id, designation, department, manager, 'employee'))
        
        if user_id:
            return jsonify({
                'success': True,
                'message': 'Employee account created successfully',
                'user_id': user_id
            })
        else:
            return jsonify({'success': False, 'message': 'Failed to create employee account'}), 500
            
    except Exception as e:
        print(f"Employee registration error: {e}")
        return jsonify({'success': False, 'message': 'Registration failed'}), 500

@app.route('/api/auth/register-vendor', methods=['POST'])
def register_vendor():
    try:
        data = request.get_json()  
        email = data.get('email')
        company_name = data.get('company_name')
        contact_person = data.get('contact_person')
        phone = data.get('phone')
        address = data.get('address')
        password = data.get('password')
        
        if not email or not company_name or not password:
            return jsonify({'success': False, 'message': 'Email, company name, and password are required'}), 400
        
        query = "SELECT id FROM vendors WHERE email = %s"
        existing_vendor = execute_query(query, (email,), fetch_one=True)
        
        if existing_vendor:
            return jsonify({'success': False, 'message': 'Vendor with this email already exists'}), 400
        
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        query = """
        INSERT INTO vendors (email, company_name, contact_person, phone, address, nda_status)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        vendor_id = execute_query(query, (email, company_name, contact_person, phone, address, 'pending'))
        
        if vendor_id:
            login_query = """
            INSERT INTO vendor_logins (vendor_id, email, password_hash, company_name, contact_person, phone, address, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """
            execute_query(login_query, (vendor_id, email, password_hash, company_name, contact_person, phone, address, True))
            
            return jsonify({
                'success': True,
                'message': 'Vendor account created successfully',
                'vendor': {
                    'id': vendor_id,
                    'email': email,
                    'company_name': company_name,
                    'contact_person': contact_person,
                    'phone': phone,
                    'address': address
                }
            })
        else:
            return jsonify({'success': False, 'message': 'Failed to create vendor account'}), 500
            
    except Exception as e:
        print(f"Vendor registration error: {e}")
        return jsonify({'success': False, 'message': 'Registration failed'}), 500

@app.route('/api/auth/admin-login', methods=['POST'])
def admin_login():
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({'success': False, 'message': 'Username and password are required'}), 400
        
        query = "SELECT * FROM admins WHERE (email = %s OR username = %s) AND is_active = 1"
        admin = execute_query(query, (username, username), fetch_one=True)
        
        if not admin:
            return jsonify({'success': False, 'message': 'Invalid username or password'}), 401
        
        if not bcrypt.checkpw(password.encode('utf-8'), admin['password_hash'].encode('utf-8')):
            return jsonify({'success': False, 'message': 'Invalid username or password'}), 401
        
        update_query = "UPDATE admins SET last_login = NOW() WHERE id = %s"
        execute_query(update_query, (admin['id'],))
        
        session['user_id'] = admin['id']
        session['user_type'] = 'admin'
        
        CURRENT_USER.update({
            'id': admin['id'],
            'name': admin['full_name'],
            'email': admin['email'],
            'username': admin['username'],
            'user_type': 'admin'
        })
        
        return jsonify({
            'success': True,
            'message': 'Login successful',
            'user': {
                'id': admin['id'],
                'name': admin['full_name'],
                'email': admin['email'],
                'username': admin['username'],
                'user_type': 'admin'
            }
        })
        
    except Exception as e:
        print(f"Admin login error: {e}")
        return jsonify({'success': False, 'message': 'Login failed'}), 500

# Health check endpoint
@app.route('/api/test/send-nda', methods=['POST'])
def test_send_nda():
    """Test endpoint for NDA sending functionality"""
    try:
        print("Testing NDA send functionality...")
        
        test_query = "SELECT COUNT(*) FROM vendors"
        result = execute_query(test_query, fetch_one=True)
        if result is None:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500
        
        print(f"âœ… Database connection working - {result['count']} vendors found")
        
        print(f"SMTP Configuration:")
        print(f"  Server: {SMTP_SERVER}")
        print(f"  Port: {SMTP_PORT}")
        print(f"  Username: {SMTP_USERNAME}")
        print(f"  Password: {'*' * len(SMTP_PASSWORD) if SMTP_PASSWORD else 'NOT SET'}")
        
        try:
            msg = MIMEMultipart()
            msg['From'] = SMTP_USERNAME
            msg['To'] = 'test@example.com'
            msg['Subject'] = 'Test NDA Email'
            msg.attach(MIMEText('Test email body', 'plain'))
            
            print("âœ… Email message created successfully")
            
        except Exception as email_test_error:
            print(f"âŒ Email message creation failed: {email_test_error}")
            return jsonify({'success': False, 'error': f'Email configuration error: {str(email_test_error)}'}), 500
        
        return jsonify({
            'success': True, 
            'message': 'NDA functionality test passed',
            'database_status': 'connected',
            'email_status': 'configured',
            'vendor_count': result['count']
        })
        
    except Exception as e:
        print(f"âŒ Test NDA error: {e}")
        return jsonify({'success': False, 'error': f'Test failed: {str(e)}'}), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint to test database connection"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({
                'status': 'error',
                'message': 'Database connection failed',
                'database': 'disconnected'
            }), 500
        
        result = execute_query("SELECT 1 as test", fetch_one=True)
        if not result:
            return jsonify({
                'status': 'error',
                'message': 'Database query failed',
                'database': 'connected_but_query_failed'
            }), 500
        
        vendors_count = execute_query("SELECT COUNT(*) FROM vendors", fetch_one=True)
        users_count = execute_query("SELECT COUNT(*) FROM users", fetch_one=True)
        
        return jsonify({
            'status': 'healthy',
            'message': 'All systems operational',
            'database': 'connected',
            'vendors_count': vendors_count['count'] if vendors_count else 0,
            'users_count': users_count['count'] if users_count else 0
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Health check failed: {str(e)}',
            'database': 'error'
        }), 500

# Dashboard routes
@app.route('/api/dashboard/stats', methods=['GET'])
def get_dashboard_stats():
    """Get dashboard statistics with proper error handling"""
    try:
        tasks_count = execute_query("SELECT COUNT(*) as count FROM tasks", fetch_one=True)
        projects_count = execute_query("SELECT COUNT(*) as count FROM projects", fetch_one=True)
        workflows_count = execute_query("SELECT COUNT(*) as count FROM workflows", fetch_one=True)
        tickets_count = execute_query("SELECT COUNT(*) as count FROM tickets", fetch_one=True)
        users_count = execute_query("SELECT COUNT(*) as count FROM users", fetch_one=True)
        
        return jsonify({
            'total_tasks': tasks_count['count'] if tasks_count else 0,
            'total_projects': projects_count['count'] if projects_count else 0,
            'total_workflows': workflows_count['count'] if workflows_count else 0,
            'total_tickets': tickets_count['count'] if tickets_count else 0,
            'active_users': users_count['count'] if users_count else 0
        })
        
    except Exception as e:
        return jsonify({
            'total_tasks': 0,
            'total_projects': 0,
            'total_workflows': 0,
            'total_tickets': 0,
            'active_users': 0
        })

@app.route('/api/auth/employee-details/<employee_id>', methods=['GET'])
def get_employee_details_by_id(employee_id):
    """Get employee details by employee ID for registration"""
    try:
        print(f"Fetching employee details for ID: {employee_id}")
        query = """
        SELECT u.id, u.name, u.email, u.employee_id, u.designation, u.department, 
               COALESCE(m.name, '') as manager_name, u.phone, u.address, u.hire_date, 
               u.salary, u.emergency_contact, u.status
        FROM users u
        LEFT JOIN users m ON u.manager = m.id
        WHERE u.user_type = 'employee' AND u.employee_id = %s
        """
        employee = execute_query(query, (employee_id,), fetch_one=True)
        print(f"Query result: {employee}")
        
        if not employee:
            print("Employee not found")
            return jsonify({'error': 'Employee not found'}), 404
        
        response_data = {
            'success': True,
            'employee': {
                'id': employee['id'],
                'name': employee['name'],
                'email': employee['email'],
                'employee_id': employee['employee_id'],
                'designation': employee['designation'],
                'department': employee['department'],
                'manager': employee['manager_name'],
                'phone': employee['phone'],
                'address': employee['address'],
                'hire_date': employee['hire_date'],
                'salary': employee['salary'],
                'emergency_contact': employee['emergency_contact'],
                'status': employee['status']
            }
        }
        print(f"Returning response: {response_data}")
        return jsonify(response_data)
        
    except Exception as e:
        print(f"Get employee details error: {e}")
        return jsonify({'error': 'Failed to get employee details'}), 500

# Attendance Management API endpoints
@app.route('/api/employee/attendance/clock-in', methods=['POST'])
def clock_in():
    """Employee clock in"""
    try:
        data = request.get_json()
        employee_id = data.get('employee_id')
        
        
        if not employee_id:
            return jsonify({'error': 'Employee ID is required'}), 400
        
        user_query = "SELECT id FROM users WHERE id = %s AND user_type = 'employee'"
        user = execute_query(user_query, (employee_id,), fetch_one=True)
        
        if not user:
            return jsonify({'error': 'Employee not found'}), 404
        
        from datetime import datetime, time, date
        
        current_time = datetime.now()
        current_date = current_time.date()
        current_time_only = current_time.time()
        
        check_query = "SELECT id, clock_in_time, clock_out_time FROM attendance WHERE employee_id = %s AND date = %s"
        existing = execute_query(check_query, (employee_id, current_date), fetch_one=True)
        
        # Allow clock-in if no existing record OR if already clocked out
        if existing and existing['clock_in_time'] and not existing['clock_out_time']:
            return jsonify({'error': 'Already clocked in today'}), 400
        
        late_threshold = time(10, 0)
        half_day_threshold = time(10, 30)
        
        if current_time_only <= late_threshold:
            status = 'Present'
            late_minutes = 0
        elif current_time_only <= half_day_threshold:
            status = 'Late'
            late_minutes = int((current_time_only.hour - 9) * 60 + current_time_only.minute - 30)
        else:
            status = 'Half Day'
            late_minutes = int((current_time_only.hour - 9) * 60 + current_time_only.minute - 30)
        
        if existing:
            # If clocked out, reset clock_out_time to null and update clock_in_time
            update_query = """
            UPDATE attendance 
            SET clock_in_time = %s, clock_out_time = NULL, status = %s, late_minutes = %s, total_hours = 0, updated_at = NOW()
            WHERE id = %s
            """
            execute_query(update_query, (current_time_only, status, late_minutes, existing['id']))
        else:
            insert_query = """
            INSERT INTO attendance (employee_id, date, clock_in_time, status, late_minutes, created_at)
            VALUES (%s, %s, %s, %s, %s, NOW())
            """
            execute_query(insert_query, (employee_id, current_date, current_time_only, status, late_minutes))
        
        # Broadcast attendance change to admin
        broadcast_database_change('attendance', 'update', {
            'employee_id': employee_id,
            'date': current_date.isoformat(),
            'action': 'clock_in',
            'clock_in_time': current_time_only.strftime('%H:%M'),
            'status': status,
            'late_minutes': late_minutes
        })
        
        return jsonify({
            'success': True, 
            'message': f'Clocked in successfully at {current_time_only.strftime("%H:%M")}',
            'status': status,
            'late_minutes': late_minutes,
            'clock_in_time': current_time_only.strftime('%H:%M')
        })
        
    except Exception as e:
        print(f"Clock in error: {e}")
        return jsonify({'error': 'Failed to clock in'}), 500

@app.route('/api/employee/attendance/clock-out', methods=['POST'])
def clock_out():
    """Employee clock out"""
    try:
        data = request.get_json()
        employee_id = data.get('employee_id')
        
        
        if not employee_id:
            return jsonify({'error': 'Employee ID is required'}), 400
        
        user_query = "SELECT id FROM users WHERE id = %s AND user_type = 'employee'"
        user = execute_query(user_query, (employee_id,), fetch_one=True)
        
        if not user:
            return jsonify({'error': 'Employee not found'}), 404
        
        from datetime import datetime, time, date
        
        current_time = datetime.now()
        current_date = current_time.date()
        current_time_only = current_time.time()
        
        check_query = "SELECT id, clock_in_time, clock_out_time, status FROM attendance WHERE employee_id = %s AND date = %s"
        attendance = execute_query(check_query, (employee_id, current_date), fetch_one=True)
        
        if not attendance or not attendance['clock_in_time']:
            return jsonify({'error': 'Please clock in first'}), 400
        
        if attendance['clock_out_time']:
            return jsonify({'error': 'Already clocked out today'}), 400
        
        clock_in_time = attendance['clock_in_time']
        
        # Handle different time formats
        if hasattr(clock_in_time, 'total_seconds'):  # timedelta object
            clock_in_hours = int(clock_in_time.total_seconds() // 3600)
            clock_in_minutes = int((clock_in_time.total_seconds() % 3600) // 60)
        else:  # time object
            clock_in_hours = clock_in_time.hour
            clock_in_minutes = clock_in_time.minute
        
        total_minutes = (current_time_only.hour - clock_in_hours) * 60 + (current_time_only.minute - clock_in_minutes)
        total_hours = round(total_minutes / 60, 2)
        
        # Ensure reasonable hours (not more than 24 hours in a day)
        total_hours = max(0, min(total_hours, 24))
        
        update_query = """
        UPDATE attendance 
        SET clock_out_time = %s, total_hours = %s, updated_at = NOW()
        WHERE id = %s
        """
        execute_query(update_query, (current_time_only, total_hours, attendance['id']))
        
        # Broadcast attendance change to admin
        broadcast_database_change('attendance', 'update', {
            'employee_id': employee_id,
            'date': current_date.isoformat(),
            'action': 'clock_out',
            'clock_out_time': current_time_only.strftime('%H:%M'),
            'total_hours': total_hours
        })
        
        return jsonify({
            'success': True, 
            'message': f'Clocked out successfully at {current_time_only.strftime("%H:%M")}',
            'total_hours': total_hours,
            'clock_out_time': current_time_only.strftime('%H:%M')
        })
        
    except Exception as e:
        print(f"Clock out error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to clock out: {str(e)}'}), 500

@app.route('/api/employee/attendance/status', methods=['GET'])
def get_attendance_status():
    """Get today's attendance status for employee"""
    try:
        employee_id = request.args.get('employee_id')
        
        if not employee_id:
            return jsonify({'error': 'Employee ID is required'}), 400
        
        user_query = "SELECT id FROM users WHERE id = %s AND user_type = 'employee'"
        user = execute_query(user_query, (employee_id,), fetch_one=True)
        
        if not user:
            return jsonify({'error': 'Employee not found'}), 404
        
        from datetime import date
        
        current_date = date.today()
        
        query = """
        SELECT clock_in_time, clock_out_time, status, total_hours, late_minutes
        FROM attendance 
        WHERE employee_id = %s AND date = %s
        """
        attendance = execute_query(query, (employee_id, current_date), fetch_one=True)
        
        if not attendance:
            return jsonify({
                'success': True,
                'attendance': {
                    'clocked_in': False,
                    'clocked_out': False,
                    'status': 'Absent',
                    'clock_in_time': None,
                    'clock_out_time': None,
                    'total_hours': 0,
                    'late_minutes': 0
                }
            })
        
        # Handle time formatting for different data types
        def format_time(time_obj):
            if not time_obj:
                return None
            if hasattr(time_obj, 'total_seconds'):  # timedelta
                hours = int(time_obj.total_seconds() // 3600)
                minutes = int((time_obj.total_seconds() % 3600) // 60)
                return f"{hours:02d}:{minutes:02d}"
            elif hasattr(time_obj, 'strftime'):  # time/datetime
                return time_obj.strftime('%H:%M')
            else:
                return str(time_obj)
        
        return jsonify({
            'success': True,
            'attendance': {
                'clocked_in': attendance['clock_in_time'] is not None,
                'clocked_out': attendance['clock_out_time'] is not None,
                'status': attendance['status'],
                'clock_in_time': format_time(attendance['clock_in_time']),
                'clock_out_time': format_time(attendance['clock_out_time']),
                'total_hours': float(attendance['total_hours']) if attendance['total_hours'] else 0,
                'late_minutes': attendance['late_minutes'] or 0
            }
        })
        
    except Exception as e:
        print(f"Get attendance status error: {e}")
        return jsonify({'error': 'Failed to get attendance status'}), 500

@app.route('/api/employee/attendance/history', methods=['GET'])
def get_attendance_history():
    """Get attendance history for employee"""
    try:
        employee_id = request.args.get('employee_id')
        month = request.args.get('month', '')
        year = request.args.get('year', '')
        
        if not employee_id:
            return jsonify({'error': 'Employee ID is required'}), 400
        
        user_query = "SELECT id FROM users WHERE employee_id = %s AND user_type = 'employee'"
        user = execute_query(user_query, (employee_id,), fetch_one=True)
        
        if not user:
            return jsonify({'error': 'Employee not found'}), 404
        
        from datetime import date, datetime
        
        if month and year:
            query = """
            SELECT date, clock_in_time, clock_out_time, status, total_hours, late_minutes
            FROM attendance 
            WHERE employee_id = %s AND YEAR(date) = %s AND MONTH(date) = %s
            ORDER BY date DESC
            """
            attendance_records = execute_query(query, (user['id'], year, month), fetch_all=True)
        else:
            current_date = date.today()
            query = """
            SELECT date, clock_in_time, clock_out_time, status, total_hours, late_minutes
            FROM attendance 
            WHERE employee_id = %s AND YEAR(date) = %s AND MONTH(date) = %s
            ORDER BY date DESC
            """
            attendance_records = execute_query(query, (user['id'], current_date.year, current_date.month), fetch_all=True)
        
        formatted_records = []
        for record in attendance_records:
            formatted_records.append({
                'date': record['date'].strftime('%Y-%m-%d'),
                'clock_in_time': record['clock_in_time'].strftime('%H:%M') if record['clock_in_time'] else None,
                'clock_out_time': record['clock_out_time'].strftime('%H:%M') if record['clock_out_time'] else None,
                'status': record['status'],
                'total_hours': float(record['total_hours']) if record['total_hours'] else 0,
                'late_minutes': record['late_minutes'] or 0
            })
        
        return jsonify({
            'success': True,
            'attendance_history': formatted_records
        })
        
    except Exception as e:
        print(f"Get attendance history error: {e}")
        return jsonify({'error': 'Failed to get attendance history'}), 500

@app.route('/api/admin/attendance', methods=['GET'])
def get_all_attendance():
    """Get all attendance records for admin"""
    try:
        from datetime import date
        
        # Get date from query parameter, default to today
        selected_date = request.args.get('date')
        if selected_date:
            current_date = date.fromisoformat(selected_date)
        else:
            current_date = date.today()
        
        query = """
        SELECT a.*, u.name as employee_name, u.employee_id
        FROM attendance a
        JOIN users u ON a.employee_id = u.id
        WHERE a.date = %s AND u.user_type = 'employee'
        ORDER BY u.name
        """
        attendance_records = execute_query(query, (current_date,), fetch_all=True)
        
        # Helper function to format time objects (handles both time and timedelta)
        def format_time(time_obj):
            if not time_obj:
                return None
            if hasattr(time_obj, 'total_seconds'):  # timedelta
                hours = int(time_obj.total_seconds() // 3600)
                minutes = int((time_obj.total_seconds() % 3600) // 60)
                return f"{hours:02d}:{minutes:02d}"
            elif hasattr(time_obj, 'strftime'):  # time/datetime
                return time_obj.strftime('%H:%M')
            else:
                return str(time_obj)
        
        formatted_records = []
        for record in attendance_records:
            formatted_records.append({
                'id': record['id'],
                'employee_id': record['employee_id'],
                'employee_name': record['employee_name'],
                'employee_code': record['employee_id'],
                'date': record['date'].strftime('%Y-%m-%d'),
                'clock_in_time': format_time(record['clock_in_time']),
                'clock_out_time': format_time(record['clock_out_time']),
                'status': record['status'],
                'total_hours': float(record['total_hours']) if record['total_hours'] else 0,
                'late_minutes': record['late_minutes'] or 0
            })
        
        return jsonify({
            'success': True,
            'attendance_records': formatted_records
        })
        
    except Exception as e:
        print(f"Get all attendance error: {e}")
        return jsonify({'error': 'Failed to get attendance records'}), 500

@app.route('/api/admin/attendance/download', methods=['GET'])
def download_attendance_report():
    """Download attendance report as CSV"""
    try:
        from datetime import date
        import csv
        from io import StringIO
        
        # Get date from query parameter, default to today
        selected_date = request.args.get('date')
        if selected_date:
            current_date = date.fromisoformat(selected_date)
        else:
            current_date = date.today()
        
        query = """
        SELECT a.*, u.name as employee_name, u.employee_id
        FROM attendance a
        JOIN users u ON a.employee_id = u.id
        WHERE a.date = %s AND u.user_type = 'employee'
        ORDER BY u.name
        """
        attendance_records = execute_query(query, (current_date,), fetch_all=True)
        
        # Helper function to format time objects (handles both time and timedelta)
        def format_time(time_obj):
            if not time_obj:
                return None
            if hasattr(time_obj, 'total_seconds'):  # timedelta
                hours = int(time_obj.total_seconds() // 3600)
                minutes = int((time_obj.total_seconds() % 3600) // 60)
                return f"{hours:02d}:{minutes:02d}"
            elif hasattr(time_obj, 'strftime'):  # time/datetime
                return time_obj.strftime('%H:%M')
            else:
                return str(time_obj)
        
        # Create CSV content
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Employee Name', 'Employee Code', 'Date', 'Clock In', 
            'Clock Out', 'Status', 'Total Hours', 'Late Minutes'
        ])
        
        # Write data rows
        for record in attendance_records:
            writer.writerow([
                record['employee_name'],
                record['employee_id'],
                record['date'].strftime('%Y-%m-%d'),
                format_time(record['clock_in_time']) or 'N/A',
                format_time(record['clock_out_time']) or 'N/A',
                record['status'],
                float(record['total_hours']) if record['total_hours'] else 0,
                record['late_minutes'] or 0
            ])
        
        # Prepare response
        csv_content = output.getvalue()
        output.close()
        
        from flask import make_response
        
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=attendance_report_{current_date}.csv'
        
        return response
        
    except Exception as e:
        print(f"Download attendance error: {e}")
        return jsonify({'error': 'Failed to download attendance report'}), 500

@app.route('/api/admin/attendance/monthly', methods=['GET'])
def download_monthly_attendance_report():
    """Download monthly attendance report as CSV"""
    try:
        from datetime import date
        import csv
        from io import StringIO
        
        # Get year and month from query parameters
        year = int(request.args.get('year', date.today().year))
        month = int(request.args.get('month', date.today().month))
        
        # Get all days in the month
        from calendar import monthrange
        days_in_month = monthrange(year, month)[1]
        
        # Query to get all attendance records for the month
        query = """
        SELECT a.*, u.name as employee_name, u.employee_id
        FROM attendance a
        JOIN users u ON a.employee_id = u.id
        WHERE YEAR(a.date) = %s AND MONTH(a.date) = %s AND u.user_type = 'employee'
        ORDER BY u.name, a.date
        """
        attendance_records = execute_query(query, (year, month), fetch_all=True)
        
        # Helper function to format time objects (handles both time and timedelta)
        def format_time(time_obj):
            if not time_obj:
                return None
            if hasattr(time_obj, 'total_seconds'):  # timedelta
                hours = int(time_obj.total_seconds() // 3600)
                minutes = int((time_obj.total_seconds() % 3600) // 60)
                return f"{hours:02d}:{minutes:02d}"
            elif hasattr(time_obj, 'strftime'):  # time/datetime
                return time_obj.strftime('%H:%M')
            else:
                return str(time_obj)
        
        # Create CSV content
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header with additional columns for monthly report
        writer.writerow([
            'Employee Name', 'Employee Code', 'Date', 'Clock In', 
            'Clock Out', 'Status', 'Total Hours', 'Late Minutes', 'Week Day'
        ])
        
        # Write data rows
        for record in attendance_records:
            date_obj = record['date']
            week_day = date_obj.strftime('%A')  # Full weekday name
            writer.writerow([
                record['employee_name'],
                record['employee_id'],
                record['date'].strftime('%Y-%m-%d'),
                format_time(record['clock_in_time']) or 'N/A',
                format_time(record['clock_out_time']) or 'N/A',
                record['status'],
                float(record['total_hours']) if record['total_hours'] else 0,
                record['late_minutes'] or 0,
                week_day
            ])
        
        # Prepare response
        csv_content = output.getvalue()
        output.close()
        
        from flask import make_response
        
        month_name = date(year, month, 1).strftime('%B')
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=monthly_attendance_report_{month_name}_{year}.csv'
        
        return response
        
    except Exception as e:
        print(f"Download monthly attendance error: {e}")
        return jsonify({'error': 'Failed to download monthly attendance report'}), 500

@app.route('/api/admin/attendance/monthly-matrix', methods=['GET'])
def download_monthly_attendance_matrix():
    """Download month-wide matrix for all employees.
    - CSV by default
    - XLSX with colors if query param format=xlsx
    """
    try:
        from datetime import date
        import csv
        from io import StringIO, BytesIO
        from calendar import monthrange

        year = int(request.args.get('year', date.today().year))
        month = int(request.args.get('month', date.today().month))
        out_format = (request.args.get('format') or 'csv').lower()

        days_in_month = monthrange(year, month)[1]

        # Get all employees
        employees = execute_query(
            "SELECT id, name, employee_id FROM users WHERE user_type = 'employee' ORDER BY name",
            fetch_all=True
        ) or []

        # Get all attendance in month
        rows = execute_query(
            """
            SELECT a.employee_id, a.date, a.status
            FROM attendance a
            JOIN users u ON a.employee_id = u.id
            WHERE YEAR(a.date) = %s AND MONTH(a.date) = %s AND u.user_type = 'employee'
            """,
            (year, month),
            fetch_all=True
        ) or []

        # Build map: (employee_id, day) -> status
        status_map = {}
        for r in rows:
            d = int(r['date'].day)
            status_map[(r['employee_id'], d)] = r['status'] or 'Absent'

        def status_full(s):
            if not s:
                return 'Absent'
            s_low = str(s).lower()
            if s_low.startswith('present'):
                return 'Present'
            if s_low.startswith('late'):
                return 'Late'
            if s_low.startswith('half'):
                return 'Half Day'
            if s_low.startswith('absent'):
                return 'Absent'
            return str(s).title()

        month_name = date(year, month, 1).strftime('%B')

        if out_format == 'xlsx':
            # Build XLSX with colors
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font

            wb = Workbook()
            ws = wb.active
            ws.title = f"{month_name}_{year}"

            header = ['Employee Name', 'Employee Code'] + [str(d) for d in range(1, days_in_month + 1)] + ['Present', 'Late', 'Half Day', 'Absent']
            ws.append(header)

            # Define color fills
            fill_present = None  # default text black
            fill_absent = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red
            fill_half = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")    # light green
            fill_late = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")    # light yellow
            font_black = Font(color="000000")
            font_red = Font(color="9C0006")

            for emp in employees:
                counts = {'Present': 0, 'Late': 0, 'Half Day': 0, 'Absent': 0}
                row_vals = [emp['name'], emp['employee_id']]
                # day statuses
                for d in range(1, days_in_month + 1):
                    st = status_full(status_map.get((emp['id'], d), 'Absent'))
                    counts[st] = counts.get(st, 0) + 1
                    row_vals.append(st)
                row_vals += [counts['Present'], counts['Late'], counts['Half Day'], counts['Absent']]
                ws.append(row_vals)

                # Apply colors for the last written row
                current_row = ws.max_row
                # Days start at column 3
                for col_idx in range(3, 3 + days_in_month):
                    cell = ws.cell(row=current_row, column=col_idx)
                    if cell.value == 'Present':
                        cell.font = font_black
                        if fill_present:
                            cell.fill = fill_present
                    elif cell.value == 'Absent':
                        cell.font = font_red
                        cell.fill = fill_absent
                    elif cell.value == 'Half Day':
                        cell.fill = fill_half
                    elif cell.value == 'Late':
                        cell.fill = fill_late

            x_out = BytesIO()
            wb.save(x_out)
            x_out.seek(0)

            from flask import send_file
            return send_file(
                x_out,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f"attendance_matrix_{month_name}_{year}.xlsx"
            )
        else:
            # Prepare CSV with full words
            output = StringIO()
            writer = csv.writer(output)
            header = ['Employee Name', 'Employee Code'] + [str(d) for d in range(1, days_in_month + 1)] + ['Present', 'Late', 'Half Day', 'Absent']
            writer.writerow(header)

            for emp in employees:
                counts = {'Present': 0, 'Late': 0, 'Half Day': 0, 'Absent': 0}
                day_values = []
                for d in range(1, days_in_month + 1):
                    st = status_full(status_map.get((emp['id'], d), 'Absent'))
                    counts[st] = counts.get(st, 0) + 1
                    day_values.append(st)
                writer.writerow([
                    emp['name'],
                    emp['employee_id'],
                    *day_values,
                    counts.get('Present', 0),
                    counts.get('Late', 0),
                    counts.get('Half Day', 0),
                    counts.get('Absent', 0)
                ])

            csv_content = output.getvalue()
            output.close()

            from flask import make_response
            response = make_response(csv_content)
            response.headers['Content-Type'] = 'text/csv'
            response.headers['Content-Disposition'] = f'attachment; filename=attendance_matrix_{month_name}_{year}.csv'
            return response
    except Exception as e:
        print(f"Download monthly attendance matrix error: {e}")
        return jsonify({'error': 'Failed to download monthly attendance matrix'}), 500


@app.route('/api/admin/attendance/summary', methods=['GET'])
def get_attendance_summary():
    """Return attendance summary stats for daily/weekly/monthly ranges"""
    try:
        from datetime import date, datetime, timedelta

        range_type = (request.args.get('range') or 'day').lower()  # day | week | month
        # Optional reference date; default today
        ref_date_param = request.args.get('date')
        if ref_date_param:
            ref_date = date.fromisoformat(ref_date_param)
        else:
            ref_date = date.today()

        # Compute range bounds
        if range_type == 'week':
            # ISO week starts Monday (weekday 0)
            start_date = ref_date - timedelta(days=ref_date.weekday())
            end_date = start_date + timedelta(days=6)
            date_filter_sql = "a.date BETWEEN %s AND %s"
            date_filter_args = (start_date, end_date)
        elif range_type == 'month':
            start_date = ref_date.replace(day=1)
            if start_date.month == 12:
                next_month = start_date.replace(year=start_date.year + 1, month=1, day=1)
            else:
                next_month = start_date.replace(month=start_date.month + 1, day=1)
            end_date = next_month - timedelta(days=1)
            date_filter_sql = "YEAR(a.date) = %s AND MONTH(a.date) = %s"
            date_filter_args = (start_date.year, start_date.month)
        else:
            # default daily
            start_date = ref_date
            end_date = ref_date
            date_filter_sql = "a.date = %s"
            date_filter_args = (ref_date,)

        # Total employees
        total_employees_row = execute_query(
            "SELECT COUNT(*) as cnt FROM users WHERE user_type = 'employee'",
            fetch_one=True
        )
        total_employees = total_employees_row['cnt'] if total_employees_row else 0

        # Aggregate by status and average hours/late minutes
        agg_query = f"""
            SELECT 
                COUNT(*) as total_records,
                SUM(CASE WHEN a.status = 'Present' THEN 1 ELSE 0 END) as present_count,
                SUM(CASE WHEN a.status = 'Late' THEN 1 ELSE 0 END) as late_count,
                SUM(CASE WHEN a.status = 'Half Day' THEN 1 ELSE 0 END) as halfday_count,
                SUM(CASE WHEN a.status = 'Absent' THEN 1 ELSE 0 END) as absent_count,
                AVG(a.total_hours) as avg_hours,
                SUM(CASE WHEN a.clock_in_time IS NOT NULL THEN 1 ELSE 0 END) as clockins
            FROM attendance a
            JOIN users u ON a.employee_id = u.id
            WHERE {date_filter_sql} AND u.user_type = 'employee'
        """
        agg = execute_query(agg_query, date_filter_args, fetch_one=True)

        # Ensure defaults
        def safe(val, default=0):
            return float(val) if (val is not None and isinstance(val, (int, float))) else (val or default)

        summary = {
            'range': range_type,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'total_employees': total_employees,
            'total_records': safe(agg.get('total_records')) if agg else 0,
            'present': safe(agg.get('present_count')) if agg else 0,
            'late': safe(agg.get('late_count')) if agg else 0,
            'half_day': safe(agg.get('halfday_count')) if agg else 0,
            'absent': safe(agg.get('absent_count')) if agg else 0,
            'avg_hours': round(safe(agg.get('avg_hours')), 2) if agg else 0,
            'clock_ins': safe(agg.get('clockins')) if agg else 0
        }

        return jsonify({'success': True, 'summary': summary})
    except Exception as e:
        print(f"Attendance summary error: {e}")
        return jsonify({'error': 'Failed to get attendance summary'}), 500


@app.route('/api/admin/attendance/weekly', methods=['GET'])
def download_weekly_attendance_report():
    """Download weekly attendance report as CSV (Monday-Sunday)"""
    try:
        from datetime import date, timedelta
        import csv
        from io import StringIO

        # Start date for the week; default to current week's Monday
        start_param = request.args.get('start')
        if start_param:
            start_date = date.fromisoformat(start_param)
        else:
            today = date.today()
            start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)

        query = """
        SELECT a.*, u.name as employee_name, u.employee_id
        FROM attendance a
        JOIN users u ON a.employee_id = u.id
        WHERE a.date BETWEEN %s AND %s AND u.user_type = 'employee'
        ORDER BY u.name, a.date
        """
        rows = execute_query(query, (start_date, end_date), fetch_all=True)

        def format_time(time_obj):
            if not time_obj:
                return None
            if hasattr(time_obj, 'total_seconds'):
                hours = int(time_obj.total_seconds() // 3600)
                minutes = int((time_obj.total_seconds() % 3600) // 60)
                return f"{hours:02d}:{minutes:02d}"
            elif hasattr(time_obj, 'strftime'):
                return time_obj.strftime('%H:%M')
            return str(time_obj)

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Employee Name', 'Employee Code', 'Date', 'Clock In', 'Clock Out', 'Status', 'Total Hours', 'Late Minutes'])
        for r in rows:
            writer.writerow([
                r['employee_name'],
                r['employee_id'],
                r['date'].strftime('%Y-%m-%d'),
                format_time(r['clock_in_time']) or 'N/A',
                format_time(r['clock_out_time']) or 'N/A',
                r['status'],
                float(r['total_hours']) if r['total_hours'] else 0,
                r['late_minutes'] or 0
            ])

        csv_content = output.getvalue()
        output.close()
        from flask import make_response
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=weekly_attendance_{start_date}_to_{end_date}.csv'
        return response
    except Exception as e:
        print(f"Download weekly attendance error: {e}")
        return jsonify({'error': 'Failed to download weekly attendance report'}), 500

# Vendor Registration API endpoints
@app.route('/api/vendor/register', methods=['POST'])
def register_vendor_detailed():
    """Register vendor for full portal access with comprehensive YellowStone form data"""
    try:
        data = request.get_json()
        
        print(f"Received registration data: {data}")
        
        # Extract comprehensive form data with fallbacks
        company_name = data.get('companyName', '')
        contact_person = data.get('contactPersonName', '')
        email = data.get('emailAddress', '')
        phone = data.get('phoneNumber', '') or data.get('phone', '')
        address = data.get('communicationAddress', '') or data.get('address', '')
        business_type = data.get('companyType', '') or data.get('businessType', '')
        services = data.get('servicesOffered', '') or data.get('services', '')
        experience = data.get('previousWorkExperience', '') or data.get('experience', '')
        certifications = data.get('certifications', '')
        vendor_references = data.get('majorCustomers', '') or data.get('references', '')
        
        print(f"Field mapping check:")
        print(f"  companyName: '{company_name}' (type: {type(company_name)})")
        print(f"  contactPersonName: '{contact_person}' (type: {type(contact_person)})")
        print(f"  emailAddress: '{email}' (type: {type(email)})")
        print(f"  phoneNumber: '{phone}' (type: {type(phone)})")
        print(f"  communicationAddress: '{address}' (type: {type(address)})")
        print(f"  companyType: '{business_type}' (type: {type(business_type)})")
        print(f"  servicesOffered: '{services}' (type: {type(services)})")
        print(f"  previousWorkExperience: '{experience}' (type: {type(experience)})")
        
        # Additional comprehensive fields
        designation = data.get('designation', '')
        registered_address = data.get('registeredOfficeAddress', '')
        fax_number = data.get('faxNumber', '')
        website = data.get('website', '')
        nature_of_business = data.get('natureOfBusiness', '')
        year_of_establishment = data.get('yearOfEstablishment', '')
        pan_number = data.get('panNumber', '')
        bank_name = data.get('bankName', '')
        account_number = data.get('accountNumber', '')
        ifsc_code = data.get('ifscCode', '')
        branch_name = data.get('branchName', '')
        annual_turnover = data.get('annualTurnover', '')
        net_worth = data.get('netWorth', '')
        total_employees = data.get('totalEmployees', '')
        technical_staff = data.get('technicalStaff', '')
        technical_capabilities = data.get('technicalCapabilities', '')
        additional_info = data.get('additionalInformation', '')
        organization_structure = data.get('organizationStructure', '')
        supplier_bank_name = data.get('supplierBankName', '')
        supplier_account_number = data.get('supplierAccountNumber', '')
        supplier_ifsc_code = data.get('supplierIfscCode', '')
        supplier_branch_name = data.get('supplierBranchName', '')
        
        print(f"Extracted fields - Company: {company_name}, Contact: {contact_person}, Email: {email}")
        
        # Check required fields (handle empty strings)
        required_fields = [company_name, contact_person, email, phone, address]
        required_field_names = ['company_name', 'contact_person', 'email', 'phone', 'address']
        
        missing_fields = []
        for field_name, field_value in zip(required_field_names, required_fields):
            if not field_value or (isinstance(field_value, str) and field_value.strip() == '') or field_value is None:
                missing_fields.append(field_name)
        
        if missing_fields:
            print(f"Missing required fields: {missing_fields}")
            return jsonify({'error': f'Missing required fields: {", ".join(missing_fields)}'}), 400
        
        # Test database connection first
        test_connection = get_db_connection()
        if not test_connection:
            print("âŒ Database connection failed")
            return jsonify({'error': 'Database connection failed. Please try again later.'}), 500
        
        # Check if vendor already exists, create if not
        check_query = "SELECT id FROM vendors WHERE email = %s"
        existing_vendor = execute_query(check_query, (email,), fetch_one=True)
        
        if existing_vendor:
            vendor_id = existing_vendor['id']
            print(f"Found existing vendor with ID: {vendor_id}")
        else:
            # Create new vendor record
            print(f"Creating new vendor record for {email}")
            vendor_insert_query = """
            INSERT INTO vendors (company_name, contact_person, email, phone, address, 
                               business_type, registration_status, portal_access, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, 'pending', 0, NOW())
            """
            vendor_id = execute_query(vendor_insert_query, (
                company_name, contact_person, email, phone, address, business_type
            ))
            print(f"Created new vendor with ID: {vendor_id}")
        
        # Check if registration already exists by email
        check_reg_query = "SELECT id FROM vendor_registrations WHERE email = %s"
        existing_reg = execute_query(check_reg_query, (email,), fetch_one=True)
        
        if existing_reg:
            print(f"Registration already exists for email {email}")
            return jsonify({'error': 'Registration already submitted for this email'}), 400
        
        print(f"No existing registration found for email {email}, proceeding with new registration")
        
        # Store comprehensive form data using existing schema
        insert_query = """
        INSERT INTO vendor_registrations 
        (company_name, contact_person, email, phone, address, business_type, 
         services, experience, certifications, vendor_references, status, created_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending', NOW())
        """
        
        print(f"Executing insert query for vendor: {company_name}")
        execute_query(insert_query, (
            company_name, contact_person, email, phone, address, business_type,
            services, experience, certifications, vendor_references
        ))
        
        print(f"Successfully inserted vendor registration for {company_name}")
        
        # Send confirmation email to vendor
        try:
            smtp = get_smtp_settings()
            msg = MIMEMultipart()
            msg['From'] = f'YellowStone XPs <{smtp["smtp_username"]}>'
            msg['To'] = email
            msg['Subject'] = 'Vendor Registration Submitted - YellowStone XPs'
            
            # Add anti-spam headers
            add_anti_spam_headers(msg, smtp)
            
            body = f"""Dear {contact_person},

We have successfully received your vendor registration for {company_name}.

REGISTRATION STATUS: COMPLETED ✓

Your registration has been submitted and is now under review by our team.

NEXT STEPS:
- Our team will review your registration details
- You will receive full portal access once approved
- You will be notified via email when your portal access is activated
- Expected review time: 3-5 business days

IMPORTANT INFORMATION:
- Company Name: {company_name}
- Contact Person: {contact_person}
- Email: {email}
- Phone: {phone}
- Registration Date: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}

If you have any questions or need to update your registration details, please contact our Vendor Relations Department.

Thank you for your interest in partnering with YellowStone Xperiences Pvt Ltd.

Best Regards,
YellowStone Xperiences Pvt Ltd
Harpreet Singh - CEO
Email: Harpreet.singh@yellowstonexps.com
Address: Plot # 2, ITC, Fourth Floor, Sector 67, Mohali -160062, Punjab, India"""
            
            msg.attach(MIMEText(body, 'plain'))
            
            if smtp['smtp_port'] == 465:
                server = smtplib.SMTP_SSL(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
            else:
                server = smtplib.SMTP(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
                server.starttls()
            
            server.login(smtp['smtp_username'], smtp['smtp_password'])
            server.sendmail(smtp['smtp_username'], email, msg.as_string())
            server.quit()
            print(f"✅ Confirmation email sent to {email}")
        except Exception as email_error:
            print(f"⚠️ Failed to send confirmation email: {email_error}")
            # Don't fail the registration if email fails
        
        # Broadcast the new registration to connected clients
        broadcast_database_change('vendor_registrations', 'insert', {
            'company_name': company_name,
            'contact_person': contact_person,
            'email': email,
            'status': 'pending'
        })
        
        return jsonify({
            'success': True,
            'message': 'Comprehensive vendor registration submitted successfully. You will receive an email once approved.'
        })
        
    except Exception as e:
        print(f"Vendor registration error: {e}")
        print(f"Error type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to submit vendor registration: {str(e)}'}), 500

# Test endpoint to check database connection
@app.route('/api/test-db', methods=['GET'])
def test_database():
    """Test database connection"""
    try:
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("SELECT 1 as test")
            result = cursor.fetchone()
            cursor.close()
            connection.close()
            return jsonify({'success': True, 'message': 'Database connection successful', 'test': result})
        else:
            return jsonify({'success': False, 'message': 'No database connection'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'}), 500

# Admin API endpoints for vendor registration management
@app.route('/api/admin/vendor-registrations', methods=['GET'])
def get_vendor_registrations():
    """Get all vendor registrations for admin review"""
    try:
        query = """
        SELECT id, company_name, contact_person, email, phone, address, 
               business_type, services, experience, certifications, 
               vendor_references, status, created_at, updated_at
        FROM vendor_registrations 
        ORDER BY created_at DESC
        """
        
        registrations = execute_query(query, fetch_all=True)
        
        return jsonify({
            'success': True,
            'registrations': registrations
        })
        
    except Exception as e:
        print(f"Get vendor registrations error: {e}")
        return jsonify({'error': 'Failed to get vendor registrations'}), 500

@app.route('/api/admin/vendor-registrations/<int:registration_id>/approve', methods=['POST'])
def approve_vendor_registration(registration_id):
    """Approve a vendor registration"""
    try:
        query = "SELECT * FROM vendor_registrations WHERE id = %s"
        registration = execute_query(query, (registration_id,), fetch_one=True)
        
        print(f"=== APPROVE DEBUG ===")
        print(f"Registration ID: {registration_id}")
        print(f"Registration found: {registration is not None}")
        if registration:
            print(f"Current status: '{registration['status']}' (type: {type(registration['status'])})")
            print(f"Status comparison: '{registration['status']}' != 'pending' = {registration['status'] != 'pending'}")
        
        if not registration:
            return jsonify({'success': False, 'message': 'Registration not found'}), 404
        
        if registration['status'] not in ['pending', None]:
            print(f"Registration status is not pending: '{registration['status']}'")
            return jsonify({'success': False, 'message': f'Registration is not pending (current status: {registration["status"]})'}), 400
        
        # Update registration status
        update_query = "UPDATE vendor_registrations SET status = 'approved', updated_at = NOW() WHERE id = %s"
        execute_query(update_query, (registration_id,))
        
        # Find vendor by email and update status
        vendor_query = "SELECT id, email FROM vendors WHERE email = %s"
        vendor = execute_query(vendor_query, (registration['email'],), fetch_one=True)
        
        if vendor:
            print(f"✅ Found vendor ID: {vendor['id']} for email: {registration['email']}")
            # Update vendor status AND company information to match registration
            vendor_update_query = """
            UPDATE vendors 
            SET registration_status = 'approved', 
                portal_access = 1,
                company_name = %s,
                contact_person = %s,
                phone = %s,
                address = %s,
                updated_at = NOW()
            WHERE id = %s
            """
            execute_query(vendor_update_query, (
                registration['company_name'],
                registration['contact_person'],
                registration['phone'],
                registration['address'],
                vendor['id']
            ))
            print(f"✅ Updated vendor information for vendor_id: {vendor['id']}")
            
            # Use vendor's actual email from the vendors table, not from registration
            vendor_email = vendor['email']
            print(f"✅ Using vendor email: {vendor_email} for login creation")
            
            # Generate password for vendor login
            vendor_password = generate_vendor_password()
            password_hash = bcrypt.hashpw(vendor_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            # DELETE any existing vendor_login records for this email to avoid confusion
            delete_query = "DELETE FROM vendor_logins WHERE email = %s"
            execute_query(delete_query, (vendor_email,))
            print(f"✅ Deleted old vendor_login records for email: {vendor_email}")
            
            # Create vendor login
            login_query = """
            INSERT INTO vendor_logins (vendor_id, email, password_hash, company_name, contact_person, phone, address)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            execute_query(login_query, (
                vendor['id'],
                vendor_email,  # Use vendor's email from vendors table
                password_hash,
                registration['company_name'],
                registration['contact_person'],
                registration['phone'],
                registration['address']
            ))
            print(f"✅ Created vendor_login for vendor_id: {vendor['id']} with email: {vendor_email}")
        else:
            # Create new vendor if not found
            vendor_insert_query = """
                INSERT INTO vendors (company_name, contact_person, email, phone, address, 
                                   business_type, registration_status, portal_access, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, 'approved', 1, NOW())
                """
            vendor_id = execute_query(vendor_insert_query, (
                registration['company_name'], registration['contact_person'], registration['email'],
                registration['phone'], registration['address'], registration['business_type']
            ))
            
            # Generate password for vendor login
            vendor_password = generate_vendor_password()
            password_hash = bcrypt.hashpw(vendor_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            # DELETE any existing vendor_login records for this email
            delete_query = "DELETE FROM vendor_logins WHERE email = %s"
            execute_query(delete_query, (registration['email'],))
            print(f"✅ Deleted old vendor_login records for email: {registration['email']}")
            
            # Create vendor login
            login_query = """
            INSERT INTO vendor_logins (vendor_id, email, password_hash, company_name, contact_person, phone, address)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            execute_query(login_query, (
                vendor_id, 
                registration['email'], 
                password_hash,
                registration['company_name'],
                registration['contact_person'],
                registration['phone'],
                registration['address']
            ))
            print(f"✅ Created new vendor with ID: {vendor_id}")
        
        # Send credentials email to vendor using correct email
        vendor_email_to_use = vendor['email'] if vendor else registration['email']
        email_sent = send_vendor_credentials_email(
            vendor_email_to_use, 
            vendor_password, 
            registration['company_name']
        )
        
        # Broadcast the database change to connected clients
        broadcast_database_change('vendor_registrations', 'update', {
            'id': registration_id,
            'status': 'approved',
            'company_name': registration['company_name']
        })
        
        email_status = "Credentials sent via email" if email_sent else "Warning: Email failed to send"
        
        return jsonify({
            'success': True,
            'message': f'Vendor registration approved successfully. {email_status}',
            'vendor_password': vendor_password,
            'email_sent': email_sent
        })
        
    except Exception as e:
        print(f"Approve vendor registration error: {e}")
        return jsonify({'error': 'Failed to approve vendor registration'}), 500

@app.route('/api/admin/vendor-registrations/<int:registration_id>/decline', methods=['POST'])
def decline_vendor_registration(registration_id):
    """Decline a vendor registration"""
    try:
        query = "SELECT * FROM vendor_registrations WHERE id = %s"
        registration = execute_query(query, (registration_id,), fetch_one=True)
        
        print(f"=== DECLINE DEBUG ===")
        print(f"Registration ID: {registration_id}")
        print(f"Registration found: {registration is not None}")
        if registration:
            print(f"Current status: '{registration['status']}' (type: {type(registration['status'])})")
            print(f"Status comparison: '{registration['status']}' != 'pending' = {registration['status'] != 'pending'}")
        
        if not registration:
            return jsonify({'success': False, 'message': 'Registration not found'}), 404
        
        if registration['status'] not in ['pending', None]:
            print(f"Registration status is not pending: '{registration['status']}'")
            return jsonify({'success': False, 'message': f'Registration is not pending (current status: {registration["status"]})'}), 400
        
        update_query = "UPDATE vendor_registrations SET status = 'rejected', updated_at = NOW() WHERE id = %s"
        execute_query(update_query, (registration_id,))
        
        # Broadcast the database change to connected clients
        broadcast_database_change('vendor_registrations', 'update', {
            'id': registration_id,
            'status': 'rejected',
            'company_name': registration['company_name']
        })
        
        return jsonify({
            'success': True,
            'message': 'Vendor registration rejected'
        })
        
    except Exception as e:
        print(f"Decline vendor registration error: {e}")
        return jsonify({'error': 'Failed to decline vendor registration'}), 500

@app.route('/api/vendor/dashboard-stats', methods=['GET'])
def get_vendor_dashboard_stats():
    """Get vendor dashboard statistics"""
    try:
        vendor_id = request.headers.get('X-Vendor-ID')
        if not vendor_id:
            return jsonify({'success': False, 'message': 'Vendor ID required'}), 400
        
        # Get vendor basic info
        vendor_query = "SELECT company_name, nda_status, registration_status FROM vendors WHERE id = %s"
        vendor_data = execute_query(vendor_query, (vendor_id,), fetch_one=True)
        
        if not vendor_data:
            return jsonify({'success': False, 'message': 'Vendor not found'}), 404
        
        # Return real data - no sample figures
        stats = {
            'active_projects': 0,
            'completed_tasks': 0,
            'performance_score': 0,
            'messages': 0,
            'pending_approvals': 0,
            'upcoming_deadlines': 0,
            'total_earnings': 0,
            'monthly_earnings': 0
        }
        
        return jsonify({
            'success': True,
            'stats': stats,
            'vendor_info': {
                'company_name': vendor_data['company_name'],
                'nda_status': vendor_data['nda_status'],
                'registration_status': vendor_data['registration_status']
            }
        })
        
    except Exception as e:
        print(f"Get vendor dashboard stats error: {e}")
        return jsonify({'success': False, 'message': 'Failed to get dashboard stats'}), 500

@app.route('/api/vendor/projects', methods=['GET'])
def get_vendor_projects():
    """Get vendor projects"""
    try:
        vendor_id = request.headers.get('X-Vendor-ID')
        if not vendor_id:
            return jsonify({'success': False, 'message': 'Vendor ID required'}), 400
        
        # Return empty projects list - no sample data
        projects = []
        
        return jsonify({
            'success': True,
            'projects': projects
        })
        
    except Exception as e:
        print(f"Get vendor projects error: {e}")
        return jsonify({'success': False, 'message': 'Failed to get projects'}), 500

@app.route('/api/vendor/tasks', methods=['GET'])
def get_vendor_tasks():
    """Get vendor tasks"""
    try:
        vendor_id = request.headers.get('X-Vendor-ID')
        if not vendor_id:
            return jsonify({'success': False, 'message': 'Vendor ID required'}), 400
        
        # Return empty tasks list - no sample data
        tasks = []
        
        return jsonify({
            'success': True,
            'tasks': tasks
        })
        
    except Exception as e:
        print(f"Get vendor tasks error: {e}")
        return jsonify({'success': False, 'message': 'Failed to get tasks'}), 500

@app.route('/api/vendor/notifications', methods=['GET'])
def get_vendor_notifications():
    """Get vendor notifications"""
    try:
        vendor_id = request.headers.get('X-Vendor-ID')
        if not vendor_id:
            return jsonify({'success': False, 'message': 'Vendor ID required'}), 400
        
        # Return empty notifications list - no sample data
        notifications = []
        
        return jsonify({
            'success': True,
            'notifications': notifications
        })
        
    except Exception as e:
        print(f"Get vendor notifications error: {e}")
        return jsonify({'success': False, 'message': 'Failed to get notifications'}), 500

@app.route('/api/vendor/profile', methods=['GET'])
def get_vendor_profile():
    """Get vendor profile data for display"""
    try:
        # Get vendor ID from session or request headers
        vendor_id = request.headers.get('X-Vendor-ID')
        print(f"Received vendor ID: {vendor_id}")
        
        if not vendor_id:
            return jsonify({'success': False, 'message': 'Vendor ID required'}), 400
        
        # Get basic vendor information
        vendor_query = """
        SELECT id, company_name, contact_person, email, phone, address, 
               business_type, registration_number, country, state, 
               registration_status, nda_status, reference_number,
               portal_access, has_full_access, created_at, updated_at,
               signature_data, company_stamp_data, signature_type, signed_date,
               company_registration_number, company_incorporation_country, company_incorporation_state
        FROM vendors WHERE id = %s
        """
        vendor_data = execute_query(vendor_query, (vendor_id,), fetch_one=True)
        print(f"Vendor data found: {vendor_data is not None}")
        
        if not vendor_data:
            return jsonify({'success': False, 'message': 'Vendor not found'}), 404
        
        # Get comprehensive registration data if available (using correct columns)
        registration_query = """
        SELECT company_name, contact_person, email, phone, address, business_type,
               services, experience, certifications, vendor_references, status,
               created_at, updated_at
        FROM vendor_registrations WHERE email = %s
        ORDER BY created_at DESC LIMIT 1
        """
        registration_data = execute_query(registration_query, (vendor_data['email'],), fetch_one=True)
        
        # Get NDA form data if available
        nda_query = """
        SELECT form_data, status, signed_at, created_at
        FROM nda_forms WHERE vendor_id = %s
        ORDER BY created_at DESC LIMIT 1
        """
        nda_data = execute_query(nda_query, (vendor_id,), fetch_one=True)
        
        # Parse NDA form data if available
        nda_form_data = {}
        if nda_data and nda_data['form_data']:
            try:
                nda_form_data = json.loads(nda_data['form_data'])
            except json.JSONDecodeError:
                nda_form_data = {}
        
        # Create comprehensive data from vendor table and registration data
        comprehensive_data = {
            'companyName': vendor_data['company_name'],
            'contactPersonName': vendor_data['contact_person'],
            'emailAddress': vendor_data['email'],
            'phoneNumber': vendor_data['phone'],
            'communicationAddress': vendor_data['address'],
            'businessType': vendor_data['business_type'] or 'Technology Services',
            'registrationNumber': vendor_data['registration_number'] or vendor_data['company_registration_number'],
            'country': vendor_data['country'] or vendor_data['company_incorporation_country'],
            'state': vendor_data['state'] or vendor_data['company_incorporation_state']
        }
        
        # Add registration data if available
        if registration_data:
            comprehensive_data.update({
                'servicesOffered': registration_data['services'],
                'experience': registration_data['experience'],
                'certifications': registration_data['certifications'],
                'vendorReferences': registration_data['vendor_references']
            })
        
        # Combine vendor data with comprehensive registration data
        profile_data = {
            'id': vendor_data['id'],
            'company_name': vendor_data['company_name'],
            'contact_person': vendor_data['contact_person'],
            'email': vendor_data['email'],
            'phone': vendor_data['phone'],
            'address': vendor_data['address'],
            'business_type': vendor_data['business_type'],
            'registration_number': vendor_data['registration_number'],
            'country': vendor_data['country'],
            'state': vendor_data['state'],
            'registration_status': vendor_data['registration_status'],
            'nda_status': vendor_data['nda_status'],
            'reference_number': vendor_data['reference_number'],
            'portal_access': vendor_data['portal_access'],
            'has_full_access': vendor_data['has_full_access'],
            'created_at': vendor_data['created_at'].isoformat() if vendor_data['created_at'] else None,
            'updated_at': vendor_data['updated_at'].isoformat() if vendor_data['updated_at'] else None,
            'comprehensive_data': comprehensive_data,
            'registration_status_detail': registration_data['status'] if registration_data else None,
            'submitted_at': registration_data['created_at'].isoformat() if registration_data and registration_data['created_at'] else None,
            'reviewed_at': registration_data['updated_at'].isoformat() if registration_data and registration_data['updated_at'] else None,
            'nda_form_data': nda_form_data,
            'nda_signed_at': nda_data['signed_at'].isoformat() if nda_data and nda_data['signed_at'] else None,
            'signature_data': vendor_data['signature_data'],
            'company_stamp_data': vendor_data['company_stamp_data'],
            'signature_type': vendor_data['signature_type'],
            'signed_date': vendor_data['signed_date'].isoformat() if vendor_data['signed_date'] else None
        }
        
        return jsonify({
            'success': True,
            'profile': profile_data
        })
        
    except Exception as e:
        print(f"Get vendor profile error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Failed to get vendor profile'}), 500
        
        # Get comprehensive registration data if available
        registration_query = """
        SELECT form_data, status, submitted_at, reviewed_at
        FROM vendor_registrations WHERE vendor_id = %s
        ORDER BY created_at DESC LIMIT 1
        """
        registration_data = execute_query(registration_query, (vendor_id,), fetch_one=True)
        
        # Get NDA form data if available
        nda_query = """
        SELECT form_data, status, signed_at, created_at
        FROM nda_forms WHERE vendor_id = %s
        ORDER BY created_at DESC LIMIT 1
        """
        nda_data = execute_query(nda_query, (vendor_id,), fetch_one=True)
        
        # Parse JSON form data if available
        comprehensive_data = {}
        if registration_data and registration_data['form_data']:
            try:
                comprehensive_data = json.loads(registration_data['form_data'])
            except json.JSONDecodeError:
                comprehensive_data = {}
        
        # Parse NDA form data if available
        nda_form_data = {}
        if nda_data and nda_data['form_data']:
            try:
                nda_form_data = json.loads(nda_data['form_data'])
            except json.JSONDecodeError:
                nda_form_data = {}
        
        # If no comprehensive registration data, use vendor table data
        if not comprehensive_data:
            comprehensive_data = {
                'companyName': vendor_data['company_name'],
                'contactPersonName': vendor_data['contact_person'],
                'emailAddress': vendor_data['email'],
                'phoneNumber': vendor_data['phone'],
                'communicationAddress': vendor_data['address'],
                'businessType': vendor_data['business_type'] or 'Technology Services',
                'registrationNumber': vendor_data['registration_number'] or vendor_data['company_registration_number'],
                'country': vendor_data['country'] or vendor_data['company_incorporation_country'],
                'state': vendor_data['state'] or vendor_data['company_incorporation_state']
            }
        
        # Combine vendor data with comprehensive registration data
        profile_data = {
            'id': vendor_data['id'],
            'company_name': vendor_data['company_name'],
            'contact_person': vendor_data['contact_person'],
            'email': vendor_data['email'],
            'phone': vendor_data['phone'],
            'address': vendor_data['address'],
            'business_type': vendor_data['business_type'],
            'registration_number': vendor_data['registration_number'],
            'country': vendor_data['country'],
            'state': vendor_data['state'],
            'registration_status': vendor_data['registration_status'],
            'nda_status': vendor_data['nda_status'],
            'reference_number': vendor_data['reference_number'],
            'portal_access': vendor_data['portal_access'],
            'has_full_access': vendor_data['has_full_access'],
            'created_at': vendor_data['created_at'].isoformat() if vendor_data['created_at'] else None,
            'updated_at': vendor_data['updated_at'].isoformat() if vendor_data['updated_at'] else None,
            'comprehensive_data': comprehensive_data,
            'registration_status_detail': registration_data['status'] if registration_data else None,
            'submitted_at': registration_data['submitted_at'].isoformat() if registration_data and registration_data['submitted_at'] else None,
            'reviewed_at': registration_data['reviewed_at'].isoformat() if registration_data and registration_data['reviewed_at'] else None,
            'nda_form_data': nda_form_data,
            'nda_signed_at': nda_data['signed_at'].isoformat() if nda_data and nda_data['signed_at'] else None,
            'signature_data': vendor_data['signature_data'],
            'company_stamp_data': vendor_data['company_stamp_data'],
            'signature_type': vendor_data['signature_type'],
            'signed_date': vendor_data['signed_date'].isoformat() if vendor_data['signed_date'] else None
        }
        
        return jsonify({
            'success': True,
            'profile': profile_data
        })
        
    except Exception as e:
        print(f"Get vendor profile error: {e}")
        return jsonify({'success': False, 'message': 'Failed to get vendor profile'}), 500

@app.route('/api/vendor/nda-form/download', methods=['GET'])
def download_vendor_nda_form():
    """Download vendor's completed NDA form as PDF"""
    try:
        vendor_id = request.headers.get('X-Vendor-ID')
        if not vendor_id:
            return jsonify({'success': False, 'message': 'Vendor ID required'}), 400
        
        # Get vendor and NDA form data
        vendor_query = """
        SELECT company_name, contact_person, email, phone, address, 
               business_type, reference_number, signature_data, 
               company_stamp_data, signature_type, signed_date
        FROM vendors WHERE id = %s
        """
        vendor_data = execute_query(vendor_query, (vendor_id,), fetch_one=True)
        
        if not vendor_data:
            return jsonify({'success': False, 'message': 'Vendor not found'}), 404
        
        # Get NDA form data
        nda_query = """
        SELECT form_data, signed_at FROM nda_forms WHERE vendor_id = %s
        ORDER BY created_at DESC LIMIT 1
        """
        nda_data = execute_query(nda_query, (vendor_id,), fetch_one=True)
        
        if not nda_data:
            return jsonify({'success': False, 'message': 'NDA form not found'}), 404
        
        # Parse NDA form data
        nda_form_data = {}
        if nda_data['form_data']:
            try:
                nda_form_data = json.loads(nda_data['form_data'])
            except json.JSONDecodeError:
                nda_form_data = {}
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph("NON-DISCLOSURE AGREEMENT", title_style))
        story.append(Spacer(1, 20))
        
        # Company Information
        story.append(Paragraph("Company Information:", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        company_info = [
            ['Company Name:', vendor_data['company_name'] or 'N/A'],
            ['Contact Person:', vendor_data['contact_person'] or 'N/A'],
            ['Email:', vendor_data['email'] or 'N/A'],
            ['Phone:', vendor_data['phone'] or 'N/A'],
            ['Address:', vendor_data['address'] or 'N/A'],
            ['Business Type:', vendor_data['business_type'] or 'N/A'],
            ['Reference Number:', vendor_data['reference_number'] or 'N/A'],
            ['Signed Date:', vendor_data['signed_date'].strftime('%Y-%m-%d') if vendor_data['signed_date'] else 'N/A']
        ]
        
        company_table = Table(company_info, colWidths=[2*inch, 4*inch])
        company_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(company_table)
        story.append(Spacer(1, 20))
        
        # NDA Terms (simplified version)
        story.append(Paragraph("Agreement Terms:", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        terms = [
            "1. Confidential Information: The parties acknowledge that they may receive confidential and proprietary information.",
            "2. Non-Disclosure: Each party agrees not to disclose confidential information to third parties.",
            "3. Use Limitation: Confidential information shall only be used for the purpose of evaluating business opportunities.",
            "4. Return of Information: Upon request, all confidential information shall be returned or destroyed.",
            "5. Term: This agreement shall remain in effect for a period of 5 years from the date of signing.",
            "6. Governing Law: This agreement shall be governed by the laws of the jurisdiction where YellowStone Group operates."
        ]
        
        for term in terms:
            story.append(Paragraph(term, styles['Normal']))
            story.append(Spacer(1, 6))
        
        story.append(Spacer(1, 20))
        
        # Signature Section
        story.append(Paragraph("Signatures:", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        signature_info = [
            ['Signature Type:', vendor_data['signature_type'] or 'Digital'],
            ['Signature Present:', 'Yes' if vendor_data['signature_data'] else 'No'],
            ['Company Stamp Present:', 'Yes' if vendor_data['company_stamp_data'] else 'No']
        ]
        
        signature_table = Table(signature_info, colWidths=[2*inch, 4*inch])
        signature_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(signature_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        # Create response
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=NDA_Form_{vendor_data["company_name"]}_{vendor_data["reference_number"]}.pdf'
        
        return response
        
    except Exception as e:
        print(f"Download NDA form error: {e}")
        return jsonify({'success': False, 'message': 'Failed to download NDA form'}), 500

# Organization Management API endpoints
# Removed duplicate route - using get_admin_employees instead

@app.route('/api/admin/employees', methods=['POST'])
def create_employee():
    """Create a new employee"""
    try:
        data = request.get_json()
        
        name = data.get('name')
        email = data.get('email')
        employee_id = data.get('employeeId')
        position = data.get('position')
        department = data.get('department')
        manager = data.get('manager')
        phone = data.get('phone', '')
        address = data.get('address', '')
        hire_date = data.get('hireDate')
        salary = data.get('salary', '')
        emergency_contact = data.get('emergencyContact', '')
        password = data.get('password', 'temp123')
        
        if not all([name, email, employee_id, position, department]):
            return jsonify({'success': False, 'message': 'Name, email, employee ID, position, and department are required'}), 400
        
        query = "SELECT id FROM users WHERE email = %s OR employee_id = %s"
        existing = execute_query(query, (email, employee_id), fetch_one=True)
        if existing:
            return jsonify({'success': False, 'message': 'Employee with this email or employee ID already exists'}), 400
        
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        query = """
        INSERT INTO users (email, password_hash, name, employee_id, designation, department, 
                          manager, phone, address, hire_date, salary, emergency_contact, 
                          status, user_type, created_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'active', 'employee', NOW())
        """
        
        execute_query(query, (email, password_hash, name, employee_id, position, department,
                             manager, phone, address, hire_date, salary, emergency_contact))
        
        return jsonify({'success': True, 'message': 'Employee created successfully'})
        
    except Exception as e:
        print(f"Create employee error: {e}")
        return jsonify({'success': False, 'message': 'Failed to create employee'}), 500

@app.route('/api/admin/employees/<int:employee_id>', methods=['GET'])
def get_admin_employee_by_id(employee_id):
    """Get specific employee by ID for admin dashboard"""
    try:
        admin_id = session.get('user_id')
        if not admin_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        # Check if user is admin
        admin_check = execute_query("SELECT role FROM admins WHERE id = %s", (admin_id,), fetch_one=True)
        if not admin_check or admin_check['role'] != 'admin':
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Get specific employee with department info
        query = """
        SELECT 
            u.id, u.employee_id, u.name, u.email, u.designation,
            u.phone, u.is_active, u.created_at,
            d.name as department_name,
            m.name as manager_name,
            m.id as manager_id
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        LEFT JOIN users m ON u.manager_id = m.id
        WHERE u.employee_id = %s AND u.is_active = 1
        """
        
        employee = execute_query(query, (employee_id,), fetch_one=True)
        
        if employee:
            return jsonify({'success': True, 'employee': employee})
        else:
            return jsonify({'success': False, 'error': 'Employee not found'}), 404
        
    except Exception as e:
        print(f"âŒ Error getting employee by ID: {e}")
        return jsonify({'error': 'Failed to get employee'}), 500


@app.route('/api/admin/employees/<int:employee_id>', methods=['PUT'])
def update_employee(employee_id):
    """Update an employee"""
    try:
        data = request.get_json()
        
        name = data.get('name')
        email = data.get('email')
        employee_id_new = data.get('employeeId')
        position = data.get('position')
        department = data.get('department')
        manager = data.get('manager')
        phone = data.get('phone', '')
        address = data.get('address', '')
        hire_date = data.get('hireDate')
        salary = data.get('salary', '')
        emergency_contact = data.get('emergencyContact', '')
        status = data.get('status', 'active')
        
        if not all([name, email, employee_id_new, position, department]):
            return jsonify({'success': False, 'message': 'Name, email, employee ID, position, and department are required'}), 400
        
        query = "SELECT id FROM users WHERE (email = %s OR employee_id = %s) AND id != %s"
        existing = execute_query(query, (email, employee_id_new, employee_id), fetch_one=True)
        if existing:
            return jsonify({'success': False, 'message': 'Email or employee ID already exists for another employee'}), 400
        
        query = """
        UPDATE users 
        SET name = %s, email = %s, employee_id = %s, designation = %s, department = %s,
            manager = %s, phone = %s, address = %s, hire_date = %s, salary = %s,
            emergency_contact = %s, status = %s, updated_at = NOW()
        WHERE id = %s AND user_type = 'employee'
        """
        
        execute_query(query, (name, email, employee_id_new, position, department, manager,
                             phone, address, hire_date, salary, emergency_contact, status, employee_id))
        
        return jsonify({'success': True, 'message': 'Employee updated successfully'})
        
    except Exception as e:
        print(f"Update employee error: {e}")
        return jsonify({'success': False, 'message': 'Failed to update employee'}), 500

@app.route('/api/admin/employees/<int:employee_id>', methods=['DELETE'])
def delete_employee(employee_id):
    """Delete an employee"""
    try:
        query = "SELECT id FROM users WHERE id = %s AND user_type = 'employee'"
        employee = execute_query(query, (employee_id,), fetch_one=True)
        if not employee:
            return jsonify({'success': False, 'message': 'Employee not found'}), 404
        
        query = "DELETE FROM users WHERE id = %s AND user_type = 'employee'"
        execute_query(query, (employee_id,))
        
        return jsonify({'success': True, 'message': 'Employee deleted successfully'})
        
    except Exception as e:
        print(f"Delete employee error: {e}")
        return jsonify({'success': False, 'message': 'Failed to delete employee'}), 500

@app.route('/api/admin/departments', methods=['GET'])
def get_all_departments():
    """Get all departments"""
    try:
        query = """
        SELECT d.id, d.name, d.description, d.head_id, d.budget, d.location, d.established_date,
               COALESCE(emp_count.employee_count, 0) as employee_count
        FROM departments d
        LEFT JOIN (
            SELECT department, COUNT(*) as employee_count
            FROM users 
            WHERE user_type = 'employee' AND department IS NOT NULL AND department != ''
            GROUP BY department
        ) emp_count ON d.name = emp_count.department
        ORDER BY d.name
        """
        departments = execute_query(query, fetch_all=True)
        
        dept_list = []
        for dept in departments:
            dept_dict = {
                'id': dept['id'],
                'name': dept['name'],
                'employeeCount': dept['employee_count'],
                'description': dept['description'] or f"{dept['name']} Department",
                'head': dept['head_id'],
                'budget': dept['budget'] or '',
                'location': dept['location'] or '',
                'establishedDate': dept['established_date'] or ''
            }
            dept_list.append(dept_dict)
        
        return jsonify(dept_list)
        
    except Exception as e:
        print(f"Get departments error: {e}")
        return jsonify({'error': 'Failed to get departments'}), 500

@app.route('/api/admin/organization-stats', methods=['GET'])
def get_organization_stats():
    """Get organization statistics"""
    try:
        query = "SELECT COUNT(*) as total FROM users WHERE user_type = 'employee'"
        total_employees = execute_query(query, fetch_one=True)['total']
        
        query = "SELECT COUNT(*) as active FROM users WHERE user_type = 'employee' AND status = 'active'"
        active_employees = execute_query(query, fetch_one=True)['active']
        
        query = "SELECT COUNT(DISTINCT department) as dept_count FROM users WHERE user_type = 'employee' AND department IS NOT NULL"
        departments = execute_query(query, fetch_one=True)['dept_count']
        
        query = "SELECT COUNT(*) as managers FROM users WHERE user_type = 'employee' AND designation LIKE '%manager%'"
        managers = execute_query(query, fetch_one=True)['managers']
        
        return jsonify({
            'totalEmployees': total_employees,
            'activeEmployees': active_employees,
            'departments': departments,
            'managers': managers
        })
        
    except Exception as e:
        print(f"Get organization stats error: {e}")
        return jsonify({'error': 'Failed to get organization stats'}), 500

# Task Management API Endpoints
@app.route('/api/admin/tasks', methods=['GET'])
def get_tasks():
    """Get all tasks"""
    try:
        query = """
        SELECT t.*, 
               CASE 
                   WHEN t.assigned_to_type = 'employee' THEN u.name
                   WHEN t.assigned_to_type = 'vendor' THEN v.company_name
                   ELSE 'Unassigned'
               END as assigned_to_name,
               CASE 
                   WHEN t.assigned_to_type = 'employee' THEN u.email
                   WHEN t.assigned_to_type = 'vendor' THEN v.email
                   ELSE NULL
               END as assigned_to_email
        FROM tasks t
        LEFT JOIN users u ON t.assigned_to_type = 'employee' AND t.assigned_to_id = u.id
        LEFT JOIN vendors v ON t.assigned_to_type = 'vendor' AND t.assigned_to_id = v.id
        ORDER BY t.created_at DESC
        """
        tasks = execute_query(query, fetch_all=True)
        return jsonify(tasks)
    except Exception as e:
        print(f"Get tasks error: {e}")
        return jsonify({'error': 'Failed to get tasks'}), 500

@app.route('/api/admin/tasks', methods=['POST'])
def create_task():
    """Create a new task"""
    try:
        data = request.get_json()
        
        title = data.get('title')
        description = data.get('description', '')
        priority = data.get('priority', 'medium')
        status = data.get('status', 'pending')
        assigned_to_type = data.get('assigned_to_type')
        assigned_to_id = data.get('assigned_to_id')
        salary_code = data.get('salary_code', '')
        due_date = data.get('due_date')
        created_by = data.get('created_by', 1)
        
        if not title:
            return jsonify({'error': 'Task title is required'}), 400
        
        priority_mapping = {
            'low': 'Low',
            'medium': 'Medium', 
            'high': 'High',
            'urgent': 'Critical'
        }
        db_priority = priority_mapping.get(priority.lower(), 'Medium')
        
        query = """
        INSERT INTO tasks (title, description, priority, status, assigned_to_type, assigned_to_id, assigned_to, due_date, salary_code, created_by, created_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
        """
        
        assigned_to = assigned_to_id if assigned_to_type == 'employee' else None
        
        execute_query(query, (title, description, db_priority, status, assigned_to_type, assigned_to_id, assigned_to, due_date, salary_code, created_by))
        
        return jsonify({'success': True, 'message': 'Task created successfully'})
        
    except Exception as e:
        print(f"Create task error: {e}")
        return jsonify({'error': 'Failed to create task'}), 500

@app.route('/api/admin/tasks/<int:task_id>', methods=['PUT'])
def update_task(task_id):
    """Update a task"""
    try:
        data = request.get_json()
        
        title = data.get('title')
        description = data.get('description', '')
        priority = data.get('priority', 'medium')
        status = data.get('status', 'pending')
        assigned_to_type = data.get('assigned_to_type')
        assigned_to_id = data.get('assigned_to_id')
        salary_code = data.get('salary_code', '')
        due_date = data.get('due_date')
        
        if not title:
            return jsonify({'error': 'Task title is required'}), 400
        
        priority_mapping = {
            'low': 'Low',
            'medium': 'Medium', 
            'high': 'High',
            'urgent': 'Critical'
        }
        db_priority = priority_mapping.get(priority.lower(), 'Medium')
        
        query = """
        UPDATE tasks 
        SET title = %s, description = %s, priority = %s, status = %s, 
            assigned_to_type = %s, assigned_to_id = %s, due_date = %s, salary_code = %s, updated_at = NOW()
        WHERE id = %s
        """
        
        execute_query(query, (title, description, db_priority, status, assigned_to_type, assigned_to_id, due_date, salary_code, task_id))
        
        return jsonify({'success': True, 'message': 'Task updated successfully'})
        
    except Exception as e:
        print(f"Update task error: {e}")
        return jsonify({'error': 'Failed to update task'}), 500

@app.route('/api/admin/tasks/<int:task_id>', methods=['DELETE'])
def delete_task(task_id):
    """Delete a task"""
    try:
        query = "DELETE FROM tasks WHERE id = %s"
        execute_query(query, (task_id,))
        
        return jsonify({'success': True, 'message': 'Task deleted successfully'})
        
    except Exception as e:
        print(f"Delete task error: {e}")
        return jsonify({'error': 'Failed to delete task'}), 500

# Project Management API Endpoints
@app.route('/api/admin/projects', methods=['GET'])
def get_projects():
    """Get all projects"""
    try:
        query = """
        SELECT p.*, 
               CASE 
                   WHEN p.assigned_to_type = 'employee' THEN u.name
                   WHEN p.assigned_to_type = 'vendor' THEN v.company_name
                   ELSE 'Unassigned'
               END as assigned_to_name,
               CASE 
                   WHEN p.assigned_to_type = 'employee' THEN u.email
                   WHEN p.assigned_to_type = 'vendor' THEN v.email
                   ELSE NULL
               END as assigned_to_email
        FROM projects p
        LEFT JOIN users u ON p.assigned_to_type = 'employee' AND p.assigned_to_id = u.id
        LEFT JOIN vendors v ON p.assigned_to_type = 'vendor' AND p.assigned_to_id = v.id
        ORDER BY p.created_at DESC
        """
        projects = execute_query(query, fetch_all=True)
        return jsonify(projects)
    except Exception as e:
        print(f"Get projects error: {e}")
        return jsonify({'error': 'Failed to get projects'}), 500

@app.route('/api/admin/projects', methods=['POST'])
def create_project():
    """Create a new project"""
    try:
        data = request.get_json()
        
        name = data.get('name')
        description = data.get('description', '')
        status = data.get('status', 'planning')
        priority = data.get('priority', 'medium')
        assigned_to_type = data.get('assigned_to_type')
        assigned_to_id = data.get('assigned_to_id')
        salary_code = data.get('salary_code', '')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        budget = data.get('budget', '')
        created_by = data.get('created_by', 1)
        
        if not name:
            return jsonify({'error': 'Project name is required'}), 400
        
        priority_mapping = {
            'low': 'low',
            'medium': 'medium', 
            'high': 'high',
            'urgent': 'urgent'
        }
        db_priority = priority_mapping.get(priority.lower(), 'medium')
        
        query = """
        INSERT INTO projects (name, description, status, priority, assigned_to_type, assigned_to_id, project_manager, start_date, end_date, budget, salary_code, created_by, created_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
        """
        
        project_manager = assigned_to_id if assigned_to_type == 'employee' else None
        
        execute_query(query, (name, description, status, db_priority, assigned_to_type, assigned_to_id, project_manager, start_date, end_date, budget, salary_code, created_by))
        
        return jsonify({'success': True, 'message': 'Project created successfully'})
        
    except Exception as e:
        print(f"Create project error: {e}")
        return jsonify({'error': 'Failed to create project'}), 500

@app.route('/api/admin/projects/<int:project_id>', methods=['PUT'])
def update_project(project_id):
    """Update a project"""
    try:
        data = request.get_json()
        
        name = data.get('name')
        description = data.get('description', '')
        status = data.get('status', 'planning')
        priority = data.get('priority', 'medium')
        assigned_to_type = data.get('assigned_to_type')
        assigned_to_id = data.get('assigned_to_id')
        salary_code = data.get('salary_code', '')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        budget = data.get('budget', '')
        
        if not name:
            return jsonify({'error': 'Project name is required'}), 400
        
        priority_mapping = {
            'low': 'low',
            'medium': 'medium', 
            'high': 'high',
            'urgent': 'urgent'
        }
        db_priority = priority_mapping.get(priority.lower(), 'medium')
        
        # Normalize dates to YYYY-MM-DD to avoid MySQL date errors
        def normalize_date(val):
            if not val:
                return None
            try:
                # ISO or date-like strings
                if isinstance(val, str):
                    if 'T' in val:
                        return val.split('T')[0]
                    if ',' in val and 'GMT' in val:
                        from datetime import datetime
                        try:
                            dt = datetime.strptime(val, '%a, %d %b %Y %H:%M:%S %Z')
                            return dt.strftime('%Y-%m-%d')
                        except Exception:
                            pass
                    # Fallback: first 10 chars if looks like date
                    if len(val) >= 10:
                        return val[:10]
                return val
            except Exception:
                return None

        start_date = normalize_date(start_date)
        end_date = normalize_date(end_date)

        query = """
        UPDATE projects 
        SET name = %s, description = %s, status = %s, priority = %s, 
            assigned_to_type = %s, assigned_to_id = %s, start_date = %s, end_date = %s, budget = %s, salary_code = %s, updated_at = NOW()
        WHERE id = %s
        """
        
        execute_query(query, (name, description, status, db_priority, assigned_to_type, assigned_to_id, start_date, end_date, budget, salary_code, project_id))
        
        return jsonify({'success': True, 'message': 'Project updated successfully'})
        
    except Exception as e:
        print(f"Update project error: {e}")
        return jsonify({'error': 'Failed to update project'}), 500

@app.route('/api/admin/projects/<int:project_id>', methods=['DELETE'])
def delete_project(project_id):
    """Delete a project"""
    try:
        query = "DELETE FROM projects WHERE id = %s"
        execute_query(query, (project_id,))
        
        return jsonify({'success': True, 'message': 'Project deleted successfully'})
        
    except Exception as e:
        print(f"Delete project error: {e}")
        return jsonify({'error': 'Failed to delete project'}), 500

# Ticket Management API Endpoints
@app.route('/api/admin/tickets', methods=['GET'])
def get_tickets():
    """Get all tickets"""
    try:
        query = """
        SELECT t.*, 
               CASE 
                   WHEN t.assigned_to_type = 'employee' THEN u.name
                   WHEN t.assigned_to_type = 'vendor' THEN v.company_name
                   ELSE 'Unassigned'
               END as assigned_to_name,
               CASE 
                   WHEN t.assigned_to_type = 'employee' THEN u.email
                   WHEN t.assigned_to_type = 'vendor' THEN v.email
                   ELSE NULL
               END as assigned_to_email
        FROM tickets t
        LEFT JOIN users u ON t.assigned_to_type = 'employee' AND t.assigned_to_id = u.id
        LEFT JOIN vendors v ON t.assigned_to_type = 'vendor' AND t.assigned_to_id = v.id
        ORDER BY t.created_at DESC
        """
        tickets = execute_query(query, fetch_all=True)
        return jsonify(tickets)
    except Exception as e:
        print(f"Get tickets error: {e}")
        return jsonify({'error': 'Failed to get tickets'}), 500

@app.route('/api/admin/tickets', methods=['POST'])
def create_ticket():
    """Create a new ticket"""
    try:
        data = request.get_json()
        print(f"Received ticket creation data: {data}")
        
        title = data.get('title')
        description = data.get('description', '')
        category = data.get('category', 'general')
        priority = data.get('priority', 'medium')
        status = data.get('status', 'open')
        assigned_to_type = data.get('assigned_to_type')
        assigned_to_id = data.get('assigned_to_id')
        salary_code = data.get('salary_code', '')
        created_by = data.get('created_by')
        
        if not created_by:
            user_query = "SELECT MIN(id) FROM users"
            first_user = execute_query(user_query, fetch_one=True)
            created_by = first_user['MIN(id)'] if first_user and first_user['MIN(id)'] else None
        
        if not created_by:
            return jsonify({'error': 'No valid user found to assign as creator'}), 400
        
        if not title:
            return jsonify({'error': 'Ticket title is required'}), 400
        
        status_map = {
            'open': 'Open',
            'in_progress': 'In Progress', 
            'resolved': 'Resolved',
            'closed': 'Closed'
        }
        
        priority_map = {
            'low': 'Low',
            'medium': 'Medium',
            'high': 'High',
            'urgent': 'Critical'
        }
        
        db_status = status_map.get(status, 'Open')
        db_priority = priority_map.get(priority, 'Medium')
        
        query = """
        INSERT INTO tickets (title, description, category, priority, status, assigned_to_type, assigned_to_id, assigned_to, salary_code, created_by, created_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
        """
        
        assigned_to = assigned_to_id if assigned_to_type == 'employee' else None
        
        print(f"Ticket creation values: assigned_to_type={assigned_to_type}, assigned_to_id={assigned_to_id}, assigned_to={assigned_to}, created_by={created_by}")
        
        execute_query(query, (title, description, category, db_priority, db_status, assigned_to_type, assigned_to_id, assigned_to, salary_code, created_by))
        
        return jsonify({'success': True, 'message': 'Ticket created successfully'})
        
    except Exception as e:
        print(f"Create ticket error: {e}")
        return jsonify({'error': 'Failed to create ticket'}), 500

@app.route('/api/admin/tickets/<int:ticket_id>', methods=['PUT'])
def update_ticket(ticket_id):
    """Update a ticket"""
    try:
        data = request.get_json()
        
        title = data.get('title')
        description = data.get('description', '')
        category = data.get('category', 'general')
        priority = data.get('priority', 'medium')
        status = data.get('status', 'open')
        assigned_to_type = data.get('assigned_to_type')
        assigned_to_id = data.get('assigned_to_id')
        salary_code = data.get('salary_code', '')
        
        if not title:
            return jsonify({'error': 'Ticket title is required'}), 400
        
        status_map = {
            'open': 'Open',
            'in_progress': 'In Progress', 
            'resolved': 'Resolved',
            'closed': 'Closed'
        }
        
        priority_map = {
            'low': 'Low',
            'medium': 'Medium',
            'high': 'High',
            'urgent': 'Critical'
        }
        
        db_status = status_map.get(status, 'Open')
        db_priority = priority_map.get(priority, 'Medium')
        
        query = """
        UPDATE tickets 
        SET title = %s, description = %s, category = %s, priority = %s, status = %s, 
            assigned_to_type = %s, assigned_to_id = %s, salary_code = %s, updated_at = NOW()
        WHERE id = %s
        """
        
        execute_query(query, (title, description, category, db_priority, db_status, assigned_to_type, assigned_to_id, salary_code, ticket_id))
        
        return jsonify({'success': True, 'message': 'Ticket updated successfully'})
        
    except Exception as e:
        print(f"Update ticket error: {e}")
        return jsonify({'error': 'Failed to update ticket'}), 500

@app.route('/api/admin/tickets/<int:ticket_id>', methods=['DELETE'])
def delete_ticket(ticket_id):
    """Delete a ticket"""
    try:
        query = "DELETE FROM tickets WHERE id = %s"
        execute_query(query, (ticket_id,))
        
        return jsonify({'success': True, 'message': 'Ticket deleted successfully'})
        
    except Exception as e:
        print(f"Delete ticket error: {e}")
        return jsonify({'error': 'Failed to delete ticket'}), 500

# Employee routes and additional endpoints continue below...
# (The rest of your existing code for employee notifications, vendor NDA submission, etc.)

@app.route('/api/vendor/submit-nda', methods=['POST'])
def submit_vendor_nda():
    """Submit vendor NDA with signature"""
    try:
        print(f"📝 Vendor NDA submission received")
        print(f"📝 Form data keys: {list(request.form.keys())}")
        
        # Handle FormData (multipart/form-data)
        reference_number = request.form.get('reference_number')
        signature = request.form.get('signature')
        company_name = request.form.get('companyName')
        contact_person = request.form.get('contactPerson')
        email = request.form.get('email')
        phone = request.form.get('phone')
        address = request.form.get('address')
        
        print(f"📝 Reference: {reference_number}, Company: {company_name}, Email: {email}")
        
        # Handle file upload for stamp
        stamp_file = request.files.get('stamp')
        stamp_data = None
        
        if stamp_file and stamp_file.filename:
            # Convert file to base64
            import base64
            stamp_data = base64.b64encode(stamp_file.read()).decode('utf-8')
        
        if not reference_number or not signature:
            return jsonify({'success': False, 'message': 'Reference number and signature are required'}), 400
        
        # Update vendor record with all submitted information
        query = """
        UPDATE vendors 
        SET signature_data = %s, 
            company_stamp_data = %s, 
            company_name = %s,
            contact_person = %s,
            phone = %s,
            address = %s,
            nda_status = 'completed',
            signed_date = NOW()
        WHERE reference_number = %s
        """
        
        print(f"📝 Updating vendor record for reference: {reference_number}")
        execute_query(query, (signature, stamp_data, company_name, contact_person, phone, address, reference_number))
        print(f"✅ Vendor record updated successfully")
        
        # Insert a record into nda_forms to appear in admin list
        try:
            # Find vendor id
            vrow = execute_query("SELECT id FROM vendors WHERE reference_number = %s", (reference_number,), fetch_one=True)
            vendor_id = vrow.get('id') if vrow else None
            if vendor_id:
                import json as _json
                form_payload = {
                    'company_name': company_name,
                    'contact_person': contact_person,
                    'email': email,
                    'phone': phone,
                    'address': address,
                    'signature_data': signature,
                    'company_stamp_data': stamp_data,
                    'signature_type': 'drawn',
                    'reference_number': reference_number
                }
                execute_query(
                    """
                    INSERT INTO nda_forms (vendor_id, form_data, status, signed_at, created_at, updated_at)
                    VALUES (%s, %s, 'signed', NOW(), NOW(), NOW())
                    """,
                    (vendor_id, _json.dumps(form_payload))
                )
            else:
                print(f"⚠️ Vendor not found for reference {reference_number}; skipping nda_forms insert")
        except Exception as insert_err:
            print(f"⚠️ Failed inserting into nda_forms: {insert_err}")

        # Send confirmation email to vendor
        try:
            smtp = get_smtp_settings()
            print(f"📧 Sending confirmation email to {email}")
            
            msg = MIMEMultipart()
            msg['From'] = f'YellowStone XPs <{smtp["smtp_username"]}>'
            msg['To'] = email
            msg['Subject'] = f'NDA Submission Confirmation - {reference_number}'
            
            # Add anti-spam headers
            add_anti_spam_headers(msg, smtp)
            
            body = f"""Dear {contact_person},

This is to confirm that we have successfully received your Non-Disclosure Agreement.

Reference Number: {reference_number}
Company Name: {company_name}
Submitted Date: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}

Your NDA has been submitted and is now under review by our team. You will be notified once the review process is complete and your vendor portal access has been approved.

Thank you for your interest in partnering with YellowStone Xperiences Pvt Ltd.

Best Regards,
YellowStone Xperiences Pvt Ltd
Harpreet Singh - CEO
Email: Harpreet.singh@yellowstonexps.com
Address: Plot # 2, ITC, Fourth Floor, Sector 67, Mohali -160062, Punjab, India"""
            
            msg.attach(MIMEText(body, 'plain'))
            
            if smtp['smtp_port'] == 465:
                server = smtplib.SMTP_SSL(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
            else:
                server = smtplib.SMTP(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
                server.starttls()
            
            server.login(smtp['smtp_username'], smtp['smtp_password'])
            text = msg.as_string()
            server.sendmail(smtp['smtp_username'], email, text)
            server.quit()
            print(f"✅ Confirmation email sent to {email}")
        except Exception as email_error:
            print(f"⚠️ Failed to send confirmation email: {email_error}")
            # Don't fail the NDA submission if email fails
        
        # Broadcast to admin views so NDA list updates in real-time
        try:
            socketio.emit('database_change', {
                'table': 'nda_forms',
                'action': 'create'
            }, room='admin')
        except:
            pass
        return jsonify({'success': True, 'message': 'NDA submitted successfully'})
        
    except Exception as e:
        print(f"Submit vendor NDA error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Failed to submit NDA: {str(e)}'}), 500

# Admin routes
@app.route('/api/admin/nda-forms', methods=['GET'])
def get_admin_nda_forms():
    """Get all NDA forms for admin view"""
    try:
        query = """
        SELECT 
            nf.id,
            JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.company_name')) as company_name,
            JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.contact_person')) as contact_person,
            JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.email')) as email,
            JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.phone')) as phone,
            JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.address')) as address,
            JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.business_type')) as business_type,
            JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.registration_number')) as company_registration_number,
            JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.country')) as company_incorporation_country,
            JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.state')) as company_incorporation_state,
            JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.signature_data')) as signature_data,
            JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.company_stamp_data')) as company_stamp_data,
            JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.signature_type')) as signature_type,
            JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.reference_number')) as reference_number,
            nf.status as nda_status,
            nf.signed_at as signed_date,
            nf.created_at,
            nf.updated_at,
            v.portal_access,
            v.has_full_access,
            CASE 
                WHEN JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.signature_data')) IS NOT NULL 
                     AND JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.signature_data')) != '' 
                     AND JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.signature_data')) != 'null'
                THEN 1
                ELSE 0
            END as has_signature,
            CASE 
                WHEN JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.company_stamp_data')) IS NOT NULL 
                     AND JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.company_stamp_data')) != '' 
                     AND JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.company_stamp_data')) != 'null'
                THEN 1
                ELSE 0
            END as has_company_stamp,
            CASE 
                WHEN v.portal_access = 1 THEN 'Granted'
                ELSE 'Pending'
            END as portal_access_status
        FROM nda_forms nf
        JOIN vendors v ON nf.vendor_id = v.id
        WHERE nf.status = 'signed'
        ORDER BY nf.created_at DESC
        """
        nda_forms = execute_query(query, fetch_all=True)
        
        if nda_forms is None:
            return jsonify([])
        return jsonify(nda_forms)
        
    except Exception as e:
        return jsonify([])

@app.route('/api/admin/nda-forms/<int:form_id>', methods=['DELETE'])
def delete_nda_form(form_id):
    """Delete an NDA form"""
    try:
        admin_id = session.get('user_id')
        if not admin_id or session.get('user_type') != 'admin':
            # Check if user has employee access for NDA
            employee_id = request.args.get('employee_id')
            if not employee_id:
                return jsonify({'error': 'Unauthorized'}), 401
            
            # Check if employee has NDA access
            access_query = """
            SELECT has_access FROM employee_access
            WHERE employee_id = %s AND access_type = 'send_nda'
            """
            access = execute_query(access_query, (employee_id,), fetch_one=True)
            
            if not access or not bool(access.get('has_access', False)):
                return jsonify({'error': 'No NDA access permission'}), 403
        
        # Delete the NDA form
        query = "DELETE FROM nda_forms WHERE id = %s"
        execute_query(query, (form_id,))
        
        return jsonify({'success': True, 'message': 'NDA form deleted successfully'})
        
    except Exception as e:
        # print(f"Error deleting NDA form: {e}")
        return jsonify({'error': 'Failed to delete NDA form'}), 500

@app.route('/api/admin/vendors', methods=['GET'])
def get_admin_vendors():
    """Get all vendors with proper error handling"""
    try:
        vendors = execute_query("SELECT * FROM vendors", fetch_all=True)
        
        if vendors is None:
            return jsonify([])
        
        if len(vendors) == 0:
            return jsonify([])
        return jsonify(vendors)
        
    except Exception as e:
        return jsonify([])

@app.route('/api/admin/vendors/<int:vendor_id>/approve-portal-access', methods=['POST'])
def approve_vendor_portal_access(vendor_id):
    """Approve portal access for a vendor"""
    try:
        print(f"✅ Approving portal access for vendor ID: {vendor_id}")
        
        # Update vendor to grant portal access
        update_query = """
        UPDATE vendors 
        SET portal_access = TRUE,
            registration_status = 'approved',
            updated_at = NOW()
        WHERE id = %s
        """
        
        result = execute_query(update_query, (vendor_id,))
        print(f"✅ Vendor portal access updated")
        
        # Get vendor details for email
        vendor_query = "SELECT company_name, email, contact_person FROM vendors WHERE id = %s"
        vendor = execute_query(vendor_query, (vendor_id,), fetch_one=True)
        
        if vendor:
            # Generate temporary password (8 characters, mix of letters and numbers)
            import secrets
            import string
            
            alphabet = string.ascii_letters + string.digits
            temp_password = ''.join(secrets.choice(alphabet) for i in range(12))
            
            # Hash the password using bcrypt
            hashed_password = bcrypt.hashpw(temp_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            # Calculate expiration date (30 days from now)
            from datetime import datetime, timedelta
            expires_at = datetime.now() + timedelta(days=30)
            
            # Insert temporary login credentials
            print(f"✅ Creating temporary login credentials for vendor {vendor_id}")
            
            # First, deactivate any existing temporary logins for this vendor (if using is_active column)
            # Note: If the table doesn't have is_active, you can DELETE old records instead:
            # delete_query = "DELETE FROM temporary_logins WHERE vendor_id = %s"
            # execute_query(delete_query, (vendor_id,))
            
            # Insert new temporary login
            insert_query = """
            INSERT INTO temporary_logins (vendor_id, email, password_hash, company_name, contact_person, reference_number, created_at, expires_at)
            VALUES (%s, %s, %s, %s, %s, %s, NOW(), %s)
            """
            # Get vendor's reference number
            vendor_details_query = "SELECT reference_number FROM vendors WHERE id = %s"
            vendor_details = execute_query(vendor_details_query, (vendor_id,), fetch_one=True)
            reference_number = vendor_details['reference_number'] if vendor_details else 'N/A'
            
            execute_query(insert_query, (vendor_id, vendor['email'], hashed_password, vendor['company_name'], vendor['contact_person'], reference_number, expires_at))
            
            print(f"✅ Temporary credentials created, expires: {expires_at}")
            print(f"✅ Temporary password: {temp_password}")
            
            print(f"✅ Sending approval email to {vendor['email']}")
            
            # Send approval email
            try:
                smtp = get_smtp_settings()
                msg = MIMEMultipart()
                msg['From'] = smtp['smtp_username']
                msg['To'] = vendor['email']
                msg['Subject'] = 'Vendor Portal Access Approved - YellowStone XPs'
                
                body = f"""Dear {vendor['contact_person']},

We are pleased to inform you that your vendor portal access has been approved!

Company: {vendor['company_name']}
Status: Portal Access Granted

TEMPORARY LOGIN CREDENTIALS (Valid for 30 days):

Username: {vendor['email']}
Password: {temp_password}

IMPORTANT:
- These credentials will expire in 30 days
- Please change your password after your first login
- Use these credentials to access the vendor portal

Portal URL: http://localhost:3000/vendor-portal

If you have any questions, please contact our Vendor Relations Department.

Thank you for partnering with YellowStone Xperiences Pvt Ltd.

Best Regards,
YellowStone Xperiences Pvt Ltd
Harpreet Singh - CEO
Email: Harpreet.singh@yellowstonexps.com"""
                
                msg.attach(MIMEText(body, 'plain'))
                
                if smtp['smtp_port'] == 465:
                    server = smtplib.SMTP_SSL(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
                else:
                    server = smtplib.SMTP(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
                    server.starttls()
                
                server.login(smtp['smtp_username'], smtp['smtp_password'])
                server.sendmail(smtp['smtp_username'], vendor['email'], msg.as_string())
                server.quit()
                print(f"✅ Approval email sent to {vendor['email']}")
            except Exception as email_error:
                print(f"⚠️ Failed to send approval email: {email_error}")
        
        return jsonify({'success': True, 'message': 'Portal access approved successfully'})
        
    except Exception as e:
        print(f"❌ Error approving portal access: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Failed to approve portal access'}), 500

@app.route('/api/admin/templates', methods=['GET'])
def get_admin_templates():
    """Get all form templates for admin view"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500
        
        cursor = connection.cursor()
        cursor.execute("""
            SELECT id, name, description, template_type, category, priority, 
                   form_fields, status, created_at, created_by_name
            FROM form_templates 
            ORDER BY created_at DESC
        """)
        
        templates = []
        for row in cursor.fetchall():
            templates.append({
                'id': row[0],
                'name': row[1],
                'description': row[2],
                'template_type': row[3],
                'category': row[4],
                'priority': row[5],
                'form_fields': row[6],
                'status': row[7],
                'created_at': row[8].isoformat() if row[8] else None,
                'created_by_name': row[9]
            })
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'templates': templates
        })
    except Exception as e:
        print(f"Error fetching templates: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/admin/templates', methods=['POST'])
def create_template():
    """Create a new form template"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['name', 'template_type']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500
        
        cursor = connection.cursor()
        
        # Insert new template
        insert_query = """
            INSERT INTO form_templates 
            (name, description, template_type, category, priority, form_fields, status, created_at, created_by_name)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), %s)
        """
        
        cursor.execute(insert_query, (
            data.get('name'),
            data.get('description', ''),
            data.get('template_type'),
            data.get('category', 'general'),
            data.get('priority', 'medium'),
            json.dumps(data.get('form_fields', {})),
            data.get('status', 'draft'),
            data.get('created_by_name', 'Admin')
        ))
        
        template_id = cursor.lastrowid
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': 'Template created successfully',
            'template_id': template_id
        })
        
    except Exception as e:
        print(f"Error creating template: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/admin/templates/<int:template_id>', methods=['DELETE'])
def delete_template(template_id):
    """Delete a form template"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500
        
        cursor = connection.cursor()
        
        # Check if template exists
        cursor.execute("SELECT id FROM form_templates WHERE id = %s", (template_id,))
        if not cursor.fetchone():
            cursor.close()
            connection.close()
            return jsonify({'success': False, 'error': 'Template not found'}), 404
        
        # Delete template
        cursor.execute("DELETE FROM form_templates WHERE id = %s", (template_id,))
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': 'Template deleted successfully'
        })
        
    except Exception as e:
        print(f"Error deleting template: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/admin/submitted-nda-forms', methods=['GET'])
def get_submitted_nda_forms():
    """Get submitted NDA forms for admin dashboard"""
    try:
        query = """
        SELECT v.id, v.company_name, v.contact_person, v.email, v.phone, v.address,
               v.nda_status, v.reference_number, v.signature_data, v.company_stamp_data,
               v.signature_type, v.signed_date, v.created_at, v.updated_at
        FROM vendors v
        WHERE v.nda_status IN ('completed', 'sent')
        ORDER BY v.updated_at DESC
        """
        nda_forms = execute_query(query, fetch_all=True)
        print(f"📊 Found {len(nda_forms) if nda_forms else 0} submitted NDA forms")
        return jsonify(nda_forms or [])
        
    except Exception as e:
        print(f"❌ Error getting submitted NDA forms: {e}")
        import traceback
        traceback.print_exc()
        return jsonify([])

@app.route('/api/admin/download-nda/<reference_number>', methods=['GET'])
def download_nda_pdf(reference_number):
    """Download NDA PDF by reference number"""
    try:
        # Get NDA form by reference number
        query = """
        SELECT nf.id, nf.form_data, nf.signed_at, nf.vendor_id
        FROM nda_forms nf
        WHERE JSON_UNQUOTE(JSON_EXTRACT(nf.form_data, '$.reference_number')) = %s
        LIMIT 1
        """
        nda_form = execute_query(query, (reference_number,), fetch_one=True)
        
        if not nda_form:
            return jsonify({'error': 'NDA not found'}), 404
        
        # Parse form data
        form_data = {}
        if nda_form['form_data']:
            try:
                form_data = json.loads(nda_form['form_data']) if isinstance(nda_form['form_data'], str) else nda_form['form_data']
            except:
                form_data = {}
        
        vendor_id = nda_form['vendor_id']
        
        # Get vendor data
        vendor_query = """
        SELECT company_name, contact_person, email, phone, address, 
               business_type, reference_number, signature_type,
               signature_data, company_stamp_data,
               CASE WHEN signature_data IS NOT NULL AND signature_data != '' THEN 1 ELSE 0 END as has_signature,
               CASE WHEN company_stamp_data IS NOT NULL AND company_stamp_data != '' THEN 1 ELSE 0 END as has_stamp
        FROM vendors
        WHERE id = %s
        """
        vendor_data = execute_query(vendor_query, (vendor_id,), fetch_one=True)
        
        if not vendor_data:
            return jsonify({'error': 'Vendor not found'}), 404
        
        vendor_data['signed_date'] = nda_form['signed_at']
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title and non-table header placements
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=12, alignment=1)
        story.append(Paragraph("NON-DISCLOSURE AGREEMENT", title_style))
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        story.append(Paragraph(f"Reference Number: {vendor_data['reference_number'] or reference_number}", ParagraphStyle('left', parent=styles['Normal'], alignment=0)))
        story.append(Paragraph(f"Vendor Email: {vendor_data['email'] or 'N/A'}", ParagraphStyle('right', parent=styles['Normal'], alignment=2)))
        story.append(Spacer(1, 10))
        
        # No top signature; it will be shown near the bottom

        # Agreement clauses – full template text and improved readability
        body = ParagraphStyle('Body', parent=styles['Normal'], fontSize=11, leading=14)
        story.append(Paragraph("<b>Confidentiality & Nondisclosure Agreement (NDA)</b>", styles['Heading2']))
        story.append(Spacer(1, 6))
        today_str = datetime.now().strftime('%d, %B, %Y')
        story.append(Paragraph(
            f"This agreement is entered into on {today_str}, by and among:", body))
        story.append(Spacer(1, 4))
        story.append(Paragraph(
            "(1) YellowStone  Xperiences  Pvt  Ltd,  a  company incorporated and existing under the laws of Company Act, India, having its registered office at Plot # 2, ITC, Fourth Floor, Sector 67, Mohali -160062, Punjab, India, registered with the company registration number U72900PB2020PTC051260, hereinafter referred to as ‘YSXP’, And",
            body
        ))
        story.append(Paragraph(
            f"(2) {vendor_data.get('company_name') or '_____________________________________________'}, a company incorporated and existing under the laws of Company Act, India, having its registered office at {vendor_data.get('address') or '________________ _________________________________________'}, registered with the company registration number {'_______________________________________'}, hereinafter referred to as ‘{vendor_data.get('company_name') or '________'}’.",
            body
        ))
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            "hereinafter also referred to individually as a ‘Party’ and collectively as the ‘Parties’.",
            body
        ))
        story.append(Spacer(1, 10))

        def section(title, text):
            story.append(Paragraph(f"<b>{title}</b>", body))
            story.append(Paragraph(text, body))
            story.append(Spacer(1, 6))

        section("1. PURPOSE",
            "In connection with discussions between YellowStone Xperiences Pvt Ltd and _______________ concerning Company’s Local representative under capacity of Strategic Partnership for cross selling of products and services. In order to pursue this Purpose, the Parties recognize that there is a need to disclose certain confidential information by the Disclosing Party (i.e., Party disclosing such information) to Receiving Party (i.e., Party receiving such information) to be used only for this Purpose and to protect such confidential information from unauthorized use and disclosure. The Disclosing Party has agreed to disclose to the Receiving Party the confidential information on a strictly confidential basis.")

        section("2. DEFINITION OF CONFIDENTIAL INFORMATION",
            "‘Confidential Information’ shall mean all information, whether disclosed before or after the Effective Date, that is disclosed in written, oral, electronic, visual or other form by either party (each, as a ‘Disclosing Party’) to the other party (each, as a ‘Receiving Party’) and either (i) marked or designated as ‘confidential’ or ‘proprietary’ at the time of disclosure or (ii) otherwise clearly indicated to be confidential at the time of disclosure. Confidential Information may include, without limitation, computer programs, software or hardware products, product development plans, drawings, models, proposed services, discoveries, ideas, concepts, equipment, code, software source documents and related technical information, algorithms, know‑how, trade secrets, formulas, processes, procedures, research, inventions (whether patentable or not), copyrights, schematics and other technical, names and expertise of employees and consultants and customer or partner information.")

        section("3. CONFIDENTIALITY OBLIGATION",
            "Receiving Party agrees to protect the Confidential Information by using the same degree of care as Receiving Party uses to protect its own confidential or proprietary information (but not less than a reasonable degree of care): (i) to prevent the unauthorized use, dissemination or publication of the Confidential Information (ii) not to divulge Confidential Information to any third party, (iii) not to make any use of such Confidential Information except for the Business Purpose, and (iv) not to copy except as reasonably required in direct support of the Business Purpose. Any copies made will include appropriate marking identifying same as constituting or containing Confidential Information of Disclosing Party; and (v) not to reverse engineer any such Confidential Information. Receiving Party shall limit the use of and access to Disclosing Party’s Confidential Information to Receiving Party’s employees and to the employees of Receiving Party’s respective parent, subsidiaries and affiliated entities or authorized representatives who have: (i) a need to know and have been notified that such information is Confidential Information to be used solely for the Business Purpose; and (ii) entered into binding confidentiality obligations no less protective of Disclosing Party than those contained in this Agreement. Receiving Party may disclose Confidential Information pursuant to any statutory or regulatory authority or court order, provided Disclosing Party is given prompt prior written notice of such requirement and the scope of such disclosure is limited to the extent possible.")

        section("4. NON-DISCLOSURE OF CONFIDENTIAL INFORMATION",
            "Receiving Party shall only permit access to Confidential Information to those of its employees or authorized representatives who have a need to know and shall not disclose it to third parties.")

        section("5. RETURN OF CONFIDENTIAL INFORMATION",
            "All Confidential Information furnished under this Agreement, shall remain the property of the Disclosing Party and shall be returned to it without retaining any copies or it shall be destroyed all documents and other materials promptly upon the termination of this Agreement under the consent of both parties or any breach of contract made by either party.")

        section("6. TERM",
            "This Agreement shall come into force from the effective date of this deed and shall continue for two (2) years from the Effective Date, unless earlier terminated by either Party with 30 days’ prior written notice.")

        section("7. GOVERNING LAW",
            "This Agreement shall be governed in all respects in accordance with the laws of Mohali, Punjab, India.")

        section("8. AMENDMENTS, FILLING, CANCELLATION, TERMINATION AND RENEWAL OF AGREEMENT",
            "Any term of this Agreement may be amended, filled, cancelled, terminated and renewed under the written mutual consent of the parties. This said all matters shall be done by giving 30 days advance written notice each other.")

        section("9. REMEDIES",
            "If a party breaches this NDA, this party must be taken the action under the laws of Mohali, Punjab, India. However, on mutual agreement of both parties the geographical location can be changed and accordingly the law of land would be followed. The parties agree that if it is in any breach of this Agreement, including without limiting any actual or threatened disclosure of Confidential Information without the prior expressed written consent of the disclosing party for which no adequate remedy under law exists, the injured party shall have right not only to terminate the Agreement but also the entitlement for equitable remedies, injunctive relief and specific performance to redress any breach or threatened breach of this Agreement made directly and indirectly by the receiving party or any of its advisors or Affiliates, or any other persons acting for or on behalf of or with the receiving party.")

        section("10. VALID NOTICES",
            "All notices under this Agreement shall be deemed to have been duly given upon the mailing of the notice, in person or by courier, sent by fax, e‑mail to the Party entitled to such notice at the address.")
        story.append(Spacer(1, 10))
        
        # Helper to clean base64 strings
        def _clean_b64(b64s):
            if not b64s:
                return None
            s = b64s
            if isinstance(s, bytes):
                s = s.decode('utf-8', errors='ignore')
            if ',' in s and s.strip().startswith('data:'):
                s = s.split(',', 1)[1]
            missing = len(s) % 4
            if missing:
                s += '=' * (4 - missing)
            return s

        # Signatures table with vendor signature inside right column
        story.append(Paragraph("Signatures:", styles['Heading2']))
        story.append(Spacer(1, 8))
        try:
            def _clean_b64(b64s):
                if not b64s:
                    return None
                s = b64s
                if isinstance(s, bytes):
                    s = s.decode('utf-8', errors='ignore')
                if ',' in s and s.strip().startswith('data:'):
                    s = s.split(',', 1)[1]
                missing = len(s) % 4
                if missing:
                    s += '=' * (4 - missing)
                return s

            sig_b64 = _clean_b64(form_data.get('signature_data') or vendor_data.get('signature_data'))
            sig_img_el = ''
            if sig_b64:
                import base64
                sig_img_el = Image(BytesIO(base64.b64decode(sig_b64)), width=2.5*inch, height=0.9*inch)
            sig_table = Table([
                ['For YSXP:', 'For Vendor:'],
                ['Signature:', sig_img_el if sig_img_el else 'Signature: _____________________________'],
                ['Name:', vendor_data.get('contact_person') or ''],
                ['Title:', ''],
                ['Date:', vendor_data['signed_date'].strftime('%Y-%m-%d') if vendor_data.get('signed_date') else ''],
            ], colWidths=[3*inch, 3*inch])
            sig_table.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
            ]))
            story.append(sig_table)
            story.append(Spacer(1, 10))
        except Exception as _e:
            print(f"Download signature render error: {_e}")

        # Company Information as lines (no table)
        info_style = ParagraphStyle('info', parent=styles['Normal'], fontSize=10, leading=13)
        story.append(Paragraph(f"Company Name: {vendor_data['company_name'] or 'N/A'}", info_style))
        story.append(Paragraph(f"Contact Person: {vendor_data['contact_person'] or 'N/A'}", info_style))
        story.append(Paragraph(f"Email: {vendor_data['email'] or 'N/A'}", info_style))
        story.append(Paragraph(f"Phone: {vendor_data['phone'] or 'N/A'}", info_style))
        story.append(Paragraph(f"Address: {vendor_data['address'] or 'N/A'}", info_style))
        story.append(Paragraph(f"Business Type: {vendor_data['business_type'] or 'N/A'}", info_style))
        story.append(Paragraph(f"Signed Date: {vendor_data['signed_date'].strftime('%Y-%m-%d') if vendor_data.get('signed_date') else 'N/A'}", info_style))
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"Timestamp: {ts}", ParagraphStyle('footer_ts', parent=styles['Normal'], alignment=2)))
        
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'NDA_{reference_number}.pdf'
        )
        
    except Exception as e:
        # print(f"Error generating PDF: {e}")
        return jsonify({'error': 'Failed to generate PDF'}), 500

@app.route('/api/admin/send-completed-nda-email', methods=['POST'])
def send_completed_nda_email():
    """Send completion email with attached NDA PDF."""
    try:
        data = request.get_json() or {}
        form_id = data.get('formId')
        reference_number = data.get('reference_number')

        vendor = None
        if form_id:
            vendor = execute_query("SELECT * FROM vendors WHERE id = %s", (form_id,), fetch_one=True)
        elif reference_number:
            vendor = execute_query("SELECT * FROM vendors WHERE reference_number = %s", (reference_number,), fetch_one=True)
        else:
            return jsonify({'success': False, 'message': 'Form ID or reference_number is required'}), 400
        if not vendor:
            return jsonify({'success': False, 'message': 'Vendor not found'}), 404

        # Try to fetch nda_forms row for richer data
        form_json = {}
        if vendor and vendor.get('reference_number'):
            try:
                nf = execute_query(
                    """
                    SELECT form_data FROM nda_forms
                    WHERE JSON_UNQUOTE(JSON_EXTRACT(form_data, '$.reference_number')) = %s
                    ORDER BY id DESC LIMIT 1
                    """,
                    (vendor['reference_number'],), fetch_one=True
                )
                if nf and nf.get('form_data'):
                    form_json = json.loads(nf['form_data']) if isinstance(nf['form_data'], str) else nf['form_data']
            except Exception as _e:
                print(f"Failed to fetch nda_forms row: {_e}")

        # Build NDA PDF in-memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        # Header with reference (left) and email (right)
        story.append(Paragraph("NON-DISCLOSURE AGREEMENT", styles['Title']))
        story.append(Spacer(1, 6))
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        left = Paragraph(f"Reference Number: {vendor.get('reference_number') or 'N/A'}", ParagraphStyle('left', parent=styles['Normal'], alignment=0))
        right = Paragraph(f"Vendor Email: {vendor.get('email') or 'N/A'}", ParagraphStyle('right', parent=styles['Normal'], alignment=2))
        story.append(left)
        story.append(right)
        story.append(Spacer(1, 12))

        # No top signature; will render near bottom below

        # NDA body – formatted per requested template
        story.append(Paragraph("<b>Confidentiality & Nondisclosure Agreement (NDA)</b>", styles['Heading2']))
        story.append(Spacer(1, 6))
        today_str = datetime.now().strftime('%d, %B, %Y')
        intro = (
            f"This agreement is entered into on {today_str}, by and among:<br/><br/>"
            f"(1) YellowStone Xperiences Pvt Ltd, a company incorporated and existing under the laws of Company Act, India, having its registered office at Plot # 2, ITC, Fourth Floor, Sector 67, Mohali -160062, Punjab, India, registered with the company registration number U72900PB2020PTC051260, hereinafter referred to as ‘YSXP’, And<br/>"
            f"(2) {vendor.get('company_name') or '________________'}, a company incorporated and existing under the laws of Company Act, India, having its registered office at {vendor.get('address') or '________________'}, registered with the company registration number {'________________'}, hereinafter referred to as ‘{vendor.get('company_name') or '________'}’.<br/><br/>"
            "Hereinafter also referred to individually as a ‘Party’ and collectively as the ‘Parties’."
        )
        story.append(Paragraph(intro, styles['Normal']))
        story.append(Spacer(1, 8))

        sections = [
            ("1. PURPOSE", "In connection with discussions between YellowStone Xperiences Pvt Ltd and the Vendor concerning Company’s local representative under capacity of Strategic Partnership for cross selling of products and services. In order to pursue this Purpose, the Parties recognize that there is a need to disclose certain confidential information by the Disclosing Party to the Receiving Party to be used only for this Purpose and to protect such confidential information from unauthorized use and disclosure. The Disclosing Party has agreed to disclose to the Receiving Party the confidential information on a strictly confidential basis."),
            ("2. DEFINITION OF CONFIDENTIAL INFORMATION", "‘Confidential Information’ shall mean all information, whether disclosed before or after the Effective Date, that is disclosed in written, oral, electronic, visual or other form by either party (each, as a ‘Disclosing Party’) to the other party (each, as a ‘Receiving Party’) and either (i) marked or designated as ‘confidential’ or ‘proprietary’ at the time of disclosure or (ii) otherwise clearly indicated to be confidential at the time of disclosure. Confidential Information may include, without limitation, computer programs, products, development plans, drawings, models, proposed services, ideas, concepts, processes, research, inventions, trade secrets, formulas, procedures, names and expertise of employees and consultants, and customer or partner information."),
            ("3. CONFIDENTIALITY OBLIGATION", "Receiving Party agrees to protect the Confidential Information by using the same degree of care as it uses for its own confidential or proprietary information (but not less than a reasonable degree of care), not to divulge Confidential Information to any third party, not to use such Confidential Information except for the Business Purpose, to copy only as reasonably required with appropriate markings, and not to reverse engineer any such Confidential Information. Access shall be limited to those with a need-to-know under obligations no less protective than this Agreement. Required disclosures by law are permitted with prompt notice and limited scope."),
            ("4. NON-DISCLOSURE OF CONFIDENTIAL INFORMATION", "Receiving Party shall only permit access to Confidential Information to those of its employees or authorized representatives who have a need to know and shall not disclose it to third parties."),
            ("5. RETURN OF CONFIDENTIAL INFORMATION", "All Confidential Information furnished under this Agreement shall remain the property of the Disclosing Party and shall be returned or destroyed (without retaining copies) promptly upon termination of this Agreement or any breach, with consent of both parties."),
            ("6. TERM", "This Agreement shall come into force from the effective date of this deed and shall continue for two (2) years from the Effective Date, unless earlier terminated by either Party with 30 days’ prior written notice."),
            ("7. GOVERNING LAW", "This Agreement shall be governed in all respects in accordance with the laws of Mohali, Punjab, India."),
            ("8. AMENDMENTS, FILLING, CANCELLATION, TERMINATION AND RENEWAL OF AGREEMENT", "Any term of this Agreement may be amended, filled, cancelled, terminated and renewed under the written mutual consent of the Parties with 30 days’ advance written notice."),
            ("9. REMEDIES", "For any breach, the injured party shall have the right to equitable remedies including injunctive relief and specific performance, in addition to termination and damages, in Mohali, Punjab, India (or as mutually agreed)."),
            ("10. VALID NOTICES", "All notices shall be deemed duly given when sent by post, courier, fax or e‑mail to the addresses below."),
        ]
        for title, text in sections:
            story.append(Paragraph(f"<b>{title}</b>", styles['Normal']))
            story.append(Paragraph(text, styles['Normal']))
            story.append(Spacer(1, 6))

        notice_info = [
            ['If to Yellowstone Xperiences Private Limited:', ''],
            ['Postal address:', 'Plot # 2, ITC, Fourth Floor, Sector 67, Mohali, Punjab, India (160062).'],
            ['E-mail:', 'hello@yellowstonexps.com, Harpreet.singh@yellowstonexps.com'],
            [f"If to {vendor.get('company_name') or 'Vendor'}:", ''],
            ['Postal address:', vendor.get('address') or '________________'],
            ['Phone:', vendor.get('phone') or '________________'],
            ['E-mail:', vendor.get('email') or '________________'],
        ]
        notice_table = Table(notice_info, colWidths=[2.2*inch, 3.8*inch])
        notice_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
            ('FONTSIZE', (0,0), (-1,-1), 9)
        ]))
        story.append(notice_table)
        story.append(Spacer(1, 8))

        story.append(Paragraph('<b>11. DISPUTE SETTLEMENT</b>', body))
        story.append(Paragraph('If any dispute arises relating to non‑disclosure of confidential information, the parties shall perform by negotiation at first. If unresolved, the injured party may sue for compensation in the court.', body))
        story.append(Spacer(1, 8))

        story.append(Paragraph('IN WITNESS WHEREOF, the parties have executed this Agreement on the date first written above by their fully authorized representatives with their free consent.', styles['Normal']))
        story.append(Spacer(1, 8))

        sig_lines = [
            ['For YSXP:', 'For Vendor:'],
            ['Signature:', 'Signature:'],
            ['Name:', vendor.get('contact_person') or ''],
            ['Title:', ''],
            ['Date:', today_str],
        ]
        sig_table = Table(sig_lines, colWidths=[3*inch, 3*inch])
        sig_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTSIZE', (0,0), (-1,-1), 9)
        ]))
        story.append(sig_table)
        
        # Helper to clean base64 strings
        def _clean_b64(b64s):
            if not b64s:
                return None
            s = b64s
            if isinstance(s, bytes):
                s = s.decode('utf-8', errors='ignore')
            if ',' in s and s.strip().startswith('data:'):
                s = s.split(',', 1)[1]
            missing = len(s) % 4
            if missing:
                s += '=' * (4 - missing)
            return s

        # Signatures block as two-column table; vendor signature in the right column
        try:
            sig_b64 = _clean_b64(form_json.get('signature_data') or vendor.get('signature_data'))
            sig_img_el = ''
            if sig_b64:
                import base64
                sig_bytes = base64.b64decode(sig_b64)
                sig_io = BytesIO(sig_bytes)
                sig_img_el = Image(sig_io, width=2.5*inch, height=0.9*inch)

            story.append(Paragraph('Signatures', styles['Heading2']))
            story.append(Spacer(1, 6))
            sig_table = Table([
                ['For YSXP:', 'For Vendor:'],
                ['Signature:', sig_img_el if sig_img_el else 'Signature: _____________________________'],
                ['Name:', vendor.get('contact_person') or ''],
                ['Title:', ''],
                ['Date:', datetime.now().strftime('%Y-%m-%d')],
            ], colWidths=[3*inch, 3*inch])
            sig_table.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
            ]))
            story.append(sig_table)
            story.append(Spacer(1, 8))
        except Exception as _img_err:
            print(f"Signature render error: {_img_err}")

        # Company details in simple lines (no table)
        info_style = ParagraphStyle('info', parent=styles['Normal'], fontSize=10, leading=13)
        story.append(Paragraph(f"Company Name: {vendor.get('company_name') or 'N/A'}", info_style))
        story.append(Paragraph(f"Contact Person: {vendor.get('contact_person') or 'N/A'}", info_style))
        story.append(Paragraph(f"Phone: {vendor.get('phone') or 'N/A'}", info_style))
        story.append(Paragraph(f"Address: {vendor.get('address') or 'N/A'}", info_style))

        # Footer timestamp right-aligned
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"Timestamp: {ts}", ParagraphStyle('footer_ts', parent=styles['Normal'], alignment=2)))

        doc.build(story)
        buffer.seek(0)

        # Prepare email with attachment using configured SMTP
        smtp = get_smtp_settings()
        if not vendor.get('email'):
            return jsonify({'success': False, 'message': 'Vendor email is missing'}), 400
        msg = MIMEMultipart()
        msg['From'] = f"YellowStone XPs <{smtp['smtp_username']}>"
        msg['To'] = vendor['email']
        # Send a copy to the sender for verification
        msg['Bcc'] = smtp['smtp_username']
        # Add anti-spam headers used elsewhere in the app
        try:
            add_anti_spam_headers(msg, smtp)
        except Exception:
            pass
        msg['Subject'] = f"Completed NDA - {vendor['company_name']} (Ref: {vendor['reference_number']})"

        body = f"""Dear {vendor.get('contact_person') or 'Vendor'},

Please find attached the completed NDA PDF containing your submitted details.

Reference Number: {vendor.get('reference_number')}
Timestamp: {ts}

Regards,
YellowStone Xperiences Pvt Ltd
"""
        msg.attach(MIMEText(body, 'plain'))

        from email.mime.application import MIMEApplication
        attachment = MIMEApplication(buffer.read(), _subtype='pdf')
        attachment.add_header('Content-Disposition', 'attachment', filename=f"NDA_{vendor.get('reference_number')}.pdf")
        msg.attach(attachment)

        try:
            recipients = [vendor['email']]
            if smtp.get('smtp_username'):
                recipients.append(smtp['smtp_username'])
            if smtp['smtp_port'] == 465:
                server = smtplib.SMTP_SSL(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
            else:
                server = smtplib.SMTP(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
                server.starttls()
            server.login(smtp['smtp_username'], smtp['smtp_password'])
            try:
                server.set_debuglevel(1)
            except Exception:
                pass
            result = server.sendmail(smtp['smtp_username'], recipients, msg.as_string())
            server.quit()
            if result:
                # result is a dict of {recipient: error}
                return jsonify({'success': False, 'message': f'Unsent recipients: {result}'}), 500
        except Exception as mail_err:
            print(f"Completed NDA email send error: {mail_err}")
            return jsonify({'success': False, 'message': f'Email send failed: {mail_err}'}), 500

        return jsonify({'success': True, 'message': 'Completion email with PDF sent successfully'})

    except Exception as e:
        print(f"Send completed NDA email error: {e}")
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'message': 'Failed to send email'}), 500

@app.route('/api/admin/send-bulk-nda', methods=['POST'])
def send_bulk_nda():
    """Send NDA to multiple vendors"""
    try:
        print("ðŸ” Bulk NDA request received")
        data = request.get_json()
        print(f"ðŸ“Š Data received: {data}")
        
        # Handle both old and new data structures
        vendors = data.get('vendors', [])
        selected_vendors = data.get('selected_vendors', [])
        new_vendors = data.get('new_vendors', [])
        
        print(f"ðŸ“‹ Vendors count: {len(vendors)}")
        print(f"ðŸ“‹ Selected vendors count: {len(selected_vendors)}")
        print(f"ðŸ“‹ New vendors count: {len(new_vendors)}")
        
        # If using selected_vendors format, fetch vendor details from database
        if selected_vendors and not vendors:
            print("ðŸ”„ Fetching vendor details for selected IDs...")
            vendor_ids_str = ','.join(map(str, selected_vendors))
            vendor_query = f"SELECT id, email, company_name FROM vendors WHERE id IN ({vendor_ids_str})"
            vendor_details = execute_query(vendor_query, fetch_all=True)
            vendors = vendor_details if vendor_details else []
            print(f"ðŸ“‹ Fetched {len(vendors)} vendor details")
        
        # Add new vendors to the list
        if new_vendors:
            vendors.extend(new_vendors)
        
        print(f"ðŸ“‹ Total vendors to process: {len(vendors)}")
        
        if not vendors:
            print("âŒ No vendors selected")
            return jsonify({'success': False, 'message': 'No vendors selected'}), 400
        
        # Process emails in background thread for immediate response
        email_thread = threading.Thread(target=send_bulk_nda_background_worker, args=(vendors,))
        email_thread.daemon = True
        email_thread.start()
        
        # Return immediately - UI responds instantly
        return jsonify({
            'success': True, 
            'message': f'Processing {len(vendors)} NDAs', 
            'vendor_count': len(vendors)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': 'Failed to process bulk NDA'}), 500

def send_bulk_nda_background_worker(vendors):
    """Background worker to send emails"""
    try:
        print(f"🔵 Bulk NDA worker started with {len(vendors)} vendors")
        # Fetch current SMTP settings
        smtp = get_smtp_settings()
        print(f"🔵 SMTP Config: Server={smtp['smtp_server']}, Port={smtp['smtp_port']}, From={smtp['smtp_username']}")
        
        # Create single SMTP connection for all emails
        if smtp['smtp_port'] == 465:
            print("Using SSL connection for port 465...")
            server = smtplib.SMTP_SSL(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
        else:
            print(f"Using STARTTLS for port {smtp['smtp_port']}...")
            server = smtplib.SMTP(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
            server.starttls()
        
        print("Logging in...")
        server.login(smtp['smtp_username'], smtp['smtp_password'])
        print(f"✅ Logged in as {smtp['smtp_username']}")
        
        for i, vendor in enumerate(vendors, 1):
            try:
                print(f"📧 Processing vendor {i}/{len(vendors)}: {vendor['email']}")
                reference_number = generate_reference_number()
                
                # Update or create vendor record
                check_query = "SELECT id FROM vendors WHERE email = %s"
                existing_vendor = execute_query(check_query, (vendor['email'],), fetch_one=True)
                
                if existing_vendor:
                    update_query = """
                    UPDATE vendors SET company_name = %s, reference_number = %s, nda_status = 'sent', updated_at = NOW()
                    WHERE email = %s
                    """
                    execute_query(update_query, (vendor['company_name'], reference_number, vendor['email']))
                else:
                    insert_query = """
                    INSERT INTO vendors (email, company_name, nda_status, reference_number, created_at)
                    VALUES (%s, %s, 'sent', %s, NOW())
                    """
                    execute_query(insert_query, (vendor['email'], vendor['company_name'], reference_number))
                
                # Send email
                try:
                    msg = MIMEMultipart()
                    msg['From'] = f'YellowStone XPs <{smtp["smtp_username"]}>'
                    msg['To'] = vendor['email']
                    msg['Subject'] = f"NDA Agreement Request - {vendor['company_name']} (Ref: {reference_number})"
                    
                    # Add anti-spam headers
                    add_anti_spam_headers(msg, smtp)
                    
                    body = f"""
Dear Sir/Madam,

Subject: Invitation to Complete Non-Disclosure Agreement - Strategic Partnership Opportunity

YellowStone Xperiences Pvt Ltd is pleased to extend an invitation to {vendor['company_name']} to participate in our Strategic Partnership Program.

REFERENCE NUMBER: {reference_number}

To facilitate this process, we have prepared a comprehensive NDA form that can be completed electronically through our secure portal:

NDA FORM: http://localhost:8000/nda-form?ref={reference_number}

IMPORTANT INSTRUCTIONS:
- Please complete all required fields accurately
- Ensure all company information is current and verified
- Digital signature is required for document execution
- Company stamp upload is optional but recommended

Should you have any questions or require assistance, please contact our Vendor Relations Department.

We look forward to establishing a mutually beneficial partnership.

Yours faithfully,

Harpreet Singh
CEO
YellowStone Xperiences Pvt Ltd
                    """
                    
                    msg.attach(MIMEText(body, 'plain'))
                    
                    text = msg.as_string()
                    print(f"  Sending email to {vendor['email']}...")
                    server.sendmail(smtp['smtp_username'], vendor['email'], text)
                    print(f"  ✅ Email sent to {vendor['email']}")
                    
                except Exception as email_error:
                    print(f"  ❌ Email failed for {vendor['email']}: {email_error}")
                    
            except Exception as vendor_error:
                print(f"  ❌ Vendor processing failed for {vendor.get('email', 'unknown')}: {vendor_error}")
        
        # Close SMTP connection
        server.quit()
        print(f"✅ Bulk NDA sending completed for {len(vendors)} vendors")
        
    except Exception as e:
        print(f"❌ Bulk NDA worker failed: {e}")
        import traceback
        traceback.print_exc()

@app.route('/api/admin/notifications', methods=['GET'])
def get_admin_notifications():
    """Get notifications for admin dashboard"""
    try:
        query = """
        SELECT id, title, message, type, is_read, created_at
        FROM notifications
        ORDER BY created_at DESC
        LIMIT 50
        """
        notifications = execute_query(query, fetch_all=True)
        
        return jsonify({
            'success': True,
            'notifications': notifications or []
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'notifications': []
        })

@app.route('/api/webhook/google-form', methods=['POST'])
def google_form_webhook():
    """Receive Google Form submissions and store them as NDA forms"""
    try:
        data = request.get_json()
        
        # Extract form data
        reference_number = data.get('reference_number', '')
        company_name = data.get('company_name', '')
        contact_person = data.get('contact_person', '')
        email = data.get('email', '')
        phone = data.get('phone', '')
        address = data.get('address', '')
        business_type = data.get('business_type', '')
        registration_number = data.get('registration_number', '')
        country = data.get('country', '')
        state = data.get('state', '')
        signature_data = data.get('signature_data', '')
        company_stamp_data = data.get('company_stamp_data', '')
        signature_type = data.get('signature_type', 'draw')
        # Handle timestamp properly
        timestamp_str = data.get('timestamp', '')
        if timestamp_str:
            try:
                # Parse ISO format timestamp from frontend
                submission_timestamp = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            except:
                submission_timestamp = datetime.now()
        else:
            submission_timestamp = datetime.now()
        
        print(f"Received Google Form submission for reference: {reference_number}")
        print(f"Timestamp received: {timestamp_str}")
        print(f"Parsed timestamp: {submission_timestamp}")
        
        # Find existing vendor by reference number
        vendor_query = "SELECT id FROM vendors WHERE reference_number = %s"
        vendor = execute_query(vendor_query, (reference_number,), fetch_one=True)
        
        if not vendor:
            print(f"âŒ Vendor not found for reference: {reference_number}")
            return jsonify({'success': False, 'message': 'Vendor not found'}), 404
        
        vendor_id = vendor[0]
        
        # Prepare form data as JSON
        form_data = {
            "company_name": company_name,
            "contact_person": contact_person,
            "email": email,
            "phone": phone,
            "address": address,
            "business_type": business_type,
            "registration_number": registration_number,
            "country": country,
            "state": state,
            "signature_data": signature_data,
            "company_stamp_data": company_stamp_data,
            "signature_type": signature_type,
            "reference_number": reference_number
        }
        
        # Check if NDA form already exists for this vendor
        existing_nda_query = "SELECT id FROM nda_forms WHERE vendor_id = %s"
        existing_nda = execute_query(existing_nda_query, (vendor_id,), fetch_one=True)
        
        if existing_nda:
            # Update existing NDA form
            update_query = """
            UPDATE nda_forms 
            SET form_data = %s, status = 'signed', signed_at = %s, updated_at = NOW()
            WHERE vendor_id = %s
            """
            
            try:
                execute_query(update_query, (
                    json.dumps(form_data), submission_timestamp, vendor_id
                ))
                print(f"âœ… NDA form updated for vendor {vendor_id}")
            except Exception as db_error:
                print(f"âŒ Database update error: {db_error}")
                return jsonify({'success': False, 'message': f'Database error: {str(db_error)}'}), 500
        else:
            # Create new NDA form
            insert_query = """
            INSERT INTO nda_forms (vendor_id, form_data, status, signed_at, created_at, updated_at)
            VALUES (%s, %s, 'signed', %s, NOW(), NOW())
            """
            
            try:
                execute_query(insert_query, (
                    vendor_id, json.dumps(form_data), submission_timestamp
                ))
                print(f"âœ… New NDA form created for vendor {vendor_id}")
            except Exception as db_error:
                print(f"âŒ Database insert error: {db_error}")
                return jsonify({'success': False, 'message': f'Database error: {str(db_error)}'}), 500
        
        # Also update vendor status
        vendor_update_query = """
        UPDATE vendors 
        SET nda_status = 'completed', signed_date = %s, updated_at = NOW()
        WHERE id = %s
        """
        
        try:
            execute_query(vendor_update_query, (submission_timestamp, vendor_id))
            print(f"âœ… Vendor status updated for {company_name}")
        except Exception as db_error:
            print(f"âš ï¸ Vendor status update error: {db_error}")
        
        return jsonify({
            'success': True, 
            'message': 'NDA form submission received and processed successfully'
        })
        
    except Exception as e:
        print(f"âŒ Google Form webhook error: {e}")
        return jsonify({'success': False, 'message': 'Failed to process form submission'}), 500

def get_smtp_settings():
    """Fetch current SMTP settings from database"""
    try:
        query = """
        SELECT setting_key, setting_value
        FROM system_settings
        WHERE setting_category = 'email' AND setting_key IN ('smtp_server', 'smtp_port', 'smtp_username', 'smtp_password', 'from_name')
        """
        settings = execute_query(query)
        smtp_config = {}
        for setting in settings:
            smtp_config[setting['setting_key']] = setting['setting_value'].strip() if setting['setting_value'] else setting['setting_value']
        
        # Return with defaults if settings not found
        return {
            'smtp_server': smtp_config.get('smtp_server', SMTP_SERVER),
            'smtp_port': int(smtp_config.get('smtp_port', SMTP_PORT)),
            'smtp_username': smtp_config.get('smtp_username', SMTP_USERNAME),
            'smtp_password': smtp_config.get('smtp_password', SMTP_PASSWORD),
            'from_name': smtp_config.get('from_name', 'YellowStone XPs')
        }
    except Exception as e:
        print(f"Error fetching SMTP settings: {e}")
        return {
            'smtp_server': SMTP_SERVER,
            'smtp_port': SMTP_PORT,
            'smtp_username': SMTP_USERNAME,
            'smtp_password': SMTP_PASSWORD,
            'from_name': 'YellowStone XPs'
        }

def add_anti_spam_headers(msg, smtp_config):
    """Add anti-spam headers to email message to prevent emails from going to spam"""
    try:
        msg['List-Unsubscribe'] = '<https://yellowstonexperiences.com/unsubscribe>'
        msg['List-Unsubscribe-Post'] = 'List-Unsubscribe=One-Click'
        msg['X-Mailer'] = 'YellowStone XPs Management System'
        msg['X-Priority'] = '3'
        msg['X-MSMail-Priority'] = 'Normal'
        msg['Importance'] = 'Normal'
        # Add Message-ID only if not set (avoid duplicates/conflicts)
        if 'Message-ID' not in msg:
            import uuid
            msg['Message-ID'] = f'<{uuid.uuid4()}@yellowstonexperiences.com>'
        msg['Return-Path'] = smtp_config.get('smtp_username', SMTP_USERNAME)
        msg['Reply-To'] = smtp_config.get('smtp_username', SMTP_USERNAME)
        # Add organization info
        msg['Organization'] = 'YellowStone Xperiences Pvt Ltd'
        msg['X-Complaints-To'] = smtp_config.get('smtp_username', SMTP_USERNAME)
    except Exception as e:
        print(f"Error adding anti-spam headers: {e}")

@app.route('/api/admin/send-nda', methods=['POST'])
def send_nda():
    """Send NDA to vendor with proper error handling"""
    try:
        data = request.get_json()
        email = data.get('email')
        company_name = data.get('company_name')
        
        print(f"NDA Send Request: email={email}, company={company_name}")
        
        if not email:
            print("âŒ Email is required")
            return jsonify({'success': False, 'error': 'Email is required'}), 400
        
        reference_number = generate_reference_number()
        print(f"Generated reference number: {reference_number}")
        
        check_query = "SELECT id FROM vendors WHERE email = %s"
        existing_vendor = execute_query(check_query, (email,), fetch_one=True)
        
        if existing_vendor:
            print(f"Updating existing vendor: {existing_vendor['id']}")
            query = """
            UPDATE vendors SET company_name = %s, reference_number = %s, updated_at = %s
            WHERE email = %s
            """
            result = execute_query(query, (company_name, reference_number, datetime.now(), email))
            print(f"âœ… Vendor update query executed successfully")
        else:
            print(f"Creating new vendor")
            query = """
            INSERT INTO vendors (email, company_name, nda_status, reference_number, created_at)
            VALUES (%s, %s, %s, %s, %s)
            """
            result = execute_query(query, (email, company_name, 'sent', reference_number, datetime.now()))
            print(f"âœ… Vendor insert query executed successfully")
        
        print(f"âœ… Vendor record saved successfully")
        
        try:
            # Fetch current SMTP settings from database
            smtp = get_smtp_settings()
            
            msg = MIMEMultipart()
            msg['From'] = f'YellowStone XPs <{smtp["smtp_username"]}>'
            msg['To'] = email
            msg['Subject'] = f"NDA Agreement Request - {company_name} (Ref: {reference_number})"
            
            # Add anti-spam headers
            add_anti_spam_headers(msg, smtp)
            
            body = f"""
Dear Sir/Madam,

Subject: Invitation to Complete Non-Disclosure Agreement - Strategic Partnership Opportunity

We hope this correspondence finds you in good health and prosperity.

YellowStone Xperiences Pvt Ltd is pleased to extend an invitation to {company_name} to participate in our Strategic Partnership Program.

REFERENCE NUMBER: {reference_number}

To facilitate this process, we have prepared a comprehensive NDA form that can be completed electronically through our secure portal:

NDA FORM LINK: http://localhost:8000/nda-form?ref={reference_number}

IMPORTANT INSTRUCTIONS:
- Please complete all required fields accurately
- Ensure all company information is current and verified
- Digital signature is required for document execution
- Company stamp upload is optional but recommended

Should you have any questions or require assistance, please contact our Vendor Relations Department.

We look forward to establishing a mutually beneficial partnership.

Yours faithfully,

Harpreet Singh
CEO
YellowStone Xperiences Pvt Ltd
Email: Harpreet.singh@yellowstonexps.com
Address: Plot # 2, ITC, Fourth Floor, Sector 67, Mohali -160062, Punjab, India
            """
            
            msg.attach(MIMEText(body, 'plain'))
            
            print(f"Attempting to send email to {email}")
            print(f"SMTP Config: Server={smtp['smtp_server']}, Port={smtp['smtp_port']}, From={smtp['smtp_username']}")
            
            # Use port 587 or 465 based on configuration
            if smtp['smtp_port'] == 465:
                print("Using SSL connection for port 465...")
                server = smtplib.SMTP_SSL(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
            else:
                print(f"Using STARTTLS for port {smtp['smtp_port']}...")
                server = smtplib.SMTP(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
                server.starttls()
            
            print("Logging in...")
            server.login(smtp['smtp_username'], smtp['smtp_password'])
            print("Sending email...")
            
            text = msg.as_string()
            server.sendmail(smtp['smtp_username'], email, text)
            server.quit()
            
            print(f"âœ… NDA email sent successfully to {email} with reference {reference_number}")
            
        except Exception as email_error:
            print(f"âŒ Email sending failed: {email_error}")
            return jsonify({'success': True, 'message': 'NDA recorded but email delivery failed', 'reference_number': reference_number, 'email_error': str(email_error)})
        
        return jsonify({'success': True, 'message': 'NDA sent successfully', 'reference_number': reference_number})
    except Exception as e:
        print(f"âŒ Send NDA error: {e}")
        return jsonify({'success': False, 'error': f'Failed to send NDA: {str(e)}'}), 500

# Static file serving
@app.route('/static/css/<path:filename>')
def serve_css(filename):
    """Serve CSS files from the React build"""
    return send_from_directory('frontend/build/static/css', filename)

@app.route('/static/js/<path:filename>')
def serve_js(filename):
    """Serve JS files from the React build"""
    return send_from_directory('frontend/build/static/js', filename)

@app.route('/static/<path:filename>')
def serve_static(filename):
    """Serve other static files from the React build"""
    return send_from_directory('frontend/build/static', filename)

# React App Routes
@app.route('/nda-form')
def serve_nda_form():
    """Serve the custom NDA form"""
    return send_file('frontend/build/nda-form.html')

@app.route('/nda-signature')
def serve_nda_signature():
    """Serve the NDA signature page"""
    return send_file('frontend/build/nda-signature.html')

@app.route('/vendor-portal')
def serve_vendor_portal():
    """Serve the vendor portal React application"""
    return send_file('frontend/build/index.html')

@app.route('/')
def serve_react_app():
    """Serve the React application"""
    return send_file('frontend/build/index.html')

@app.route('/app')
def serve_react_app_alt():
    """Alternative route to serve the React application"""
    return send_file('frontend/build/index.html')

@app.route('/api/admin/broadcast-change', methods=['POST'])
def manual_broadcast_change():
    """Manually trigger a database change broadcast (for testing)"""
    try:
        data = request.get_json()
        
        table_name = data.get('table')
        action = data.get('action')
        change_data = data.get('data')
        
        if not all([table_name, action]):
            return jsonify({'success': False, 'error': 'Missing required fields: table, action'}), 400
        
        # Broadcast the change
        broadcast_database_change(table_name, action, change_data)
        
        return jsonify({
            'success': True,
            'message': f'Broadcasted {action} on {table_name}',
            'data': change_data
        })
        
    except Exception as e:
        print(f"Manual broadcast error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# WebSocket event handlers for real-time updates
@socketio.on('connect')
def handle_connect():
    """Handle client connection"""
    print(f"Client connected: {request.sid}")
    emit('connected', {'message': 'Connected to real-time updates'})

@socketio.on('disconnect')
def handle_disconnect():
    """Handle client disconnection"""
    print(f"Client disconnected: {request.sid}")

@socketio.on('join_room')
def handle_join_room(data):
    """Handle client joining a specific room for targeted updates"""
    room = data.get('room', 'general')
    join_room(room)
    print(f"Client {request.sid} joined room: {room}")
    emit('joined_room', {'room': room})

@socketio.on('leave_room')
def handle_leave_room(data):
    """Handle client leaving a specific room"""
    room = data.get('room', 'general')
    leave_room(room)
    print(f"Client {request.sid} left room: {room}")
    emit('left_room', {'room': room})

def send_vendor_credentials_email(vendor_email, vendor_password, company_name):
    """Send vendor login credentials via email"""
    try:
        smtp = get_smtp_settings()
        msg = MIMEMultipart()
        msg['From'] = f'YellowStone XPs <{smtp["smtp_username"]}>'
        msg['To'] = vendor_email
        msg['Subject'] = "Your Vendor Portal Access - YellowStone XPs"
        
        # Add anti-spam headers
        add_anti_spam_headers(msg, smtp)
        
        body = f"""Dear Vendor,

Congratulations! Your registration for {company_name} has been approved.

You now have full access to the YellowStone XPs Vendor Portal.

LOGIN CREDENTIALS:

Username: {vendor_email}
Password: {vendor_password}

Portal URL: http://localhost:3000/vendor-portal

IMPORTANT SECURITY NOTES:
- Please change your password after your first login
- Keep your credentials secure
- Contact support if you have any issues
- Do not share your login credentials with anyone

If you need assistance, please contact our Vendor Relations Department.

Best regards,
YellowStone Xperiences Pvt Ltd
Harpreet Singh - CEO
Email: Harpreet.singh@yellowstonexps.com
Address: Plot # 2, ITC, Fourth Floor, Sector 67, Mohali -160062, Punjab, India
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        if smtp['smtp_port'] == 465:
            server = smtplib.SMTP_SSL(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
        else:
            server = smtplib.SMTP(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
            server.starttls()
        
        server.login(smtp['smtp_username'], smtp['smtp_password'])
        text = msg.as_string()
        server.sendmail(smtp['smtp_username'], vendor_email, text)
        server.quit()
        
        print(f"âœ… Credentials email sent to {vendor_email}")
        return True
        
    except Exception as e:
        print(f"âŒ Failed to send credentials email: {e}")
        return False

def broadcast_database_change(table_name, action, data=None, room='admin'):
    """Broadcast database changes to connected clients"""
    try:
        message = {
            'table': table_name,
            'action': action,  # 'insert', 'update', 'delete'
            'data': data,
            'timestamp': datetime.now().isoformat()
        }
        
        # Broadcast to specific room (admin dashboard)
        socketio.emit('database_change', message, room=room)
        
        # Also broadcast to general room for any other listeners
        socketio.emit('database_change', message, room='general')
        
        print(f"Broadcasted database change: {table_name} - {action} - {data}")
        
    except Exception as e:
        print(f"Error broadcasting database change: {e}")

# Frontend API endpoints (without /admin prefix)
@app.route('/api/tasks', methods=['GET'])
def get_tasks_frontend():
    """Get all tasks for frontend"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500
        
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM tasks ORDER BY created_at DESC")
        tasks = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'tasks': tasks
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/employee/assigned-tenders', methods=['GET'])
def get_employee_assigned_tenders():
    """Return tenders assigned to an employee (coordinator or in team)."""
    try:
        employee_id = request.args.get('employee_id', type=int)
        if not employee_id:
            return jsonify({'success': False, 'error': 'Employee ID required'}), 400

        query = """
            SELECT t.*
            FROM created_tenders t
            WHERE t.project_coordinator_id = %s
               OR (t.project_team_ids IS NOT NULL AND JSON_CONTAINS(t.project_team_ids, JSON_ARRAY(%s)))
        """
        try:
            tenders = execute_query(query, (employee_id, employee_id), fetch_all=True) or []
        except Exception:
            like_value = f"%{employee_id}%"
            tenders = execute_query(
                "SELECT * FROM created_tenders WHERE project_coordinator_id = %s OR project_team_ids LIKE %s",
                (employee_id, like_value),
                fetch_all=True
            ) or []

        return jsonify({'success': True, 'tenders': tenders})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/projects', methods=['GET'])
def get_projects_frontend():
    """Get all projects for frontend"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500
        
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM projects ORDER BY created_at DESC")
        projects = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'projects': projects
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/tickets', methods=['GET'])
def get_tickets_frontend():
    """Get all tickets for frontend"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500
        
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM tickets ORDER BY created_at DESC")
        tickets = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'tickets': tickets
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/project-updates', methods=['GET'])
def get_project_updates():
    """Get all project updates"""
    try:
        return jsonify([])
    except Exception as e:
        return jsonify({'error': 'Failed to get project updates'}), 500

@app.route('/api/admin/task-updates', methods=['GET'])
def get_task_updates():
    """Get all task updates"""
    try:
        return jsonify([])
    except Exception as e:
        return jsonify({'error': 'Failed to get task updates'}), 500

@app.route('/api/admin/ticket-updates', methods=['GET'])
def get_ticket_updates():
    """Get ticket updates.
    If 'ticket_id' query param is provided, returns updates for that ticket.
    Otherwise returns empty (UI does not need the full list).
    """
    try:
        ticket_id = request.args.get('ticket_id', type=int)
        if not ticket_id:
            return jsonify([])

        execute_query("""
            CREATE TABLE IF NOT EXISTS ticket_updates (
                id INT AUTO_INCREMENT PRIMARY KEY,
                ticket_id INT NOT NULL,
                employee_id INT NULL,
                status VARCHAR(50) NULL,
                update_message TEXT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE,
                FOREIGN KEY (employee_id) REFERENCES users(id) ON DELETE SET NULL
            )
        """)

        query = """
            SELECT tu.id, tu.status, tu.update_message, tu.created_at,
                   u.name AS employee_name
            FROM ticket_updates tu
            LEFT JOIN users u ON tu.employee_id = u.id
            WHERE tu.ticket_id = %s
            ORDER BY tu.created_at DESC, tu.id DESC
        """
        raw_updates = execute_query(query, (ticket_id,), fetch_all=True) or []

        # Normalize timestamps to ISO with UTC marker so browsers convert to local correctly
        updates = []
        for u in raw_updates:
            created_at = u.get('created_at')
            if created_at is not None:
                try:
                    # If naive datetime, treat as UTC and append 'Z'
                    iso_value = created_at.isoformat()
                    if 'Z' not in iso_value and '+' not in iso_value:
                        iso_value = iso_value + 'Z'
                    u['created_at'] = iso_value
                except Exception:
                    # Fallback to string
                    u['created_at'] = str(created_at)
            updates.append(u)

        return jsonify({'success': True, 'updates': updates})
    except Exception:
        return jsonify({'error': 'Failed to get ticket updates'}), 500

@app.route('/api/admin/tickets/<int:ticket_id>/updates', methods=['GET'])
def get_ticket_updates_for_ticket(ticket_id):
    """Get updates log for a specific ticket"""
    try:
        # Ensure table exists, then fetch updates
        execute_query("""
            CREATE TABLE IF NOT EXISTS ticket_updates (
                id INT AUTO_INCREMENT PRIMARY KEY,
                ticket_id INT NOT NULL,
                employee_id INT NULL,
                status VARCHAR(50) NULL,
                update_message TEXT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE,
                FOREIGN KEY (employee_id) REFERENCES users(id) ON DELETE SET NULL
            )
        """)

        query = """
            SELECT tu.id, tu.status, tu.update_message, tu.created_at,
                   u.name AS employee_name
            FROM ticket_updates tu
            LEFT JOIN users u ON tu.employee_id = u.id
            WHERE tu.ticket_id = %s
            ORDER BY tu.created_at DESC, tu.id DESC
        """
        updates = execute_query(query, (ticket_id,), fetch_all=True) or []
        return jsonify({'success': True, 'updates': updates})
    except Exception as e:
        print(f"Error fetching ticket updates for ticket {ticket_id}: {e}")
        return jsonify({'success': False, 'error': 'Failed to get updates'}), 500

# Employee-specific API endpoints
@app.route('/api/employee/notifications', methods=['GET'])
def get_employee_notifications():
    """Get notifications for a specific employee"""
    try:
        employee_id = request.args.get('employee_id')
        if not employee_id:
            return jsonify({'success': False, 'error': 'Employee ID required'}), 400
        
        # Return empty notifications for now
        return jsonify({
            'success': True,
            'notifications': []
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employee/announcements', methods=['GET'])
def get_employee_announcements():
    """Get announcements for a specific employee"""
    try:
        employee_id = request.args.get('employee_id')
        if not employee_id:
            return jsonify({'success': False, 'error': 'Employee ID required'}), 400
        
        # Return empty announcements for now
        return jsonify({
            'success': True,
            'announcements': []
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/api/debug/check-tickets-table', methods=['GET'])
def check_tickets_table():
    """Debug endpoint to check tickets table structure"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Database connection failed'}), 500
        
        cursor = connection.cursor()
        cursor.execute("DESCRIBE tickets")
        table_structure = cursor.fetchall()
        
        cursor.execute("SELECT * FROM tickets LIMIT 1")
        sample_data = cursor.fetchone()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'table_structure': table_structure,
            'sample_data': sample_data
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/debug/employee-data', methods=['GET'])
def debug_employee_data():
    """Debug endpoint to check employee data"""
    try:
        employee_id = request.args.get('employee_id')
        if not employee_id:
            return jsonify({'error': 'Employee ID required'}), 400
        
        # Get all tickets
        tickets_query = "SELECT * FROM tickets"
        all_tickets = execute_query(tickets_query, fetch_all=True)
        
        # Get all projects  
        projects_query = "SELECT * FROM projects"
        all_projects = execute_query(projects_query, fetch_all=True)
        
        # Get all tasks
        tasks_query = "SELECT * FROM tasks"
        all_tasks = execute_query(tasks_query, fetch_all=True)
        
        return jsonify({
            'employee_id': employee_id,
            'all_tickets': all_tickets,
            'all_projects': all_projects,
            'all_tasks': all_tasks,
            'tickets_for_employee': [t for t in all_tickets if str(t.get('assigned_to_id')) == str(employee_id) or str(t.get('created_by')) == str(employee_id)],
            'projects_for_employee': [p for p in all_projects if str(p.get('assigned_to_id')) == str(employee_id)],
            'tasks_for_employee': [t for t in all_tasks if str(t.get('assigned_to_id')) == str(employee_id)]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/fix-employee-data', methods=['POST'])
def fix_employee_data():
    """Fix employee data by assigning existing data to the correct employee"""
    try:
        data = request.get_json()
        employee_id = data.get('employee_id')
        if not employee_id:
            return jsonify({'error': 'Employee ID required'}), 400
        
        # Update existing tickets to be assigned to this employee
        tickets_query = "UPDATE tickets SET assigned_to_id = %s WHERE assigned_to_id = 17"
        execute_query(tickets_query, (employee_id,))
        
        # Update existing projects to be assigned to this employee
        projects_query = "UPDATE projects SET assigned_to_id = %s WHERE assigned_to_id = 17"
        execute_query(projects_query, (employee_id,))
        
        # Update existing tasks to be assigned to this employee
        tasks_query = "UPDATE tasks SET assigned_to_id = %s WHERE assigned_to_id = 17"
        execute_query(tasks_query, (employee_id,))
        
        return jsonify({
            'success': True,
            'message': f'Updated all data to be assigned to employee {employee_id}'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/employee/create-ticket', methods=['POST'])
def create_employee_ticket():
    """Create a ticket for an employee"""
    try:
        data = request.get_json()
        employee_id = data.get('employee_id')
        title = data.get('title', 'Untitled Ticket')
        description = data.get('description', '')
        
        # Map frontend priority values to database values (based on actual database values)
        priority_mapping = {
            'urgent': 'Critical',
            'high': 'Critical', 
            'medium': 'Medium',
            'low': 'Low'
        }
        priority = priority_mapping.get(data.get('priority', '').lower(), 'Medium')
        
        # Map frontend category values to database values
        category_mapping = {
            'it support': 'technical',
            'technical': 'technical',
            'general': 'general',
            'bug': 'bug',
            'feature': 'feature'
        }
        category = category_mapping.get(data.get('category', '').lower(), 'general')
        
        if not employee_id:
            return jsonify({'success': False, 'error': 'Employee ID required'}), 400
        
        # Get employee details
        employee_query = "SELECT name, email FROM users WHERE id = %s"
        employee = execute_query(employee_query, (employee_id,), fetch_one=True)
        
        if not employee:
            return jsonify({'success': False, 'error': 'Employee not found'}), 404
        
        # Insert ticket into database
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500
        
        cursor = connection.cursor()
        ticket_query = """
        INSERT INTO tickets (title, description, priority, status, category, assigned_to, created_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        
        # Get employee name for created_by field
        employee_name = employee['name']
        
        cursor.execute(ticket_query, (
            title,
            description,
            priority,
            'open',
            category,
            category,  # assigned_to will temporarily contain the category
            employee_name  # created_by will contain the employee's name
        ))
        
        ticket_id = cursor.lastrowid
        connection.commit()
        cursor.close()
        connection.close()
        
        if ticket_id:
            print(f"Ticket created successfully: ID {ticket_id} by employee {employee_id}")
            
            # Broadcast the change to admin dashboard
            broadcast_database_change('tickets', 'insert', {
                'id': ticket_id,
                'title': title,
                'priority': priority,
                'status': 'open',
                'assigned_to_name': employee['name'],
                'assigned_to_email': employee['email']
            })
            
            return jsonify({
                'success': True,
                'message': 'Ticket created successfully',
                'ticket_id': ticket_id
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to create ticket'}), 500
            
    except Exception as e:
        print(f"Error creating employee ticket: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employee/projects/<int:project_id>/update', methods=['POST', 'PUT'])
def update_employee_project(project_id):
    """Update project status and comments for employee"""
    try:
        data = request.get_json()
        status = data.get('status')
        comments = data.get('comments', '')
        
        # Map status to database format
        status_map = {
            'planning': 'planning',
            'in_progress': 'in_progress',
            'on_hold': 'on_hold',
            'completed': 'completed',
            'cancelled': 'cancelled'
        }
        
        db_status = status_map.get(status, status)  # Use original if not in map
        
        # Update project status and comments
        update_query = """
        UPDATE projects 
        SET status = %s, updated_at = NOW()
        WHERE id = %s
        """
        
        execute_query(update_query, (db_status, project_id))
        
        # Broadcast the change
        broadcast_database_change('projects', 'update', {
            'id': project_id,
            'status': db_status,
            'comments': comments
        })
        
        return jsonify({
            'success': True,
            'message': 'Project updated successfully'
        })
        
    except Exception as e:
        print(f"âŒ Error updating employee project: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employee/tasks/<int:task_id>/update', methods=['POST', 'PUT'])
def update_employee_task(task_id):
    """Update task status and comments for employee"""
    try:
        data = request.get_json()
        status = data.get('status')
        comments = data.get('comments', '')
        
        # Map status to database format
        status_map = {
            'pending': 'pending',
            'in_progress': 'in_progress',
            'completed': 'completed',
            'cancelled': 'cancelled'
        }
        
        db_status = status_map.get(status, status)  # Use original if not in map
        
        # Ensure task_updates table exists
        try:
            execute_query("""
                CREATE TABLE IF NOT EXISTS task_updates (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    task_id INT NOT NULL,
                    employee_id INT NULL,
                    status VARCHAR(50) NULL,
                    update_message TEXT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE CASCADE,
                    FOREIGN KEY (employee_id) REFERENCES users(id) ON DELETE SET NULL
                )
            """)
        except Exception:
            pass

        # Update task status and comments
        update_query = """
        UPDATE tasks 
        SET status = %s, updated_at = NOW()
        WHERE id = %s
        """
        
        execute_query(update_query, (db_status, task_id))
        
        # Insert progress log
        try:
            employee_id = (request.get_json() or {}).get('employee_id')
            insert_update = """
                INSERT INTO task_updates (task_id, employee_id, status, update_message)
                VALUES (%s, %s, %s, %s)
            """
            execute_query(insert_update, (task_id, employee_id, db_status, comments))
        except Exception:
            pass

        # Broadcast the change
        broadcast_database_change('tasks', 'update', {
            'id': task_id,
            'status': db_status,
            'comments': comments
        })
        
        return jsonify({
            'success': True,
            'message': 'Task updated successfully'
        })
        
    except Exception as e:
        print(f"âŒ Error updating employee task: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employee/tickets/<int:ticket_id>/update', methods=['POST', 'PUT'])
def update_employee_ticket(ticket_id):
    """Update ticket status and comments for employee"""
    try:
        data = request.get_json()
        status = data.get('status')
        comments = data.get('comments', '')
        
        print(f"📝 Updating ticket {ticket_id} with status: {status}, comments: {comments}")
        
        # Map status to database format (same as admin endpoint)
        status_map = {
            'open': 'Open',
            'in_progress': 'In Progress', 
            'resolved': 'Resolved',
            'closed': 'Closed'
        }
        
        db_status = status_map.get(status, status)  # Use original if not in map
        print(f"📝 Mapped status '{status}' to '{db_status}'")
        
        # Ensure ticket_updates table exists for logging
        try:
            execute_query("""
                CREATE TABLE IF NOT EXISTS ticket_updates (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    ticket_id INT NOT NULL,
                    employee_id INT NULL,
                    status VARCHAR(50) NULL,
                    update_message TEXT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE,
                    FOREIGN KEY (employee_id) REFERENCES users(id) ON DELETE SET NULL
                )
            """)
        except Exception as _tbl_err:
            # If table creation fails, proceed without blocking the status update
            pass

        # Update ticket status and comments
        update_query = """
        UPDATE tickets 
        SET status = %s, updated_at = NOW()
        WHERE id = %s
        """
        
        execute_query(update_query, (db_status, ticket_id))
        
        # Insert progress log
        try:
            employee_id = (request.get_json() or {}).get('employee_id')
            insert_update = """
                INSERT INTO ticket_updates (ticket_id, employee_id, status, update_message)
                VALUES (%s, %s, %s, %s)
            """
            execute_query(insert_update, (ticket_id, employee_id, db_status, comments))
        except Exception as _ins_err:
            # Non-fatal
            pass
        print(f"✅ Ticket {ticket_id} updated successfully in database")
        
        # Broadcast the change
        broadcast_database_change('tickets', 'update', {
            'id': ticket_id,
            'status': db_status,
            'comments': comments
        })
        
        return jsonify({
            'success': True,
            'message': 'Ticket updated successfully'
        })
        
    except Exception as e:
        print(f"âŒ Error updating employee ticket: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/tasks/<int:task_id>/updates', methods=['GET'])
def get_task_updates_for_task(task_id):
    """Get updates log for a specific task"""
    try:
        execute_query("""
            CREATE TABLE IF NOT EXISTS task_updates (
                id INT AUTO_INCREMENT PRIMARY KEY,
                task_id INT NOT NULL,
                employee_id INT NULL,
                status VARCHAR(50) NULL,
                update_message TEXT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE CASCADE,
                FOREIGN KEY (employee_id) REFERENCES users(id) ON DELETE SET NULL
            )
        """)

        query = """
            SELECT tu.id, tu.status, tu.update_message, tu.created_at,
                   u.name AS employee_name
            FROM task_updates tu
            LEFT JOIN users u ON tu.employee_id = u.id
            WHERE tu.task_id = %s
            ORDER BY tu.created_at DESC, tu.id DESC
        """
        raw_updates = execute_query(query, (task_id,), fetch_all=True) or []

        updates = []
        for u in raw_updates:
            created_at = u.get('created_at')
            if created_at is not None:
                try:
                    iso_value = created_at.isoformat()
                    if 'Z' not in iso_value and '+' not in iso_value:
                        iso_value = iso_value + 'Z'
                    u['created_at'] = iso_value
                except Exception:
                    u['created_at'] = str(created_at)
            updates.append(u)

        return jsonify({'success': True, 'updates': updates})
    except Exception as e:
        print(f"Error fetching task updates for task {task_id}: {e}")
        return jsonify({'success': False, 'error': 'Failed to get updates'}), 500

@app.route('/api/admin/task-updates', methods=['GET'])
def get_task_updates_query():
    """Query param variant: /api/admin/task-updates?task_id=123"""
    try:
        task_id = request.args.get('task_id', type=int)
        if not task_id:
            return jsonify([])
        return get_task_updates_for_task(task_id)
    except Exception as e:
        return jsonify({'success': False, 'error': 'Failed to get updates'}), 500

@app.route('/api/employee/upload-profile-picture', methods=['POST'])
def upload_profile_picture():
    """Upload profile picture for the current user"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if file and file.filename.lower().endswith(tuple(ALLOWED_IMAGE_EXTENSIONS)):
            # Generate unique filename
            filename = f"{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
            file_path = os.path.join(PROFILE_PHOTOS_FOLDER, filename)
            
            # Save the file
            file.save(file_path)
            
            # Update database with new profile picture path
            update_query = "UPDATE users SET profile_picture = %s WHERE id = %s"
            execute_query(update_query, (filename, user_id))
            
            return jsonify({
                'success': True,
                'message': 'Profile picture uploaded successfully',
                'filename': filename
            })
        else:
            return jsonify({'error': 'Invalid file type. Only PNG, JPG, JPEG, and GIF are allowed.'}), 400
            
    except Exception as e:
        print(f"âŒ Upload profile picture error: {e}")
        return jsonify({'error': 'Failed to upload profile picture'}), 500


@app.route('/api/employee/upload-lead-report', methods=['POST'])
def upload_lead_report():
    """Upload lead report file and process it"""
    try:
        # print(f"Upload request received. Files: {list(request.files.keys())}")
        # print(f"Form data: {list(request.form.keys())}")
        
        # Relax auth: accept employee_id from args/form if session missing
        user_id = session.get('user_id') or request.args.get('employee_id') or request.form.get('employee_id')
        try:
            user_id = int(user_id)
        except Exception:
            user_id = None
        if not user_id:
            # final fallback: derive user from email in file if available (skipped for now)
            user_id = 0
        if not user_id:
            user_id = 0
        
        if 'file' not in request.files:
            # print("No 'file' key in request.files")
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        # print(f"File details: name='{file.filename}', content_type='{file.content_type}', size={len(file.read())}")
        file.seek(0)  # Reset file pointer after reading
        
        if file.filename == '':
            # print("Empty filename")
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file extension
        allowed_extensions = {'.csv', '.xlsx', '.xls'}
        file_ext = '.' + file.filename.split('.')[-1].lower() if '.' in file.filename else ''
        print(f"File extension: '{file_ext}'")
        
        if file_ext not in allowed_extensions:
            print(f"Invalid file extension: {file_ext}")
            return jsonify({'error': f'Invalid file type. Only CSV, XLS, and XLSX files are allowed. Got: {file_ext}'}), 400
        
        # Use a single timestamp for this batch so reports can group these rows together
        batch_uploaded_at = datetime.now()

        # Read the file based on extension
        try:
            print(f"Attempting to read file with pandas...")
            import pandas as pd
            print(f"Pandas imported successfully")
            
            if file_ext == '.csv':
                print(f"Reading CSV file...")
                df = pd.read_csv(file)
            else:
                print(f"Reading Excel file...")
                df = pd.read_excel(file)
            
            print(f"File read successfully. Shape: {df.shape}")
            print(f"Columns: {list(df.columns)}")
            
        except ImportError as e:
            print(f"Pandas import error: {e}")
            return jsonify({'error': 'Pandas library not available. Please install: pip install pandas openpyxl'}), 400
        except Exception as e:
            print(f"File reading error: {e}")
            return jsonify({'error': f'Error reading file: {str(e)}'}), 400
        
        # Expected columns
        expected_columns = [
            'Company Name', 'Project Name', 'Key Account Manager', 
            'Project Coordinator', 'Client End Manager', 'Client Email',
            'Location', 'Start Date', 'Expected Project Start Date',
            'Last Interacted Date', 'Lead Status', 'Lead Source', 'Remarks'
        ]
        
        # Check if all required columns are present
        # print(f"Checking columns...")
        # print(f"Expected: {expected_columns}")
        # print(f"Found: {list(df.columns)}")
        
        missing_columns = [col for col in expected_columns if col not in df.columns]
        if missing_columns:
            # print(f"Missing columns: {missing_columns}")
            return jsonify({
                'error': f'Missing required columns: {", ".join(missing_columns)}',
                'expected_columns': expected_columns,
                'found_columns': list(df.columns),
                'missing_columns': missing_columns
            }), 400
        
        # print(f"All required columns present")
        
        # Process each row
        processed_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Prepare data for insertion
                lead_data = {
                    'company_name': str(row['Company Name']).strip() if pd.notna(row['Company Name']) else '',
                    'project_name': str(row['Project Name']).strip() if pd.notna(row['Project Name']) else '',
                    'key_account_manager': str(row['Key Account Manager']).strip() if pd.notna(row['Key Account Manager']) else '',
                    'project_coordinator': str(row['Project Coordinator']).strip() if pd.notna(row['Project Coordinator']) else '',
                    'client_end_manager': str(row['Client End Manager']).strip() if pd.notna(row['Client End Manager']) else '',
                    'client_email': str(row['Client Email']).strip() if pd.notna(row['Client Email']) else '',
                    'location': str(row['Location']).strip() if pd.notna(row['Location']) else '',
                    'start_date': str(row['Start Date']).strip() if pd.notna(row['Start Date']) else '',
                    'expected_project_start_date': str(row['Expected Project Start Date']).strip() if pd.notna(row['Expected Project Start Date']) else '',
                    'last_interacted_date': str(row['Last Interacted Date']).strip() if pd.notna(row['Last Interacted Date']) else '',
                    'lead_status': str(row['Lead Status']).strip() if pd.notna(row['Lead Status']) else '',
                    'lead_source': str(row['Lead Source']).strip() if pd.notna(row['Lead Source']) else '',
                    'remarks': str(row['Remarks']).strip() if pd.notna(row['Remarks']) else '',
                    'uploaded_by': user_id,
                    'uploaded_at': batch_uploaded_at
                }
                
                # Insert into database
                insert_query = """
                INSERT INTO lead_generation_reports 
                (company_name, project_name, key_account_manager, project_coordinator, 
                 client_end_manager, client_email, location, start_date, expected_project_start_date,
                 last_interacted_date, lead_status, lead_source, remarks, uploaded_by, uploaded_at)
                VALUES (%(company_name)s, %(project_name)s, %(key_account_manager)s, %(project_coordinator)s,
                        %(client_end_manager)s, %(client_email)s, %(location)s, %(start_date)s, %(expected_project_start_date)s,
                        %(last_interacted_date)s, %(lead_status)s, %(lead_source)s, %(remarks)s, %(uploaded_by)s, %(uploaded_at)s)
                """
                
                print(f"Inserting row {index + 1}: {lead_data['company_name']}")
                execute_query(insert_query, lead_data)
                processed_count += 1
                print(f"Row {index + 1} inserted successfully")
                
            except Exception as e:
                print(f"Error inserting row {index + 1}: {str(e)}")
                errors.append(f"Row {index + 1}: {str(e)}")
        
        print(f"Upload completed. Processed: {processed_count}/{len(df)} rows")
        
        # Count total leads from this file by the batch timestamp to show in reports list
        file_count = execute_query("SELECT COUNT(*) as cnt FROM lead_generation_reports WHERE uploaded_by = %s AND uploaded_at = %s", (user_id, batch_uploaded_at), fetch_one=True)
        count_leads = file_count['cnt'] if file_count else processed_count
        
        return jsonify({
            'success': True,
            'message': f'Successfully processed {processed_count} leads',
            'processed_count': processed_count,
            'total_rows': len(df),
            'total_leads': count_leads,
            'uploaded_at': batch_uploaded_at,
            'filename': file.filename,
            'errors': errors[:10] if errors else []  # Limit errors to first 10
        })
        
    except Exception as e:
        print(f"Upload lead report error: {e}")
        return jsonify({'error': 'Failed to upload lead report'}), 500

@app.route('/api/employee/uploaded-reports', methods=['GET', 'OPTIONS'])
def get_uploaded_reports():
    """Get all uploaded reports for the current user"""
    # print(f"Uploaded reports request: {request.method} - {request.url}")
    
    if request.method == 'OPTIONS':
        # print("Handling OPTIONS request")
        return '', 200
    
    try:
        # Get current user ID
        current_user_id = session.get('user_id')
        # print(f"Current user_id: {current_user_id}")
        
        # Get query parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        search = request.args.get('search', '', type=str)
        
        # print(f"Loading uploaded reports for user {current_user_id}, search: '{search}', page: {page}")
        
        # Build query to get uploaded reports with lead counts
        # Query groups by uploaded_at and filename to create report entries
        base_query = """
        SELECT 
            DATE(lr.uploaded_at) as uploaded_date,
            lr.uploaded_at,
            u.name as uploaded_by_name,
            COUNT(*) as total_leads,
            MIN(lr.uploaded_at) as min_uploaded_at,
            COUNT(*) OVER() as total_count
        FROM lead_generation_reports lr
        LEFT JOIN users u ON lr.uploaded_by = u.id
        WHERE 1=1
        """
        
        params = []
        
        # Add search filter
        if search:
            search_filter = """
            AND (ur.report_name LIKE %s OR ur.original_filename LIKE %s OR ur.description LIKE %s)
            """
            base_query += search_filter
            search_param = f"%{search}%"
            params.extend([search_param, search_param, search_param])
        
        # Filter to current employee if provided
        if current_user_id:
            base_query += " AND lr.uploaded_by = %s"
            params.append(current_user_id)

        # Add grouping and ordering
        base_query += """
        GROUP BY lr.uploaded_at, u.name
        ORDER BY lr.uploaded_at DESC LIMIT %s OFFSET %s
        """
        offset = (page - 1) * per_page
        params.extend([per_page, offset])
        
        # print(f"Query: {base_query}")
        # print(f"Params: {params}")
        
        # Execute query
        try:
            reports = execute_query(base_query, params, fetch_all=True)
            # print(f"Query executed successfully, got {len(reports) if reports else 0} results")
            
            # Extract total count from first row if available
            total = reports[0]['total_count'] if reports and len(reports) > 0 else 0
            
        except Exception as e:
            # print(f"Error executing query: {e}")
            return jsonify({'error': f'Database error: {str(e)}'}), 500
        
        response_data = {
            'success': True,
            'reports': reports or [],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        }
        
        # print(f"Returning response with {len(reports or [])} reports")
        return jsonify(response_data)
        
    except Exception as e:
        # print(f"Get uploaded reports error: {e}")
        return jsonify({'error': 'Failed to fetch uploaded reports'}), 500

@app.route('/api/employee/uploaded-reports/<path:report_id>', methods=['DELETE', 'OPTIONS'])
def delete_uploaded_report(report_id):
    """Delete an uploaded report batch by its uploaded_at value for the current employee."""
    if request.method == 'OPTIONS':
        return '', 200
    try:
        current_user_id = session.get('user_id') or request.args.get('employee_id')
        try:
            current_user_id = int(current_user_id)
        except Exception:
            current_user_id = None
        if not current_user_id:
            return jsonify({'error': 'Not authenticated'}), 401

        key_raw = unquote(report_id)

        # Primary delete with raw value
        execute_query("DELETE FROM lead_generation_reports WHERE uploaded_by = %s AND uploaded_at = %s", (current_user_id, key_raw))

        # If frontend sent a formatted label instead of exact timestamp, attempt to parse common formats
        for fmt in ("%a, %d %b %Y %H:%M:%S GMT", "%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                dt_obj = datetime.strptime(key_raw, fmt)
                formatted = dt_obj.strftime("%Y-%m-%d %H:%M:%S")
                execute_query("DELETE FROM lead_generation_reports WHERE uploaded_by = %s AND uploaded_at = %s", (current_user_id, formatted))
                break
            except Exception:
                continue

        return jsonify({'success': True})
    except Exception as e:
        print(f"Error deleting uploaded report: {e}")
        return jsonify({'error': 'Failed to delete report'}), 500

@app.route('/api/employee/uploaded-reports', methods=['DELETE', 'POST', 'OPTIONS'])
def delete_uploaded_report_fallback():
    """Fallback deletion endpoint that accepts id via query/body to avoid path encoding issues."""
    if request.method == 'OPTIONS':
        return '', 200
    try:
        current_user_id = session.get('user_id') or request.args.get('employee_id') or (request.get_json(silent=True) or {}).get('employee_id')
        try:
            current_user_id = int(current_user_id)
        except Exception:
            current_user_id = None
        if not current_user_id:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.get_json(silent=True) or {}
        key_raw = request.args.get('id') or request.args.get('uploaded_at') or data.get('id') or data.get('uploaded_at')
        if not key_raw:
            return jsonify({'error': 'Missing report id'}), 400
        key_raw = unquote(str(key_raw))

        execute_query("DELETE FROM lead_generation_reports WHERE uploaded_by = %s AND uploaded_at = %s", (current_user_id, key_raw))

        for fmt in ("%a, %d %b %Y %H:%M:%S GMT", "%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                dt_obj = datetime.strptime(key_raw, fmt)
                formatted = dt_obj.strftime("%Y-%m-%d %H:%M:%S")
                execute_query("DELETE FROM lead_generation_reports WHERE uploaded_by = %s AND uploaded_at = %s", (current_user_id, formatted))
                break
            except Exception:
                continue

        return jsonify({'success': True})
    except Exception as e:
        print(f"Error deleting uploaded report (fallback): {e}")
        return jsonify({'error': 'Failed to delete report'}), 500

@app.route('/api/employee/lead-reports', methods=['GET', 'OPTIONS'])
def get_lead_reports():
    """Get all lead reports for the current user"""
    # print(f"Lead reports request: {request.method} - {request.url}")
    
    if request.method == 'OPTIONS':
        # print("Handling OPTIONS request")
        return '', 200
    
    try:
        # Get current user ID
        current_user_id = session.get('user_id')
        # print(f"Current user_id: {current_user_id}")
        
        # Get query parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        search = request.args.get('search', '', type=str)
        
        # print(f"Loading reports for user {current_user_id}, search: '{search}', page: {page}")
        
        # Build query - show all reports if current user has no reports, otherwise show user's reports
        base_query = """
        SELECT lr.*, u.name as uploaded_by_name,
               COUNT(*) OVER() as total_count
        FROM lead_generation_reports lr
        LEFT JOIN users u ON lr.uploaded_by = u.id
        WHERE 1=1
        """
        
        params = []
        
        # Add search filter
        if search:
            search_filter = """
            AND (lr.company_name LIKE %s OR lr.project_name LIKE %s OR 
                 lr.key_account_manager LIKE %s OR lr.client_email LIKE %s)
            """
            base_query += search_filter
            search_param = f"%{search}%"
            params.extend([search_param, search_param, search_param, search_param])
        
        # Add ordering and pagination
        base_query += " ORDER BY lr.uploaded_at DESC LIMIT %s OFFSET %s"
        offset = (page - 1) * per_page
        params.extend([per_page, offset])
        
        # print(f"Single query: {base_query}")
        # print(f"Params: {params}")
        
        # Execute single query
        try:
            reports = execute_query(base_query, params, fetch_all=True)
            # print(f"Query executed successfully, got {len(reports) if reports else 0} results")
            
            # Extract total count from first row if available
            total = reports[0]['total_count'] if reports and len(reports) > 0 else 0
            
        except Exception as e:
            # print(f"Error executing query: {e}")
            return jsonify({'error': f'Database error: {str(e)}'}), 500
        
        response_data = {
            'success': True,
            'reports': reports or [],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        }
        
        # print(f"Returning response with {len(reports or [])} reports")
        return jsonify(response_data)
        
    except Exception as e:
        # print(f"Get lead reports error: {e}")
        return jsonify({'error': 'Failed to fetch lead reports'}), 500

@app.route('/api/test-table-exists', methods=['GET'])
def test_table_exists():
    """Test if lead_generation_reports table exists"""
    try:
        test_query = "SELECT COUNT(*) as count FROM lead_generation_reports LIMIT 1"
        result = execute_query(test_query, fetch_one=True)
        if result:
            return jsonify({
                'success': True, 
                'table_exists': True, 
                'count': result['count']
            })
        else:
            return jsonify({
                'success': True, 
                'table_exists': False
            })
    except Exception as e:
        return jsonify({
            'success': False, 
            'error': str(e),
            'table_exists': False
        })

@app.route('/api/employee/remove-profile-picture', methods=['POST'])
def remove_profile_picture():
    """Remove profile picture for the current user"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        # Get current user's profile picture path
        query = "SELECT profile_picture FROM users WHERE id = %s"
        user = execute_query(query, (user_id,), fetch_one=True)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        profile_picture_path = user.get('profile_picture')
        
        # Remove profile picture from database
        update_query = "UPDATE users SET profile_picture = NULL WHERE id = %s"
        execute_query(update_query, (user_id,))
        
        # Delete the actual file if it exists
        if profile_picture_path:
            try:
                file_path = os.path.join(PROFILE_PHOTOS_FOLDER, profile_picture_path)
                if os.path.exists(file_path):
                    os.remove(file_path)
                    print(f"âœ… Profile picture file deleted: {file_path}")
            except Exception as e:
                print(f"âš ï¸ Could not delete profile picture file: {e}")
        
        return jsonify({
            'success': True,
            'message': 'Profile picture removed successfully'
        })
        
    except Exception as e:
        print(f"âŒ Remove profile picture error: {e}")
        return jsonify({'error': 'Failed to remove profile picture'}), 500

@app.route('/api/employee/profile-picture/<filename>')
def get_profile_picture(filename):
    """Serve profile picture files"""
    try:
        return send_from_directory(PROFILE_PHOTOS_FOLDER, filename)
    except Exception as e:
        # print(f"âŒ Error serving profile picture: {e}")
        return jsonify({'error': 'Profile picture not found'}), 404

# ==================== EMPLOYEE LEAD MANAGEMENT ENDPOINTS ====================

@app.route('/api/employee/assigned-leads', methods=['GET'])
def get_employee_assigned_leads():
    """Get leads assigned to the current employee"""
    try:
        employee_id = request.args.get('employee_id')
        if not employee_id:
            return jsonify({'success': False, 'error': 'Employee ID required'}), 400
        
        # print(f"🔍 Debug assigned leads - employee_id: {employee_id}")
        
        # Get employee's assigned leads with lead details
        query = """
        SELECT
            la.id as lead_assignment_id,
            la.employee_id,
            la.status,
            la.progress_notes,
            la.created_at,
            la.assigned_at,
            la.due_date,
            la.notes as assignment_notes,
            lgr.id as lead_id,
            lgr.company_name,
            lgr.project_name,
            lgr.client_email,
            lgr.location,
            lgr.lead_status,
            lgr.priority,
            lgr.estimated_value,
            lgr.industry,
            u.name as assigned_by_name
        FROM lead_assignments la
        JOIN lead_generation_reports lgr ON la.lead_id = lgr.id
        LEFT JOIN users u ON la.assigned_by = u.id
        WHERE la.employee_id = %s
        ORDER BY la.assigned_at DESC
        """
        
        leads = execute_query(query, (employee_id,))
        
        # print(f"🔍 Debug: Employee {employee_id} assigned leads: {leads}")
        if leads:
            for lead in leads:
                # print(f"🔍 Debug: Lead {lead.get('lead_id')} assigned_by_name: {lead.get('assigned_by_name')}")
                pass
        
        return jsonify({'success': True, 'leads': leads})
        
    except Exception as e:
        print(f"âŒ Error getting employee assigned leads: {e}")
        return jsonify({'error': 'Failed to get assigned leads'}), 500

@app.route('/api/employee/leads/<int:assignment_id>/progress', methods=['POST'])
def update_lead_progress(assignment_id):
    """Update lead progress by employee"""
    try:
        employee_id = request.args.get('employee_id')
        if not employee_id:
            return jsonify({'success': False, 'error': 'Employee ID required'}), 400
        
        data = request.get_json()
        status = data.get('status')
        progress_notes = data.get('progress_notes')
        
        # Verify the assignment belongs to this employee
        assignment_check = execute_query(
            "SELECT id, lead_id FROM lead_assignments WHERE id = %s AND employee_id = %s",
            (assignment_id, employee_id),
            fetch_one=True
        )
        
        if not assignment_check:
            return jsonify({'error': 'Assignment not found or not authorized'}), 404
        
        # Update the assignment
        update_query = """
        UPDATE lead_assignments 
        SET status = %s, 
            progress_notes = %s,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = %s AND employee_id = %s
        """
        
        execute_query(update_query, (status, progress_notes, assignment_id, employee_id))
        
        # Also insert into lead_progress_updates for tracking history
        insert_progress_query = """
        INSERT INTO lead_progress_updates (lead_assignment_id, employee_id, status, progress_notes)
        VALUES (%s, %s, %s, %s)
        """
        execute_query(insert_progress_query, (assignment_id, employee_id, status, progress_notes))
        
        # Keep the master lead status in sync with assignment status so admin table reflects it
        try:
            lead_id = assignment_check.get('lead_id')
            if lead_id:
                execute_query(
                    "UPDATE lead_generation_reports SET lead_status = %s, updated_at = CURRENT_TIMESTAMP WHERE id = %s",
                    (status, lead_id)
                )
        except Exception as e:
            print(f"Warning: failed to sync lead status for lead_id {assignment_check.get('lead_id')}: {e}")
        
        # Broadcast realtime updates for admin dashboards and lists
        try:
            broadcast_database_change('lead_assignments', 'update', {
                'assignment_id': assignment_id,
                'employee_id': employee_id,
                'status': status
            })
            if assignment_check.get('lead_id'):
                broadcast_database_change('leads', 'update', {
                    'lead_id': assignment_check.get('lead_id'),
                    'status': status
                })
        except Exception as e:
            print(f"Warning: failed to broadcast lead changes: {e}")
        
        return jsonify({'success': True, 'message': 'Lead progress updated successfully'})
        
    except Exception as e:
        print(f"âŒ Error updating lead progress: {e}")
        return jsonify({'error': 'Failed to update lead progress'}), 500

@app.route('/api/admin/leads/<int:lead_id>/progress-updates', methods=['GET'])
def get_lead_progress_updates(lead_id):
    """Get progress updates for a specific lead"""
    try:
        query = """
        SELECT 
            lpu.*,
            u.name as employee_name,
            u.email as employee_email,
            la.lead_id
        FROM lead_progress_updates lpu
        JOIN lead_assignments la ON lpu.lead_assignment_id = la.id
        JOIN users u ON lpu.employee_id = u.id
        WHERE la.lead_id = %s
        ORDER BY lpu.created_at DESC
        """
        
        updates = execute_query(query, (lead_id,), fetch_all=True)
        return jsonify({'success': True, 'updates': updates or []})
        
    except Exception as e:
        print(f"Error getting progress updates: {e}")
        return jsonify({'success': False, 'updates': []})

# ==================== LEAD MANAGEMENT ENDPOINTS ====================

@app.route('/api/admin/leads', methods=['GET'])
def get_admin_leads():
    """Get all leads for admin dashboard"""
    try:
        admin_id = session.get('user_id') or request.args.get('admin_id')
        user_type = session.get('user_type') or 'admin'
        if not admin_id:
            admin_id = 1  # Use default admin ID
        
        # Get all leads with LATEST assignment information (by assigned_at) 
        query = """
        SELECT
            lgr.*,
            la.id as assignment_id,
            la.employee_id as assigned_to,
            la.status as assignment_status,
            la.due_date as assignment_due_date,
            la.notes as assignment_notes,
            u.name as assigned_employee_name,
            u.designation as assigned_employee_designation,
            assigner.name as assigned_by_name
        FROM lead_generation_reports lgr
        LEFT JOIN (
            SELECT la1.*
            FROM lead_assignments la1
            INNER JOIN (
                SELECT lead_id, MAX(assigned_at) as latest_assignment_time
                FROM lead_assignments
                GROUP BY lead_id
            ) la2 ON la1.lead_id = la2.lead_id AND la1.assigned_at = la2.latest_assignment_time
        ) la ON lgr.id = la.lead_id
        LEFT JOIN users u ON la.employee_id = u.id
        LEFT JOIN users assigner ON la.assigned_by = assigner.id
        ORDER BY lgr.created_at DESC
        """
        
        leads = execute_query(query)
        
        # Debug: Check what assignments exist for lead 199
        debug_query = "SELECT * FROM lead_assignments WHERE lead_id = 199 ORDER BY id DESC"
        debug_assignments = execute_query(debug_query)
        # print(f"ðŸ” Debug: All assignments for lead 199: {debug_assignments}")
        
        # print(f"ðŸ” Debug: Raw leads from DB: {leads}")
        
        # Convert to list format for frontend (query already handles latest assignment)
        result_leads = []
        for lead in leads:
            result_leads.append({
                'id': lead['id'],
                'company_name': lead['company_name'],
                'project_name': lead['project_name'],
                'key_account_manager': lead['key_account_manager'],
                'project_coordinator': lead['project_coordinator'],
                'client_end_manager': lead['client_end_manager'],
                'client_email': lead['client_email'],
                'location': lead['location'],
                'start_date': lead['start_date'],
                'expected_project_start_date': lead['expected_project_start_date'],
                'last_interacted_date': lead['last_interacted_date'],
                'lead_status': lead['lead_status'],
                'lead_source': lead['lead_source'],
                'remarks': lead['remarks'],
                'uploaded_by': lead['uploaded_by'],
                'uploaded_at': lead['uploaded_at'],
                'created_at': lead['created_at'],
                'updated_at': lead['updated_at'],
                'assigned_to': lead['assigned_to'],
                'assignment_status': lead['assignment_status'],
                'assignment_due_date': lead['assignment_due_date'],
                'assignment_notes': lead['assignment_notes'],
                'assigned_employee_name': lead['assigned_employee_name'],
                'assigned_employee_designation': lead['assigned_employee_designation']
            })
        
        return jsonify(result_leads)
        
    except Exception as e:
        print(f"âŒ Error getting leads: {e}")
        return jsonify({'error': 'Failed to get leads'}), 500

@app.route('/api/admin/leads', methods=['POST'])
def create_lead():
    """Create a new lead"""
    try:
        admin_id = session.get('user_id') or request.args.get('admin_id')
        user_type = session.get('user_type') or 'admin'
        if not admin_id:
            admin_id = 1  # Use default admin ID
        
        data = request.get_json()
        
        # Check if there are any users in the database to use as uploaded_by
        user_check = execute_query("SELECT id FROM users LIMIT 1", fetch_one=True)
        uploaded_by_id = user_check['id'] if user_check else None
        
        # Insert new lead
        query = """
        INSERT INTO lead_generation_reports 
        (company_name, project_name, key_account_manager, project_coordinator, 
         client_end_manager, client_email, client_phone, location, industry,
         priority, estimated_value, start_date, 
         expected_project_start_date, last_interacted_date, lead_status, 
         lead_source, remarks, uploaded_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        values = (
            data.get('company_name'),
            data.get('project_name'),
            data.get('key_account_manager'),
            data.get('project_coordinator'),
            data.get('client_end_manager'),
            data.get('client_email'),
            data.get('client_phone'),
            data.get('location'),
            data.get('industry'),
            data.get('priority'),
            data.get('estimated_value'),
            data.get('start_date'),
            data.get('expected_project_start_date'),
            data.get('last_interacted_date'),
            data.get('lead_status', 'new'),
            data.get('lead_source'),
            data.get('remarks'),
            uploaded_by_id  # Use first available user ID or None
        )
        
        execute_query(query, values)
        
        # Get the ID of the created lead
        lead_id = execute_query("SELECT LAST_INSERT_ID() as id", fetch_one=True)['id']
        
        return jsonify({'success': True, 'message': 'Lead created successfully', 'lead_id': lead_id})
        
    except Exception as e:
        print(f"âŒ Error creating lead: {e}")
        return jsonify({'error': 'Failed to create lead'}), 500

@app.route('/api/admin/leads/<int:lead_id>/assign', methods=['POST'])
def assign_lead(lead_id):
        # print(f"🔍 Debug: assign_lead function called for lead_id: {lead_id}")
    """Assign a lead to an employee"""
    try:
        # Relaxed auth: accept session or query/body admin_id, else fallback to any user
        admin_id = session.get('user_id') or request.args.get('admin_id') or (request.get_json(silent=True) or {}).get('admin_id')
        
        data = request.get_json(silent=True) or {}
        employee_id = data.get('employee_id')
        due_date = data.get('due_date')
        notes = data.get('notes')
        
        if not employee_id:
            return jsonify({'error': 'Employee ID is required'}), 400
        
        # Check if lead exists and get company name
        lead_check = execute_query("SELECT id, company_name FROM lead_generation_reports WHERE id = %s", (lead_id,))
        if not lead_check:
            return jsonify({'error': 'Lead not found'}), 404
        
        lead_data = lead_check[0]
        
        # Check if employee exists by employee_id field (accept numeric/trimmed)
        employee_check = execute_query("SELECT id, name FROM users WHERE employee_id = %s", (str(employee_id).strip(),))
        if not employee_check:
            return jsonify({'error': f'Employee with ID {employee_id} not found'}), 404
        
        employee = employee_check[0]
        actual_employee_id = employee['id']  # Get the database ID
        
        # Determine assigned_by id robustly
        assigned_by_id = None
        if admin_id:
            admin_user_check = execute_query("SELECT id FROM users WHERE id = %s", (admin_id,), fetch_one=True)
            if admin_user_check:
                assigned_by_id = admin_id
        if not assigned_by_id:
            any_user_check = execute_query("SELECT id FROM users LIMIT 1", fetch_one=True)
            assigned_by_id = (any_user_check['id'] if any_user_check else actual_employee_id)
        
        # Handle empty or UI-formatted due_date (convert DD-MM-YYYY -> YYYY-MM-DD)
        if not due_date or str(due_date).strip() == '':
            due_date = None
        else:
            try:
                ds = str(due_date)
                if '-' in ds:
                    parts = ds.split('-')
                    # If DD-MM-YYYY, flip
                    if len(parts[0]) == 2 and len(parts[2]) == 4:
                        due_date = f"{parts[2]}-{parts[1]}-{parts[0]}"
                # else leave as is (assuming YYYY-MM-DD)
            except Exception:
                due_date = None
        
        # Insert assignment
        assignment_query = """
        INSERT INTO lead_assignments (lead_id, employee_id, assigned_by, due_date, notes)
        VALUES (%s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
        assigned_by = VALUES(assigned_by),
        due_date = VALUES(due_date),
        notes = VALUES(notes),
        status = 'assigned',
        updated_at = CURRENT_TIMESTAMP
        """
        
        execute_query(assignment_query, (lead_id, actual_employee_id, assigned_by_id, due_date, notes))

        # Ensure master lead status reflects assignment
        try:
            execute_query(
                "UPDATE lead_generation_reports SET lead_status = 'assigned', updated_at = CURRENT_TIMESTAMP WHERE id = %s",
                (lead_id,)
            )
        except Exception as se:
            print(f"Warning: failed to sync lead status on assignment: {se}")

        # Broadcast change so admin/employee UIs refresh
        try:
            broadcast_database_change('lead_assignments', 'insert', {
                'lead_id': lead_id,
                'employee_id': actual_employee_id,
                'assigned_by': assigned_by_id,
                'due_date': due_date
            })
            broadcast_database_change('leads', 'update', {
                'lead_id': lead_id,
                'status': 'assigned'
            })
        except Exception as be:
            print(f"Warning: failed to broadcast assignment: {be}")
        
        # print(f"ðŸ” Debug: Lead {lead_id} assigned to employee {actual_employee_id}")
        # print(f"ðŸ” Debug: Assignment query executed successfully")
        
        return jsonify({'success': True, 'message': 'Lead assigned successfully'})
        
    except Exception as e:
        print(f"âŒ Error assigning lead: {e}")
        return jsonify({'error': 'Failed to assign lead'}), 500

# Open variant without admin prefix to avoid strict admin guards in some environments
@app.route('/api/leads/<int:lead_id>/assign', methods=['POST', 'OPTIONS'])
def assign_lead_open(lead_id):
    """Open assign endpoint used as a fallback by the frontend. Mirrors assign_lead logic."""
    if request.method == 'OPTIONS':
        response = jsonify({})
        response.headers.add('Access-Control-Allow-Origin', 'http://localhost:3000')
        response.headers.add('Access-Control-Allow-Methods', 'POST')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
        return response
    try:
        admin_id = session.get('user_id') or request.args.get('admin_id') or (request.get_json(silent=True) or {}).get('admin_id')
        print(f"🔍 Open assign endpoint called for lead_id: {lead_id}, admin_id: {admin_id}")
        data = request.get_json(silent=True) or {}
        employee_id = data.get('employee_id')
        due_date = data.get('due_date')
        notes = data.get('notes')

        if not employee_id:
            return jsonify({'error': 'Employee ID is required'}), 400

        lead_check = execute_query("SELECT id, company_name FROM lead_generation_reports WHERE id = %s", (lead_id,))
        if not lead_check:
            return jsonify({'error': 'Lead not found'}), 404

        employee_check = execute_query("SELECT id, name FROM users WHERE employee_id = %s", (str(employee_id).strip(),))
        if not employee_check:
            return jsonify({'error': f'Employee with ID {employee_id} not found'}), 404
        employee = employee_check[0]
        actual_employee_id = employee['id']

        assigned_by_id = None
        if admin_id:
            admin_user_check = execute_query("SELECT id FROM users WHERE id = %s", (admin_id,), fetch_one=True)
            if admin_user_check:
                assigned_by_id = admin_id
        if not assigned_by_id:
            any_user_check = execute_query("SELECT id FROM users LIMIT 1", fetch_one=True)
            assigned_by_id = (any_user_check['id'] if any_user_check else actual_employee_id)

        if not due_date or str(due_date).strip() == '':
            due_date = None
        else:
            try:
                ds = str(due_date)
                if '-' in ds:
                    parts = ds.split('-')
                    if len(parts[0]) == 2 and len(parts[2]) == 4:
                        due_date = f"{parts[2]}-{parts[1]}-{parts[0]}"
            except Exception:
                due_date = None

        assignment_query = """
        INSERT INTO lead_assignments (lead_id, employee_id, assigned_by, due_date, notes)
        VALUES (%s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
        assigned_by = VALUES(assigned_by),
        due_date = VALUES(due_date),
        notes = VALUES(notes),
        status = 'assigned',
        updated_at = CURRENT_TIMESTAMP
        """
        execute_query(assignment_query, (lead_id, actual_employee_id, assigned_by_id, due_date, notes))

        try:
            execute_query(
                "UPDATE lead_generation_reports SET lead_status = 'assigned', updated_at = CURRENT_TIMESTAMP WHERE id = %s",
                (lead_id,)
            )
        except Exception as se:
            print(f"Warning: failed to sync lead status on assignment (open): {se}")

        try:
            broadcast_database_change('lead_assignments', 'insert', {
                'lead_id': lead_id,
                'employee_id': actual_employee_id,
                'assigned_by': assigned_by_id,
                'due_date': due_date
            })
            broadcast_database_change('leads', 'update', {
                'lead_id': lead_id,
                'status': 'assigned'
            })
        except Exception as be:
            print(f"Warning: failed to broadcast assignment (open): {be}")

        return jsonify({'success': True, 'message': 'Lead assigned successfully'})
    except Exception as e:
        print(f"❌ Error assigning lead (open endpoint): {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to assign lead'}), 500

@app.route('/api/admin/leads/<int:lead_id>/status', methods=['PUT'])
def update_lead_status(lead_id):
    """Update lead status"""
    try:
        admin_id = session.get('user_id')
        if not admin_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        # Check if user is admin
        admin_check = execute_query("SELECT role FROM admins WHERE id = %s", (admin_id,), fetch_one=True)
        if not admin_check or admin_check['role'] != 'admin':
            return jsonify({'error': 'Unauthorized'}), 403
        
        data = request.get_json()
        new_status = data.get('status')
        
        if not new_status:
            return jsonify({'error': 'Status is required'}), 400
        
        # Update lead status
        query = "UPDATE lead_generation_reports SET lead_status = %s, updated_at = CURRENT_TIMESTAMP WHERE id = %s"
        execute_query(query, (new_status, lead_id))
        
        return jsonify({'success': True, 'message': 'Lead status updated successfully'})
        
    except Exception as e:
        print(f"âŒ Error updating lead status: {e}")
        return jsonify({'error': 'Failed to update lead status'}), 500

@app.route('/api/admin/leads/<int:lead_id>', methods=['DELETE'])
def delete_lead(lead_id):
    """Delete a lead"""
    try:
        # Delete lead (cascade will handle assignments)
        query = "DELETE FROM lead_generation_reports WHERE id = %s"
        execute_query(query, (lead_id,))
        
        return jsonify({'success': True, 'message': 'Lead deleted successfully'})
        
    except Exception as e:
        # print(f"âŒ Error deleting lead: {e}")
        return jsonify({'error': 'Failed to delete lead'}), 500


@app.route('/api/admin/employees', methods=['GET'])
def get_admin_employees():
    """Get all employees for admin dashboard"""
    try:
        admin_id = session.get('user_id') or request.args.get('admin_id')
        user_type = session.get('user_type') or 'admin'
        if not admin_id:
            admin_id = 1  # Use default admin ID
        
        # Get all employees with department info
        query = """
        SELECT 
            u.id, u.employee_id, u.name, u.email, 
            COALESCE(u.designation, '') as position,
            COALESCE(u.department, '') as department,
            COALESCE(u.manager, '') as manager,
            COALESCE(u.phone, '') as phone,
            u.created_at as hireDate,
            COALESCE(SUBSTRING(u.name, 1, 1), '') as avatar,
            u.status
        FROM users u
        WHERE u.user_type = 'employee'
        ORDER BY u.name
        """
        
        employees = execute_query(query)
        return jsonify(employees)
        
    except Exception as e:
        print(f"âŒ Error getting employees: {e}")
        return jsonify({'error': 'Failed to get employees'}), 500

@app.route('/api/admin/employees/search', methods=['GET'])
def search_admin_employees():
    """Lightweight employee search by name (for assigns)."""
    try:
        q = (request.args.get('query') or '').strip()
        limit = int(request.args.get('limit', 20))
        if not q:
            return jsonify([])
        like = f"%{q}%"
        query = (
            "SELECT id, name, email FROM users "
            "WHERE user_type = 'employee' AND name LIKE %s ORDER BY name LIMIT %s"
        )
        results = execute_query(query, (like, limit), fetch_all=True) or []
        return jsonify(results)
    except Exception as e:
        print(f"Employee search error: {e}")
        return jsonify([])

@app.route('/api/admin/employees/<employee_id>/details', methods=['GET'])
def get_admin_employee_details_by_id(employee_id):
    """Get employee details by employee_id for assignment verification"""
    try:
        # Relaxed auth: accept admin_id via session, query or JSON; default to 1 (dev)
        admin_id = session.get('user_id') or request.args.get('admin_id')
        if not admin_id:
            data = request.get_json(silent=True) or {}
            admin_id = data.get('admin_id')
        if not admin_id:
            admin_id = 1
        
        # Get employee details with manager info
        query = """
        SELECT 
            u.id, u.employee_id, u.name, u.email, u.designation,
            u.phone, u.created_at
        FROM users u
        WHERE u.employee_id = %s
        """
        
        employee = execute_query(query, (employee_id,), fetch_one=True)
        
        if not employee:
            return jsonify({'error': f'Employee with ID {employee_id} not found'}), 404
        
        return jsonify(employee)
        
    except Exception as e:
        print(f"âŒ Error getting employee details: {e}")
        return jsonify({'error': 'Failed to get employee details'}), 500


# Company Management Endpoints
@app.route('/api/admin/companies', methods=['GET'])
def get_companies():
    """Get all companies for admin"""
    try:
        admin_id = session.get('user_id') or request.args.get('admin_id')
        user_type = session.get('user_type') or 'admin'
        if not admin_id:
            admin_id = 1  # Use default admin ID
        
        query = """
        SELECT c.id, c.company_name, c.email, c.contact_person, c.phone, c.industry, c.website, c.status, c.email_type,
               c.created_at, c.updated_at,
               GROUP_CONCAT(DISTINCT se.email_type ORDER BY se.sent_at) as sent_email_types,
               GROUP_CONCAT(DISTINCT se.status ORDER BY se.sent_at) as email_statuses
        FROM companies c
        LEFT JOIN scheduled_emails se ON c.id = se.company_id AND se.status = 'sent'
        GROUP BY c.id
        ORDER BY c.created_at DESC
        """
        
        companies = execute_query(query)
        return jsonify({'success': True, 'companies': companies})
        
    except Exception as e:
        print(f"❌ Error getting companies: {e}")
        return jsonify({'error': 'Failed to get companies'}), 500


@app.route('/api/admin/companies', methods=['POST'])
def create_company():
    """Create a new company"""
    try:
        # Relaxed auth: accept admin_id from session, query, or body; default to 1 (dev)
        admin_id = session.get('user_id') or request.args.get('admin_id')
        if not admin_id:
            data = request.get_json(silent=True) or {}
            admin_id = data.get('admin_id')
        if not admin_id:
            admin_id = 1
        
        data = request.get_json()
        company_name = data.get('company_name')
        email = data.get('email')
        contact_person = data.get('contact_person', '')
        phone = data.get('phone', '')
        industry = data.get('industry', '')
        website = data.get('website', '')
        email_type = data.get('email_type', 'intro')
        
        if not company_name or not email:
            return jsonify({'error': 'Company name and email are required'}), 400
        
        query = """
        INSERT INTO companies (company_name, email, contact_person, phone, industry, website, email_type)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        
        execute_query(query, (company_name, email, contact_person, phone, industry, website, email_type))
        
        return jsonify({'success': True, 'message': 'Company created successfully'})
        
    except Exception as e:
        print(f"❌ Error creating company: {e}")
        return jsonify({'error': 'Failed to create company'}), 500


@app.route('/api/admin/companies/<int:company_id>', methods=['DELETE'])
def delete_company(company_id):
    """Delete a company"""
    try:
        # Relaxed auth: accept admin_id from session, query, or body; default to 1 (dev)
        admin_id = session.get('user_id') or request.args.get('admin_id')
        if not admin_id:
            data = request.get_json(silent=True) or {}
            admin_id = data.get('admin_id')
        if not admin_id:
            admin_id = 1
        
        query = "DELETE FROM companies WHERE id = %s"
        execute_query(query, (company_id,))
        
        return jsonify({'success': True, 'message': 'Company deleted successfully'})
        
    except Exception as e:
        # print(f"❌ Error deleting company: {e}")
        return jsonify({'error': 'Failed to delete company'}), 500


# Email Scheduling Endpoints
@app.route('/api/admin/scheduled-emails', methods=['GET'])
def get_scheduled_emails():
    """Get all scheduled emails"""
    try:
        admin_id = session.get('user_id') or request.args.get('admin_id')
        if not admin_id:
            admin_id = 1  # Use default admin ID
        
        # print(f"🔍 Debug: Getting scheduled emails for admin_id: {admin_id}")
        
        query = """
        SELECT se.id, se.company_id, c.company_name, c.email, c.contact_person,
               se.subject, se.email_body, se.scheduled_time, se.status, 
               se.sent_at, se.error_message, se.created_at,
               u.name as created_by_name
        FROM scheduled_emails se
        JOIN companies c ON se.company_id = c.id
        LEFT JOIN users u ON se.created_by = u.id
        ORDER BY se.scheduled_time DESC
        """
        
        emails = execute_query(query)
        # print(f"🔍 Debug: Found {len(emails) if emails else 0} scheduled emails")
        
        # Format datetime objects for JSON serialization
        if emails:
            for email in emails:
                if email['scheduled_time']:
                    email['scheduled_time'] = email['scheduled_time'].strftime('%Y-%m-%d %H:%M:%S')
                if email.get('sent_at'):
                    email['sent_at'] = email['sent_at'].strftime('%Y-%m-%d %H:%M:%S')
                if email.get('created_at'):
                    email['created_at'] = email['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        
        # Debug: Print first email to see the format
        # if emails:
        #     print(f"DEBUG: First email scheduled_time: {emails[0]['scheduled_time']}")
        #     print(f"DEBUG: First email sent_at: {emails[0].get('sent_at')}")
        
        return jsonify({'success': True, 'emails': emails})
        
    except Exception as e:
        # print(f"❌ Error getting scheduled emails: {e}")
        return jsonify({'error': 'Failed to get scheduled emails'}), 500


@app.route('/api/admin/scheduled-emails', methods=['POST'])
def schedule_email():
    """Schedule an email to be sent"""
    try:
        # Relaxed auth: accept admin_id via session, form, query, or JSON; default to 1
        admin_id = session.get('user_id') or request.form.get('admin_id') or request.args.get('admin_id')
        if not admin_id:
            data = request.get_json(silent=True) or {}
            admin_id = data.get('admin_id')
        if not admin_id:
            admin_id = 1
        
        data = request.get_json(silent=True) or {}
        company_id = data.get('company_id')
        scheduled_time = data.get('scheduled_time')
        email_type = data.get('email_type', 'intro')  # Get email_type from request, default to 'intro'
        
        if not company_id or not scheduled_time:
            return jsonify({'error': 'Company ID and scheduled time are required'}), 400
        
        # Get company details
        company_query = "SELECT company_name, email, contact_person, industry, email_type FROM companies WHERE id = %s"
        company = execute_query(company_query, (company_id,), fetch_one=True)
        
        if not company:
            return jsonify({'error': 'Company not found'}), 404
        
        # Check if admin_id exists in users table, if not use any existing user
        user_check_query = "SELECT id FROM users WHERE id = %s"
        user_exists = execute_query(user_check_query, (admin_id,), fetch_one=True)
        
        if not user_exists:
            # If admin_id doesn't exist, find any user to use as created_by
            any_user_query = "SELECT id FROM users LIMIT 1"
            any_user = execute_query(any_user_query, fetch_one=True)
            created_by_id = any_user['id'] if any_user else None
        else:
            created_by_id = admin_id
        
        if not created_by_id:
            return jsonify({'error': 'No valid user found for email scheduling'}), 400
        
        # Generate email template based on selected email_type and company industry
        template = get_email_template(
            email_type,  # Use email_type from request instead of company's email_type
            company['industry'], 
            company['company_name'], 
            company['contact_person']
        )
        
        subject = template['subject']
        email_body = template['body']
        
        query = """
        INSERT INTO scheduled_emails (company_id, subject, email_body, scheduled_time, created_by, email_type)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        
        execute_query(query, (company_id, subject, email_body, scheduled_time, created_by_id, email_type))
        
        return jsonify({'success': True, 'message': 'Email scheduled successfully'})
        
    except Exception as e:
        print(f"❌ Error scheduling email: {e}")
        return jsonify({'error': 'Failed to schedule email'}), 500


@app.route('/api/admin/scheduled-emails/<int:email_id>', methods=['DELETE'])
def cancel_scheduled_email(email_id):
    """Cancel a scheduled email"""
    try:
        admin_id = session.get('user_id')
        if not admin_id or session.get('user_type') != 'admin':
            return jsonify({'error': 'Unauthorized'}), 401
        
        query = "UPDATE scheduled_emails SET status = 'cancelled' WHERE id = %s AND status = 'pending'"
        execute_query(query, (email_id,))
        
        return jsonify({'success': True, 'message': 'Email cancelled successfully'})
        
    except Exception as e:
        # print(f"❌ Error cancelling email: {e}")
        return jsonify({'error': 'Failed to cancel email'}), 500


@app.route('/api/admin/scheduled-emails/bulk', methods=['POST'])
def schedule_bulk_emails():
    """Schedule emails for multiple companies with batch processing"""
    try:
        admin_id = session.get('user_id') or request.form.get('admin_id')
        try:
            if request.is_json and request.json:
                admin_id = admin_id or request.json.get('admin_id')
        except:
            pass
        if not admin_id:
            admin_id = 1  # Use default admin ID
        
        # Check if request has files (FormData with attachment)
        attachment_file = request.files.get('attachment')
        if attachment_file and attachment_file.filename:
            # Handle FormData request with attachment
            import json
            company_ids = json.loads(request.form.get('company_ids'))
            scheduled_time = request.form.get('scheduled_time')
            batch_size = int(request.form.get('batch_size', 10))
            batch_interval = int(request.form.get('batch_interval', 10))
            email_type = request.form.get('email_type', 'intro')
            
            # Save the attachment file
            from werkzeug.utils import secure_filename
            import os
            
            filename = secure_filename(attachment_file.filename)
            # Create uploads directory if it doesn't exist
            upload_dir = os.path.join('uploads', 'email_attachments')
            os.makedirs(upload_dir, exist_ok=True)
            
            # Save file
            filepath = os.path.join(upload_dir, filename)
            attachment_file.save(filepath)
            attachment_path = filepath
        else:
            # Handle JSON request (no attachment)
            data = request.get_json()
            company_ids = data.get('company_ids', [])
            scheduled_time = data.get('scheduled_time')
            batch_size = data.get('batch_size', 10)
            batch_interval = data.get('batch_interval', 10)
            email_type = data.get('email_type', 'intro')
            attachment_path = None
        
        if not company_ids or not scheduled_time:
            return jsonify({'error': 'Company IDs and scheduled time are required'}), 400
        
        # print(f"📧 Scheduling bulk emails: {len(company_ids)} companies, batch_size={batch_size}, interval={batch_interval} minutes")
        
        # Check if admin_id exists in users table, if not use any existing user
        user_check_query = "SELECT id FROM users WHERE id = %s"
        user_exists = execute_query(user_check_query, (admin_id,), fetch_one=True)
        
        if not user_exists:
            any_user_query = "SELECT id FROM users LIMIT 1"
            any_user = execute_query(any_user_query, fetch_one=True)
            created_by_id = any_user['id'] if any_user else None
        else:
            created_by_id = admin_id
        
        if not created_by_id:
            return jsonify({'error': 'No valid user found for email scheduling'}), 400
        
        # Get company details for all selected companies
        placeholders = ','.join(['%s'] * len(company_ids))
        companies_query = f"SELECT id, company_name, email, contact_person, industry, email_type FROM companies WHERE id IN ({placeholders})"
        companies = execute_query(companies_query, company_ids)
        
        if not companies:
            return jsonify({'error': 'No companies found'}), 404
        
        # Calculate batch scheduling
        from datetime import datetime, timedelta
        
        # Parse the initial scheduled time and convert to UTC for storage
        initial_time = datetime.fromisoformat(scheduled_time.replace('Z', ''))
        
        # Convert to UTC for consistent storage (assuming input is local time)
        # This ensures the database stores times in UTC
        initial_time = initial_time.replace(tzinfo=None)  # Remove timezone info for MySQL storage
        
        scheduled_count = 0
        batch_number = 1
        
        # Process companies in batches
        for i in range(0, len(companies), batch_size):
            batch_companies = companies[i:i + batch_size]
            
            # Calculate scheduled time for this batch
            if batch_number == 1:
                batch_scheduled_time = initial_time
            else:
                # Add interval minutes for subsequent batches
                batch_scheduled_time = initial_time + timedelta(minutes=(batch_number - 1) * batch_interval)
            
            # print(f"📦 Batch {batch_number}: {len(batch_companies)} companies scheduled for {batch_scheduled_time}")
            
            # Schedule emails for each company in this batch
            for company in batch_companies:
                # Generate email template based on the selected email_type and company's industry
                template = get_email_template(
                    email_type,  # Use the selected email type from the form
                    company['industry'], 
                    company['company_name'], 
                    company['contact_person']
                )
                
                subject = template['subject']
                email_body = template['body']
                
                query = """
                INSERT INTO scheduled_emails (company_id, subject, email_body, scheduled_time, created_by, email_type, attachment_path)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """
                
                execute_query(query, (company['id'], subject, email_body, batch_scheduled_time, created_by_id, email_type, attachment_path))
                scheduled_count += 1
            
            batch_number += 1
        
        total_batches = (len(companies) + batch_size - 1) // batch_size
        total_duration = (total_batches - 1) * batch_interval if total_batches > 1 else 0
        
        return jsonify({
            'success': True, 
            'message': f'{scheduled_count} emails scheduled in {total_batches} batches over {total_duration} minutes',
            'scheduled_count': scheduled_count,
            'total_batches': total_batches,
            'batch_size': batch_size,
            'batch_interval': batch_interval,
            'total_duration_minutes': total_duration
        })
        
    except Exception as e:
        print(f"❌ Error scheduling bulk emails: {e}")
        return jsonify({'error': 'Failed to schedule bulk emails'}), 500


@app.route('/api/admin/scheduled-emails/<int:email_id>/send-now', methods=['POST'])
def send_email_now(email_id):
    """Send a scheduled email immediately"""
    try:
        # Relaxed auth: accept admin_id from session, query or JSON; default to 1 for dev
        admin_id = session.get('user_id') or request.args.get('admin_id')
        if not admin_id:
            data = request.get_json(silent=True) or {}
            admin_id = data.get('admin_id')
        if not admin_id:
            admin_id = 1
        
        # Get the scheduled email details
        email_query = """
        SELECT se.id, se.company_id, se.subject, se.email_body, se.scheduled_time,
               c.company_name, c.email, c.contact_person
        FROM scheduled_emails se
        JOIN companies c ON se.company_id = c.id
        WHERE se.id = %s AND se.status = 'pending'
        """
        
        email_data = execute_query(email_query, (email_id,), fetch_one=True)
        
        if not email_data:
            return jsonify({'error': 'Email not found or already sent'}), 404
        
        print(f"📧 Sending email to {email_data['company_name']} ({email_data['email']})")
        
        # Send actual email using SMTP
        try:
            # Fetch current SMTP settings
            smtp = get_smtp_settings()
            
            msg = MIMEMultipart()
            msg['From'] = f'YellowStone XPs <{smtp["smtp_username"]}>'
            msg['To'] = email_data['email']
            msg['Subject'] = email_data['subject']
            
            # Add anti-spam headers
            add_anti_spam_headers(msg, smtp)
            
            # Add email body
            msg.attach(MIMEText(email_data['email_body'], 'plain'))
            
            # Connect to SMTP server and send email
            print(f"SMTP Config: Server={smtp['smtp_server']}, Port={smtp['smtp_port']}, From={smtp['smtp_username']}")
            
            if smtp['smtp_port'] == 465:
                print("Using SSL connection for port 465...")
                server = smtplib.SMTP_SSL(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
            else:
                print(f"Using STARTTLS for port {smtp['smtp_port']}...")
                server = smtplib.SMTP(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
                server.starttls()
            
            # Enable verbose SMTP session logging
            try:
                server.set_debuglevel(1)
            except Exception:
                pass

            print("Logging in...")
            server.login(smtp['smtp_username'], smtp['smtp_password'])
            print("Sending email...")
            text = msg.as_string()
            # BCC sender to verify delivery
            refused = server.sendmail(smtp['smtp_username'], [email_data['email'], smtp['smtp_username']], text)
            print(f"SMTP sendmail refused map: {refused}")
            server.quit()
            
            print(f"✅ Email sent successfully to {email_data['email']}")
            
            # Mark as sent in database
            update_query = """
            UPDATE scheduled_emails 
            SET status = 'sent', sent_at = CURRENT_TIMESTAMP
            WHERE id = %s
            """
            
            execute_query(update_query, (email_id,))
            
            return jsonify({
                'success': True, 
                'message': f'Email sent successfully to {email_data["company_name"]} ({email_data["email"]})'
            })
            
        except Exception as email_error:
            print(f"❌ SMTP Error: {email_error}")
            
            # Mark as failed in database
            update_query = """
            UPDATE scheduled_emails 
            SET status = 'failed', error_message = %s
            WHERE id = %s
            """
            
            execute_query(update_query, (str(email_error), email_id,))
            
            return jsonify({
                'success': False, 
                'message': f'Failed to send email: {str(email_error)}'
            }), 500
        
    except Exception as e:
        print(f"❌ Error sending email now: {e}")
        return jsonify({'error': 'Failed to send email'}), 500


@app.route('/api/admin/scheduled-emails/send-all-pending', methods=['POST'])
def send_all_pending_emails():
    """Send all pending emails immediately"""
    try:
        # Relaxed auth: accept admin_id from session, query or JSON; default to 1 for dev
        admin_id = session.get('user_id') or request.args.get('admin_id')
        if not admin_id:
            data = request.get_json(silent=True) or {}
            admin_id = data.get('admin_id')
        if not admin_id:
            admin_id = 1
        
        # Get all pending emails
        pending_query = """
        SELECT se.id, se.company_id, se.subject, se.email_body,
               c.company_name, c.email, c.contact_person
        FROM scheduled_emails se
        JOIN companies c ON se.company_id = c.id
        WHERE se.status = 'pending'
        """
        
        pending_emails = execute_query(pending_query)
        
        if not pending_emails:
            return jsonify({'success': True, 'message': 'No pending emails to send'})
        
        print(f"📧 Sending {len(pending_emails)} pending emails...")
        
        sent_count = 0
        failed_count = 0
        failed_emails = []
        
        # Connect to SMTP server once for all emails
        try:
            # Fetch current SMTP settings
            smtp = get_smtp_settings()
            
            print(f"SMTP Config: Server={smtp['smtp_server']}, Port={smtp['smtp_port']}")
            
            if smtp['smtp_port'] == 465:
                print("Using SSL connection for port 465...")
                server = smtplib.SMTP_SSL(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
            else:
                print(f"Using STARTTLS for port {smtp['smtp_port']}...")
                server = smtplib.SMTP(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
                server.starttls()
            
            # Enable verbose SMTP session logging
            try:
                server.set_debuglevel(1)
            except Exception:
                pass

            print("Logging in...")
            server.login(smtp['smtp_username'], smtp['smtp_password'])
            print("Connected successfully")
            
            for email_data in pending_emails:
                try:
                    msg = MIMEMultipart()
                    msg['From'] = f'YellowStone XPs <{smtp["smtp_username"]}>'
                    msg['To'] = email_data['email']
                    msg['Subject'] = email_data['subject']
                    
                    # Add anti-spam headers
                    add_anti_spam_headers(msg, smtp)
                    
                    # Add email body
                    msg.attach(MIMEText(email_data['email_body'], 'plain'))
                    
                    # Send email
                    text = msg.as_string()
                    refused = server.sendmail(smtp['smtp_username'], [email_data['email'], smtp['smtp_username']], text)
                    print(f"SMTP sendmail refused map: {refused}")
                    
                    # Mark as sent in database
                    update_query = """
                    UPDATE scheduled_emails 
                    SET status = 'sent', sent_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                    """
                    
                    execute_query(update_query, (email_data['id'],))
                    
                    sent_count += 1
                    print(f"✅ Email sent to {email_data['company_name']} ({email_data['email']})")
                    
                except Exception as email_error:
                    print(f"❌ Failed to send email to {email_data['email']}: {email_error}")
                    
                    # Mark as failed in database
                    update_query = """
                    UPDATE scheduled_emails 
                    SET status = 'failed', error_message = %s
                    WHERE id = %s
                    """
                    
                    execute_query(update_query, (str(email_error), email_data['id'],))
                    
                    failed_count += 1
                    failed_emails.append({
                        'company': email_data['company_name'],
                        'email': email_data['email'],
                        'error': str(email_error)
                    })
            
            server.quit()
            
            return jsonify({
                'success': True, 
                'message': f'Sent {sent_count} emails successfully, {failed_count} failed',
                'sent_count': sent_count,
                'failed_count': failed_count,
                'failed_emails': failed_emails
            })
            
        except Exception as smtp_error:
            print(f"❌ SMTP Connection Error: {smtp_error}")
            return jsonify({
                'success': False, 
                'message': f'Failed to connect to email server: {str(smtp_error)}'
            }), 500
        
    except Exception as e:
        print(f"❌ Error sending all pending emails: {e}")
        return jsonify({'error': 'Failed to send emails'}), 500


# Email threading functions
def generate_message_id():
    """Generate a unique Message-ID for email threading"""
    # Use the same domain as the configured sender for better deliverability
    domain = "yellowstonexperiences.com"
    unique_id = str(uuid.uuid4())
    return f"<{unique_id}@{domain}>"

def get_previous_email_message_id(company_id, email_type):
    """Get the Message-ID of the previous email sent to this company"""
    try:
        # Find the most recent email sent to this company
        query = """
        SELECT message_id, email_type, sent_at
        FROM scheduled_emails 
        WHERE company_id = %s AND status = 'sent' AND message_id IS NOT NULL
        ORDER BY sent_at DESC
        LIMIT 1
        """
        
        result = execute_query(query, (company_id,))
        if result and len(result) > 0:
            return result[0]['message_id']
        return None
    except Exception as e:
        # print(f"Error getting previous email Message-ID: {e}")
        return None

def get_email_thread_references(company_id):
    """Get all Message-IDs in the email thread for this company"""
    try:
        query = """
        SELECT message_id
        FROM scheduled_emails 
        WHERE company_id = %s AND status = 'sent' AND message_id IS NOT NULL
        ORDER BY sent_at ASC
        """
        
        results = execute_query(query, (company_id,))
        if results:
            return [row['message_id'] for row in results]
        return []
    except Exception as e:
        # print(f"Error getting email thread references: {e}")
        return []


# Email template system
def get_email_template(email_type, industry, company_name, contact_person):
    """Generate email template based on email type and industry"""
    
    # Extract first name from contact person or company name
    first_name = contact_person.split(' ')[0] if contact_person else company_name.split(' ')[0]
    
    # Industry-specific modifications
    industry_context = ""
    if industry.lower() in ['technology', 'it services', 'software']:
        industry_context = "technology sector"
    elif industry.lower() in ['healthcare', 'medical']:
        industry_context = "healthcare industry"
    elif industry.lower() in ['finance', 'banking', 'financial services']:
        industry_context = "financial services sector"
    elif industry.lower() in ['retail', 'ecommerce']:
        industry_context = "retail and e-commerce"
    elif industry.lower() in ['real estate', 'property']:
        industry_context = "real estate sector"
    else:
        industry_context = "your industry"
    
    templates = {
        'intro': {
            'subject': 'Strengthening your digital presence across the Middle East',
            'body': f"""Dear {first_name},

Warm greetings from Yellowstone Xps.

We specialize in helping enterprises across the Middle East enhance their digital marketing performance through data-driven strategies and ROI-focused campaigns.

Our expertise covers:
• Performance Marketing (Google, Meta, LinkedIn Ads)
• SEO & Content Strategy
• Social Media Growth & Branding
• Lead Generation & Conversion Optimization

We've supported multiple regional brands in improving online visibility, generating qualified leads, and scaling brand engagement.

Would you be open to a short conversation to explore how we can support your digital goals?

Thanks & Regards,
Saloni Sharma | BDM | YellowStone XPs - The Digital Hub
Contact: +91-6284951313 (Handphone) : +971-507054123 (WhatsApp)
www.yellowstonexps.com | contact@yellowstonexps.com
IT/ITES | Digital Services | Software Development"""
        },
        'first_followup': {
            'subject': 'Re: Following up on your digital marketing initiatives',
            'body': f"""Dear {first_name},

Just following up on my previous note — I understand how busy schedules can get, but I wanted to reconnect regarding digital marketing support from Yellowstone Xps.

We're currently helping several enterprise clients in the UAE and Saudi Arabia optimize their marketing spend and improve ROI through a blend of performance campaigns, social media, and automation tools.

Would you like me to share a few success stories or campaign case studies from similar industries?

Looking forward to your thoughts.

Thanks & Regards,
Saloni Sharma | BDM | YellowStone XPs - The Digital Hub
Contact: +91-6284951313 (Handphone) : +971-507054123 (WhatsApp)
www.yellowstonexps.com | contact@yellowstonexps.com
IT/ITES | Digital Services | Software Development"""
        },
        'second_followup': {
            'subject': 'Re: Achieving measurable ROI through digital marketing',
            'body': f"""Dear {first_name},

I hope you're doing well.

I wanted to share some insights about how many enterprises we work with have overcome challenges in generating quality leads or tracking ROI across multiple channels - areas where our team's data-driven approach has brought significant improvements.

At Yellowstone Xps, we focus on measurable growth - every campaign is optimized for performance, not just visibility.

I thought it might be valuable to discuss how we can contribute to your digital objectives. Would you be available for a brief call in the coming days?

Thanks & Regards,
Saloni Sharma | BDM | YellowStone XPs - The Digital Hub
Contact: +91-6284951313 (Handphone) : +971-507054123 (WhatsApp)
www.yellowstonexps.com | contact@yellowstonexps.com
IT/ITES | Digital Services | Software Development"""
        },
        'third_followup': {
            'subject': 'Re: Your reliable partner for digital growth in the Middle East',
            'body': f"""Dear {first_name},

I hope this message finds you well.

I wanted to share some thoughts about how having a strategic digital partner can help businesses scale faster and stay ahead of market trends, even when you have an in-house team or existing marketing vendors.

At Yellowstone Xps, we bring a full-service digital approach - combining creative storytelling with analytical precision - to help our enterprise clients reach the right audience, at the right time.

I'd be happy to share a short capability deck or client portfolio for your review if you're interested.

Thanks & Regards,
Saloni Sharma | BDM | YellowStone XPs - The Digital Hub
Contact: +91-6284951313 (Handphone) : +971-507054123 (WhatsApp)
www.yellowstonexps.com | contact@yellowstonexps.com
IT/ITES | Digital Services | Software Development"""
        },
        'final_followup': {
            'subject': 'Re: Staying connected for future digital initiatives',
            'body': f"""Dear {first_name},

I hope this message finds you well.

I understand you may not have an immediate need for digital marketing support right now - and that's perfectly fine.

At Yellowstone Xps, we regularly help enterprises across the Middle East plan and execute successful digital campaigns - from lead generation and SEO to full-scale performance marketing.

I'd be happy to stay in touch and share occasional insights or success stories that could be relevant to your future marketing plans.

Would it be alright if I kept you updated from time to time?

Thanks & Regards,
Saloni Sharma | BDM | YellowStone XPs - The Digital Hub
Contact: +91-6284951313 (Handphone) : +971-507054123 (WhatsApp)
www.yellowstonexps.com | contact@yellowstonexps.com
IT/ITES | Digital Services | Software Development"""
        }
    }
    
    return templates.get(email_type, templates['intro'])

# Background task to automatically send scheduled emails
def check_and_send_scheduled_emails():
    """Background task that checks for pending emails and sends them when their time arrives"""
    while True:
        try:
            # print("🕐 Checking for scheduled emails to send...")
            
            # Find emails that should be sent (scheduled_time <= NOW() and status = 'pending')
            # Let MySQL evaluate current time to avoid app/server timezone drift
            query = """
            SELECT se.id, se.company_id, se.subject, se.email_body, se.scheduled_time,
                   se.email_type, se.attachment_path, c.company_name, c.email, c.contact_person
            FROM scheduled_emails se
            JOIN companies c ON se.company_id = c.id
            WHERE se.status = 'pending' AND se.scheduled_time <= NOW()
            ORDER BY se.scheduled_time ASC
            """
            
            emails_to_send = execute_query(query)
            
            if emails_to_send:
                # print(f"📧 Found {len(emails_to_send)} emails ready to send")
                
                for email_data in emails_to_send:
                    try:
                        # print(f"📤 Sending email to {email_data['company_name']} ({email_data['email']})")
                        
                        # Generate unique Message-ID for this email
                        message_id = generate_message_id()
                        
                        # Get previous email Message-ID for threading
                        previous_message_id = get_previous_email_message_id(email_data['company_id'], email_data.get('email_type', 'intro'))
                        
                        # Get all Message-IDs in the thread for References header
                        thread_references = get_email_thread_references(email_data['company_id'])
                        
                        # Get SMTP settings
                        smtp = get_smtp_settings()
                        
                        # Send actual email using SMTP
                        msg = MIMEMultipart()
                        msg['From'] = f'YellowStone XPs <{smtp["smtp_username"]}>'
                        msg['To'] = email_data['email']
                        msg['Subject'] = email_data['subject']
                        msg['Message-ID'] = message_id
                        
                        # Add anti-spam headers
                        add_anti_spam_headers(msg, smtp)
                        
                        # Add threading headers for follow-up emails
                        if previous_message_id:
                            msg['In-Reply-To'] = previous_message_id
                        
                        if thread_references:
                            # Add all previous Message-IDs to References header
                            references_str = ' '.join(thread_references)
                            msg['References'] = references_str
                        
                        # Add email body
                        msg.attach(MIMEText(email_data['email_body'], 'plain'))
                        
                        # Attach file if attachment_path exists
                        if email_data.get('attachment_path'):
                            try:
                                import os
                                if os.path.exists(email_data['attachment_path']):
                                    with open(email_data['attachment_path'], 'rb') as f:
                                        attachment_data = f.read()
                                        from email.mime.base import MIMEBase
                                        from email import encoders
                                        
                                        part = MIMEBase('application', 'octet-stream')
                                        part.set_payload(attachment_data)
                                        encoders.encode_base64(part)
                                        part.add_header(
                                            'Content-Disposition',
                                            'attachment; filename= %s' % os.path.basename(email_data['attachment_path'])
                                        )
                                        msg.attach(part)
                            except Exception as attach_error:
                                # print(f"⚠️ Could not attach file: {attach_error}")
                                pass
                        
                        # Fetch current SMTP settings
                        smtp = get_smtp_settings()
                        
                        # Connect to SMTP server and send email (handle SSL vs STARTTLS)
                        if smtp['smtp_port'] == 465:
                            server = smtplib.SMTP_SSL(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
                        else:
                            server = smtplib.SMTP(smtp['smtp_server'], smtp['smtp_port'], timeout=30)
                            server.starttls()
                        try:
                            server.set_debuglevel(1)
                        except Exception:
                            pass
                        server.login(smtp['smtp_username'], smtp['smtp_password'])
                        text = msg.as_string()
                        refused = server.sendmail(smtp['smtp_username'], [email_data['email'], smtp['smtp_username']], text)
                        # Optional: log any refusal map for diagnostics
                        print(f"SMTP sendmail refused map (scheduler): {refused}")
                        server.quit()
                        
                        # print(f"✅ Email sent successfully to {email_data['email']}")
                        
                        # Mark as sent in database with Message-ID and threading info
                        update_query = """
                        UPDATE scheduled_emails 
                        SET status = 'sent', sent_at = CURRENT_TIMESTAMP, 
                            message_id = %s, in_reply_to = %s, references_header = %s
                        WHERE id = %s
                        """
                        
                        references_str = ' '.join(thread_references) if thread_references else None
                        execute_query(update_query, (message_id, previous_message_id, references_str, email_data['id'],))
                        
                        # Emit WebSocket notification for real-time update (commented out to avoid WSGI errors in background thread)
                        # socketio.emit('email_sent', {
                        #     'email_id': email_data['id'],
                        #     'company_name': email_data['company_name'],
                        #     'status': 'sent',
                        #     'sent_at': datetime.now().isoformat()
                        # }, room='admin')
                        
                    except Exception as email_error:
                        # print(f"❌ SMTP Error sending to {email_data['email']}: {email_error}")
                        
                        # Mark as failed in database
                        update_query = """
                        UPDATE scheduled_emails 
                        SET status = 'failed', error_message = %s
                        WHERE id = %s
                        """
                        
                        execute_query(update_query, (str(email_error), email_data['id'],))
                        
                        # Emit WebSocket notification for failed email (commented out to avoid WSGI errors in background thread)
                        # socketio.emit('email_failed', {
                        #     'email_id': email_data['id'],
                        #     'company_name': email_data['company_name'],
                        #     'status': 'failed',
                        #     'error_message': str(email_error)
                        # }, room='admin')
            else:
                # print("📭 No emails ready to send")
                pass
                
        except Exception as e:
            # print(f"❌ Error in email scheduler: {e}")
            pass
        
        # Wait before checking again
        time.sleep(15)

# Employee Access Management API Endpoints
@app.route('/api/admin/employee-access', methods=['GET'])
def get_employee_access():
    """Get all employees with their access permissions"""
    try:
        admin_id = session.get('user_id') or request.args.get('admin_id')
        user_type = session.get('user_type') or 'admin'  # Default to admin if no session
        if not admin_id:
            # Allow the request to proceed for admin portal
            admin_id = 1  # Use default admin ID
        
        # Get all employees
        query = """
        SELECT id, name, email, employee_id, department, designation
        FROM users
        WHERE user_type = 'employee' AND status = 'active'
        ORDER BY name ASC
        """
        employees = execute_query(query, fetch_all=True)
        
        # Get access permissions for each employee
        for employee in employees:
            access_query = """
            SELECT access_type, has_access
            FROM employee_access
            WHERE employee_id = %s
            """
            access_permissions = execute_query(access_query, (employee['id'],), fetch_all=True)
            
            # Initialize access object
            employee['access'] = {
                'send_nda': False,
                'send_leads': False,
                'hr_access': False
            }
            
            # Set access based on database
            for perm in access_permissions:
                if perm['access_type'] in employee['access']:
                    # Convert MySQL boolean (1/0) to Python boolean
                    employee['access'][perm['access_type']] = bool(perm['has_access'])
        
        return jsonify({'success': True, 'employees': employees})
        
    except Exception as e:
        # print(f"❌ Error getting employee access: {e}")
        return jsonify({'error': 'Failed to get employee access'}), 500

@app.route('/api/admin/employee-access/<int:employee_id>', methods=['PUT'])
def update_employee_access(employee_id):
    """Update access permissions for an employee"""
    try:
        # Try to get admin_id from session, request args, or JSON body
        admin_id = session.get('user_id')
        if not admin_id:
            data = request.get_json() if request.is_json else {}
            admin_id = data.get('admin_id') or request.form.get('admin_id') or request.args.get('admin_id')
        
        # Default to admin_id = 1 if not found (for development/testing)
        if not admin_id:
            admin_id = 1
        
        data = request.get_json()
        access_type = data.get('access_type')
        has_access = data.get('has_access', False)
        
        if not access_type:
            return jsonify({'error': 'Access type required'}), 400
        
        # Insert or update access
        query = """
        INSERT INTO employee_access (employee_id, access_type, has_access)
        VALUES (%s, %s, %s)
        ON DUPLICATE KEY UPDATE has_access = VALUES(has_access)
        """
        execute_query(query, (employee_id, access_type, has_access))
        
        return jsonify({'success': True, 'message': 'Access updated successfully'})
        
    except Exception as e:
        # print(f"❌ Error updating employee access: {e}")
        return jsonify({'error': 'Failed to update employee access'}), 500

@app.route('/api/admin/employee-access/bulk', methods=['PUT'])
def update_bulk_employee_access():
    """Update access permissions for multiple employees"""
    try:
        admin_id = session.get('user_id')
        if not admin_id or session.get('user_type') != 'admin':
            return jsonify({'error': 'Unauthorized'}), 401
        
        data = request.get_json()
        updates = data.get('updates', [])
        
        for update in updates:
            employee_id = update.get('employee_id')
            access_type = update.get('access_type')
            has_access = update.get('has_access', False)
            
            if not employee_id or not access_type:
                continue
            
            query = """
            INSERT INTO employee_access (employee_id, access_type, has_access)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE has_access = VALUES(has_access)
            """
            execute_query(query, (employee_id, access_type, has_access))
        
        return jsonify({'success': True, 'message': 'Access permissions updated successfully'})
        
    except Exception as e:
        # print(f"❌ Error updating bulk employee access: {e}")
        return jsonify({'error': 'Failed to update employee access'}), 500

@app.route('/api/employee/access-permissions', methods=['GET'])
def get_employee_access_permissions():
    """Get access permissions for the logged-in employee"""
    try:
        # Try to get employee_id from query parameter FIRST, then session
        employee_id = request.args.get('employee_id')
        
        # If not in query parameter, try to get from session
        if not employee_id:
            employee_id = session.get('user_id')
        
        # Convert to int if it's a string
        if employee_id:
            try:
                employee_id = int(employee_id)
            except (ValueError, TypeError):
                pass
        
        # If still no employee_id, return unauthorized
        if not employee_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        # Get access permissions for this employee
        query = """
        SELECT access_type, has_access
        FROM employee_access
        WHERE employee_id = %s
        """
        permissions = execute_query(query, (employee_id,), fetch_all=True)
        
        # Build access object
        access = {
            'send_nda': False,
            'send_leads': False,
            'hr_access': False
        }
        
        for perm in permissions:
            if perm['access_type'] in access:
                # Convert MySQL boolean (1/0) to Python boolean
                access[perm['access_type']] = bool(perm['has_access'])
        
        return jsonify({'success': True, 'access': access})
        
    except Exception as e:
        # print(f"❌ Error getting employee access: {e}")
        return jsonify({'error': 'Failed to get employee access'}), 500

# Settings API Endpoints
@app.route('/api/admin/settings', methods=['GET'])
def get_settings():
    """Get all system settings"""
    try:
        admin_id = session.get('user_id') or request.args.get('admin_id')
        user_type = session.get('user_type') or 'admin'
        if not admin_id:
            admin_id = 1  # Use default admin ID
        
        query = """
        SELECT setting_key, setting_value, setting_category, description
        FROM system_settings
        ORDER BY setting_category, setting_key
        """
        
        settings = execute_query(query)
        
        # Group settings by category
        settings_dict = {}
        for setting in settings:
            category = setting['setting_category']
            if category not in settings_dict:
                settings_dict[category] = {}
            settings_dict[category][setting['setting_key']] = setting['setting_value']
        
        return jsonify({'success': True, 'settings': settings_dict})
        
    except Exception as e:
        # print(f"❌ Error getting settings: {e}")
        return jsonify({'error': 'Failed to get settings'}), 500

@app.route('/api/admin/settings', methods=['PUT'])
def update_settings():
    """Update system settings"""
    try:
        admin_id = session.get('user_id') or request.args.get('admin_id')
        user_type = session.get('user_type') or 'admin'
        if not admin_id:
            admin_id = 1  # Use default admin ID
        
        data = request.get_json()
        settings = data.get('settings', {})
        
        # Update each setting in the database
        for category, category_settings in settings.items():
            for key, value in category_settings.items():
                query = """
                INSERT INTO system_settings (setting_key, setting_value, setting_category)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE setting_value = VALUES(setting_value)
                """
                execute_query(query, (key, str(value), category))
        
        # If email settings are updated, update the global SMTP variables
        if 'email' in settings:
            global SMTP_SERVER, SMTP_PORT, SMTP_USERNAME, SMTP_PASSWORD
            if 'smtp_server' in settings['email']:
                SMTP_SERVER = settings['email']['smtp_server']
            if 'smtp_port' in settings['email']:
                SMTP_PORT = int(settings['email']['smtp_port'])
            if 'smtp_username' in settings['email']:
                SMTP_USERNAME = settings['email']['smtp_username']
            if 'smtp_password' in settings['email']:
                SMTP_PASSWORD = settings['email']['smtp_password']
        
        return jsonify({'success': True, 'message': 'Settings updated successfully'})
        
    except Exception as e:
        # print(f"❌ Error updating settings: {e}")
        return jsonify({'error': 'Failed to update settings'}), 500

@app.route('/api/admin/settings/<string:key>', methods=['PUT'])
def update_setting(key):
    """Update a single setting"""
    try:
        admin_id = session.get('user_id') or request.args.get('admin_id')
        user_type = session.get('user_type') or 'admin'
        if not admin_id:
            admin_id = 1  # Use default admin ID
        
        data = request.get_json()
        value = data.get('value')
        category = data.get('category', 'system')
        
        query = """
        INSERT INTO system_settings (setting_key, setting_value, setting_category)
        VALUES (%s, %s, %s)
        ON DUPLICATE KEY UPDATE setting_value = VALUES(setting_value)
        """
        execute_query(query, (key, str(value), category))
        
        # Update global SMTP variables if email settings are changed
        if category == 'email':
            global SMTP_SERVER, SMTP_PORT, SMTP_USERNAME, SMTP_PASSWORD
            if key == 'smtp_server':
                SMTP_SERVER = value
            elif key == 'smtp_port':
                SMTP_PORT = int(value)
            elif key == 'smtp_username':
                SMTP_USERNAME = value
            elif key == 'smtp_password':
                SMTP_PASSWORD = value
        
        return jsonify({'success': True, 'message': 'Setting updated successfully'})
        
    except Exception as e:
        # print(f"❌ Error updating setting: {e}")
        return jsonify({'error': 'Failed to update setting'}), 500

@app.route('/api/admin/settings/test-email', methods=['POST'])
def test_email():
    """Test email configuration"""
    print("🔵 TEST EMAIL ENDPOINT CALLED")
    try:
        admin_id = session.get('user_id') or request.args.get('admin_id')
        user_type = session.get('user_type') or 'admin'
        if not admin_id:
            admin_id = 1  # Use default admin ID
        
        data = request.get_json()
        print(f"🔵 DATA RECEIVED: {data}")
        if not data:
            print("⚠️ No JSON data received in test-email request")
            return jsonify({'error': 'Request body is required'}), 400
        
        test_email_address = data.get('email')
        print(f"🔵 Test email address: {test_email_address}")
        
        if not test_email_address or test_email_address.strip() == '':
            print(f"⚠️ No email address provided in test-email request. Data received: {data}")
            return jsonify({'error': 'Please enter a test email address in the "Enter test email address" field'}), 400
        
        print("🔵 Fetching SMTP settings from database...")
        # Get current SMTP settings
        smtp_query = """
        SELECT setting_key, setting_value
        FROM system_settings
        WHERE setting_category = 'email' AND setting_key IN ('smtp_server', 'smtp_port', 'smtp_username', 'smtp_password', 'from_name')
        """
        smtp_settings = execute_query(smtp_query)
        print(f"🔵 SMTP settings fetched: {len(smtp_settings) if smtp_settings else 0} records")
        
        # Check if SMTP settings exist
        if not smtp_settings or len(smtp_settings) == 0:
            print("⚠️ No SMTP settings found in database")
            return jsonify({'error': '⚠️ SMTP settings not saved. Please click "Save Changes" first before testing.'}), 400
        
        # Build SMTP config
        smtp_config = {}
        for setting in smtp_settings:
            smtp_config[setting['setting_key']] = setting['setting_value'].strip() if setting['setting_value'] else setting['setting_value']
        print(f"🔵 Built SMTP config: {list(smtp_config.keys())}")
        
        # Check if required settings exist
        required_settings = ['smtp_server', 'smtp_port', 'smtp_username', 'smtp_password']
        print(f"🔵 Checking required settings. Values:")
        for req in required_settings:
            value = smtp_config.get(req, 'NOT SET')
            print(f"  - {req}: {value[:20] + '...' if value and len(value) > 20 else value}")
        
        missing_settings = [req for req in required_settings if not smtp_config.get(req)]
        if missing_settings:
            print(f"⚠️ Missing SMTP settings: {missing_settings}")
            return jsonify({'error': f'Missing required SMTP settings: {", ".join(missing_settings)}. Please save your settings first.'}), 400
        
        print("✅ All SMTP settings validated successfully")
        
        # Send test email
        print("🔵 Creating email message...")
        msg = MIMEMultipart()
        msg['From'] = smtp_config.get('smtp_username', SMTP_USERNAME)
        msg['To'] = test_email_address
        msg['Subject'] = 'Test Email from YellowStone XPs'
        print(f"🔵 Email From: {msg['From']}, To: {msg['To']}")
        
        body = """This is a test email from YellowStone XPs Management Portal.
        
Your email configuration is working correctly!

This email was sent to verify your SMTP settings.

Regards,
YellowStone XPs Team"""
        
        msg.attach(MIMEText(body, 'plain'))
        
        smtp_server = smtp_config.get('smtp_server', SMTP_SERVER)
        smtp_port = int(smtp_config.get('smtp_port', SMTP_PORT))
        
        # Correct port if user entered common IMAP/POP3 ports
        if smtp_port == 995 or smtp_port == 993 or smtp_port == 110 or smtp_port == 143:
            smtp_port = 587  # Default SMTP port with STARTTLS
        
        # Validate SMTP server hostname can be resolved
        print(f"🔵 Validating SMTP server hostname: {smtp_server}")
        try:
            import socket
            resolved_ip = socket.gethostbyname(smtp_server)
            print(f"✅ SMTP server resolved to: {resolved_ip}")
        except socket.gaierror as e:
            print(f"⚠️ Failed to resolve SMTP server: {e}")
            # Don't return error here - some networks have DNS issues but SMTP still works
            print(f"⚠️ Continuing anyway - will try to connect to SMTP server")
        
        # Connect to SMTP server
        print(f"Attempting to connect to {smtp_server}:{smtp_port}")
        
        # Use SSL for port 465, STARTTLS for other ports
        if smtp_port == 465:
            # SSL connection
            server = smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=30)
        else:
            # STARTTLS connection
            server = smtplib.SMTP(smtp_server, smtp_port, timeout=30)
            if smtp_port in [587, 25, 2525]:  # Only use STARTTLS for these ports
                server.starttls()
        
        server.login(smtp_config.get('smtp_username', SMTP_USERNAME), 
                    smtp_config.get('smtp_password', SMTP_PASSWORD))
        server.sendmail(smtp_config.get('smtp_username', SMTP_USERNAME), test_email_address, msg.as_string())
        server.quit()
        
        return jsonify({'success': True, 'message': f'Test email sent successfully using port {smtp_port}'})
        
    except Exception as e:
        error_message = str(e)
        print(f"Error sending test email: {error_message}")
        
        # Provide helpful error messages
        if 'getaddrinfo failed' in error_message or '11001' in error_message:
            detailed_error = (
                'Could not reach the SMTP server. '
                'Please verify:\n\n'
                '1. SMTP Server address is correct\n'
                '2. Port number is correct (587 for STARTTLS, 465 for SSL)\n'
                '3. Internet connection is working\n\n'
                'Common SMTP servers:\n'
                '- Gmail: smtp.gmail.com (port 587)\n'
                '- Outlook: smtp-mail.outlook.com (port 587)\n'
                '- Yahoo: smtp.mail.yahoo.com (port 587)\n'
                '- Custom: Check with your email provider'
            )
        elif 'authentication' in error_message.lower() or '535' in error_message:
            detailed_error = 'Authentication failed. Check your username and password (use app password for Gmail).'
        elif 'connection refused' in error_message or 'connection reset' in error_message:
            detailed_error = 'Connection refused. Check if the port is correct and if your firewall allows the connection.'
        else:
            detailed_error = f'Error: {error_message}'
        
        return jsonify({'error': detailed_error}), 500

@app.route('/api/employee/change-password', methods=['POST'])
def change_employee_password():
    """Change employee password"""
    try:
        data = request.get_json()
        employee_id = data.get('employee_id')
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        
        if not employee_id or not current_password or not new_password:
            return jsonify({'success': False, 'message': 'All fields are required'}), 400
        
        # Get user from users table
        query = "SELECT id, password_hash, user_type FROM users WHERE id = %s"
        user = execute_query(query, (employee_id,), fetch_one=True)
        
        if not user:
            return jsonify({'success': False, 'message': 'Employee not found'}), 404
        
        # Verify current password
        if not bcrypt.checkpw(current_password.encode('utf-8'), user['password_hash'].encode('utf-8')):
            return jsonify({'success': False, 'message': 'Current password is incorrect'}), 401
        
        # Hash new password
        password_hash = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        # Update password
        update_query = "UPDATE users SET password_hash = %s WHERE id = %s"
        execute_query(update_query, (password_hash, employee_id))
        
        return jsonify({'success': True, 'message': 'Password changed successfully'})
        
    except Exception as e:
        # print(f"Error changing password: {e}")
        return jsonify({'success': False, 'message': 'Failed to change password'}), 500

@app.route('/api/auth/employee-forgot-password', methods=['POST'])
def employee_forgot_password():
    """Reset employee password without requiring old password"""
    try:
        data = request.get_json()
        employee_id_str = data.get('employee_id')
        new_password = data.get('new_password')
        
        if not employee_id_str or not new_password:
            return jsonify({'success': False, 'message': 'Employee ID and new password are required'}), 400
        
        if len(new_password) < 6:
            return jsonify({'success': False, 'message': 'Password must be at least 6 characters long'}), 400
        
        # Get employee by employee_id from users table
        query = "SELECT id, user_type FROM users WHERE employee_id = %s AND user_type = 'employee'"
        user = execute_query(query, (employee_id_str,), fetch_one=True)
        
        if not user:
            return jsonify({'success': False, 'message': 'Employee ID not found'}), 404
        
        # Hash new password
        password_hash = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        # Update password in database
        update_query = "UPDATE users SET password_hash = %s WHERE employee_id = %s AND user_type = 'employee'"
        execute_query(update_query, (password_hash, employee_id_str))
        
        return jsonify({'success': True, 'message': 'Password reset successfully'})
        
    except Exception as e:
        # print(f"Error resetting password: {e}")
        return jsonify({'success': False, 'message': 'Failed to reset password'}), 500

# ==================== TENDERS MANAGEMENT ENDPOINTS ====================

@app.route('/api/admin/tenders', methods=['GET'])
def get_tenders():
    """Get all tenders"""
    try:
        tender_type = request.args.get('type', 'all')  # 'all', 'government', 'private'

        # Ensure tenders table exists with correct schema
        try:
            # First try to create the table
            execute_query(
                """
                CREATE TABLE IF NOT EXISTS tenders (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    tender_number VARCHAR(50) NOT NULL UNIQUE,
                    title VARCHAR(255) NOT NULL,
                    tender_name VARCHAR(255) NULL,
                    short_name VARCHAR(100) NULL,
                    description TEXT NULL,
                    tender_type ENUM('government', 'private') NOT NULL,
                    category VARCHAR(255) NULL,
                    organization_name VARCHAR(255) NULL,
                    budget_amount DECIMAL(15,2) NULL,
                    currency VARCHAR(3) NOT NULL DEFAULT 'INR',
                    published_date DATE NULL,
                    rfp_date DATE NULL,
                    submission_deadline DATE NULL,
                    rfq_date DATE NULL,
                    opening_date DATE NULL,
                    status VARCHAR(50) NOT NULL DEFAULT 'tender_submitted',
                    query_text TEXT NULL,
                    contact_person VARCHAR(255) NULL,
                    contact_email VARCHAR(255) NULL,
                    contact_phone VARCHAR(50) NULL,
                    location VARCHAR(255) NULL,
                    eligibility_criteria TEXT NULL,
                    documents_required TEXT NULL,
                    important_documents TEXT NULL,
                    project_coordinator_id INT NULL,
                    project_team_ids TEXT NULL,
                    created_by INT NULL,
                    updated_by INT NULL,
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (project_coordinator_id) REFERENCES users(id) ON DELETE SET NULL,
                    FOREIGN KEY (created_by) REFERENCES users(id) ON DELETE SET NULL,
                    FOREIGN KEY (updated_by) REFERENCES users(id) ON DELETE SET NULL
                )
                """
            )
            # If table exists with old schema, alter the status column
            try:
                execute_query("ALTER TABLE tenders MODIFY COLUMN status VARCHAR(50) NOT NULL DEFAULT 'tender_submitted'")
            except Exception:
                pass  # Column might already be correct or doesn't exist yet
        except Exception as e:
            print(f"Note: Tenders table may already exist or error creating it: {e}")
        
        # Create created_tenders table (no unique constraints)
        try:
            execute_query(
                """
                CREATE TABLE IF NOT EXISTS created_tenders (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    tender_number VARCHAR(50) NOT NULL,
                    title VARCHAR(255) NOT NULL,
                    tender_name VARCHAR(255) NULL,
                    short_name VARCHAR(100) NULL,
                    description TEXT NULL,
                    tender_type ENUM('government', 'private') NOT NULL,
                    category VARCHAR(255) NULL,
                    organization_name VARCHAR(255) NULL,
                    budget_amount DECIMAL(15,2) NULL,
                    currency VARCHAR(3) NOT NULL DEFAULT 'INR',
                    published_date DATE NULL,
                    rfp_date DATE NULL,
                    submission_deadline DATE NULL,
                    rfq_date DATE NULL,
                    opening_date DATE NULL,
                    status VARCHAR(50) NOT NULL DEFAULT 'tender_submitted',
                    query_text TEXT NULL,
                    contact_person VARCHAR(255) NULL,
                    contact_email VARCHAR(255) NULL,
                    contact_phone VARCHAR(50) NULL,
                    location VARCHAR(255) NULL,
                    eligibility_criteria TEXT NULL,
                    documents_required TEXT NULL,
                    important_documents TEXT NULL,
                    project_coordinator_id INT NULL,
                    project_team_ids TEXT NULL,
                    created_by INT NULL,
                    updated_by INT NULL,
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (project_coordinator_id) REFERENCES users(id) ON DELETE SET NULL,
                    FOREIGN KEY (created_by) REFERENCES users(id) ON DELETE SET NULL,
                    FOREIGN KEY (updated_by) REFERENCES users(id) ON DELETE SET NULL
                )
                """
            )
            # Ensure extra optional date columns exist
            try:
                exists_col = execute_query("""
                    SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME='created_tenders' AND COLUMN_NAME='query_submission_date'
                """, fetch_one=True) or {'cnt': 0}
                if int(exists_col.get('cnt', 0)) == 0:
                    execute_query("ALTER TABLE created_tenders ADD COLUMN query_submission_date DATE NULL AFTER opening_date")
            except Exception:
                pass
            try:
                exists_col2 = execute_query("""
                    SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME='created_tenders' AND COLUMN_NAME='appendix_ab_submission_date'
                """, fetch_one=True) or {'cnt': 0}
                if int(exists_col2.get('cnt', 0)) == 0:
                    execute_query("ALTER TABLE created_tenders ADD COLUMN appendix_ab_submission_date DATE NULL AFTER query_submission_date")
            except Exception:
                pass
            # Ensure assigned_vendor_id exists for vendor assignment (idempotent)
            try:
                col_check = execute_query(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'created_tenders'
                      AND COLUMN_NAME = 'assigned_vendor_id'
                    """,
                    fetch_one=True
                ) or { 'cnt': 0 }
                if int(col_check.get('cnt', 0)) == 0:
                    execute_query(
                        "ALTER TABLE created_tenders ADD COLUMN assigned_vendor_id INT NULL AFTER project_team_ids"
                    )

                fk_check = execute_query(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.KEY_COLUMN_USAGE
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'created_tenders'
                      AND COLUMN_NAME = 'assigned_vendor_id'
                      AND REFERENCED_TABLE_NAME = 'vendors'
                    """,
                    fetch_one=True
                ) or { 'cnt': 0 }
                if int(fk_check.get('cnt', 0)) == 0:
                    try:
                        execute_query(
                            "ALTER TABLE created_tenders ADD CONSTRAINT fk_created_tenders_vendor FOREIGN KEY (assigned_vendor_id) REFERENCES vendors(id) ON DELETE SET NULL"
                        )
                    except Exception:
                        pass
            except Exception:
                pass
        except Exception as e:
            print(f"Note: created_tenders table may already exist or error creating it: {e}")

        # Ensure tender_updates exists for logging daily updates by employees
        # No foreign key on tender_id to avoid constraint issues (we'll validate in code)
        try:
            execute_query(
                """
                CREATE TABLE IF NOT EXISTS tender_updates (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    tender_id INT NOT NULL,
                    employee_id INT NULL,
                    update_message TEXT NULL,
                    status VARCHAR(50) NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (employee_id) REFERENCES users(id) ON DELETE SET NULL
                )
                """
            )
            # Remove foreign key on tender_id if it exists (causes issues with table switching)
            try:
                # Get constraint names first
                constraints = execute_query("""
                    SELECT CONSTRAINT_NAME 
                    FROM information_schema.KEY_COLUMN_USAGE 
                    WHERE TABLE_NAME = 'tender_updates' 
                    AND TABLE_SCHEMA = DATABASE()
                    AND COLUMN_NAME = 'tender_id'
                    AND REFERENCED_TABLE_NAME IS NOT NULL
                """, fetch_all=True)
                # Drop each constraint that references tender_id
                if constraints:
                    for constraint in constraints:
                        constraint_name = constraint.get('CONSTRAINT_NAME')
                        if constraint_name:
                            try:
                                execute_query(f"ALTER TABLE tender_updates DROP FOREIGN KEY {constraint_name}")
                            except Exception:
                                pass
            except Exception:
                pass  # Table might not exist or no constraints
        except Exception:
            pass

        query = """
        SELECT 
            t.*, 
            u.name as created_by_name,
            u2.name as updated_by_name,
            ucoor.name as project_coordinator_name,
            v.company_name as assigned_vendor_name,
            DATEDIFF(t.submission_deadline, CURDATE()) as days_left,
            tu.update_message AS today_update_message,
            tu.employee_name AS today_update_by
        FROM created_tenders t
        LEFT JOIN users u ON t.created_by = u.id
        LEFT JOIN users u2 ON t.updated_by = u2.id
        LEFT JOIN users ucoor ON t.project_coordinator_id = ucoor.id
        LEFT JOIN vendors v ON t.assigned_vendor_id = v.id
        LEFT JOIN (
            SELECT 
                y.tender_id, 
                y.update_message, 
                COALESCE(u3.name, 'Unknown') AS employee_name
            FROM tender_updates y
            INNER JOIN (
                SELECT tender_id, MAX(id) AS latest_id
                FROM tender_updates
                WHERE DATE(created_at) = CURDATE()
                GROUP BY tender_id
            ) x ON y.id = x.latest_id
            LEFT JOIN users u3 ON y.employee_id = u3.id
        ) tu ON tu.tender_id = t.id
        """
        
        if tender_type != 'all':
            query += " WHERE t.tender_type = %s"
            query += " ORDER BY t.created_at DESC"
            tenders = execute_query(query, (tender_type,), fetch_all=True)
        else:
            query += " ORDER BY t.created_at DESC"
            tenders = execute_query(query, fetch_all=True)
        
        # Debug: Print first tender's update_by if exists
        if tenders and len(tenders) > 0:
            first_tender = tenders[0]
            print(f"DEBUG: First tender update_by: {first_tender.get('today_update_by')}, update_message: {first_tender.get('today_update_message')}")
        
        return jsonify(tenders if tenders else [])
        
    except Exception as e:
        print(f"Error getting tenders: {e}")
        return jsonify([])

@app.route('/api/admin/tenders/export', methods=['GET'])
def export_tenders_report():
    """Export tenders as CSV with columns required by the admin list, including today's latest update log."""
    try:
        query = """
        SELECT 
            t.id,
            t.category,
            COALESCE(t.tender_name, t.title) AS tender_name,
            t.tender_number,
            t.status,
            t.submission_deadline,
            tu.update_message AS today_update_message,
            tu.employee_name AS today_update_by,
            tu.updated_at AS today_update_at
        FROM created_tenders t
        LEFT JOIN (
            SELECT 
                y.tender_id, 
                y.update_message, 
                COALESCE(u.name, 'Unknown') AS employee_name,
                y.created_at AS updated_at
            FROM tender_updates y
            INNER JOIN (
                SELECT tender_id, MAX(id) AS latest_id
                FROM tender_updates
                WHERE DATE(created_at) = CURDATE()
                GROUP BY tender_id
            ) x ON y.id = x.latest_id
            LEFT JOIN users u ON y.employee_id = u.id
        ) tu ON tu.tender_id = t.id
        ORDER BY t.created_at DESC
        """

        rows = execute_query(query, fetch_all=True) or []

        from io import StringIO
        import csv as _csv
        import datetime as _dt

        sio = StringIO()
        writer = _csv.writer(sio)
        writer.writerow(['No.', 'Department', 'Tender Name', 'Tender Number', 'Status', 'Last Date', 'Today Update (Log)', 'Updated By', 'Log Time'])
        for i, r in enumerate(rows, start=1):
            last_date = r.get('submission_deadline')
            if hasattr(last_date, 'strftime'):
                last_date = last_date.strftime('%Y-%m-%d')
            log_time = r.get('today_update_at')
            if hasattr(log_time, 'strftime'):
                log_time = log_time.strftime('%Y-%m-%d %H:%M:%S')
            # Fallback: also include the time inline in the log cell for compatibility with older spreadsheets
            inline_log = r.get('today_update_message') or ''
            if log_time:
                inline_log = f"{inline_log} (at {log_time})" if inline_log else f"(at {log_time})"
            writer.writerow([
                i,
                r.get('category') or '',
                r.get('tender_name') or '',
                r.get('tender_number') or '',
                r.get('status') or '',
                last_date or '',
                inline_log,
                r.get('today_update_by') or '',
                log_time or ''
            ])

        csv_data = sio.getvalue()
        sio.close()

        filename = f"tenders_report_{_dt.date.today().isoformat()}.csv"
        response = make_response(csv_data)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    except Exception as e:
        print(f"Error exporting tenders: {e}")
        return jsonify({'success': False, 'message': 'Failed to export'}), 500

@app.route('/api/employee/tenders/<int:tender_id>/update', methods=['POST'])
def update_employee_tender(tender_id):
    """Allow employees to submit updates/comments for assigned tenders"""
    try:
        data = request.get_json()
        employee_id = data.get('employee_id')
        update_message = data.get('update_message', '').strip()
        status = data.get('status')
        
        if not employee_id:
            return jsonify({'success': False, 'message': 'Employee ID is required'}), 400
        
        if not update_message:
            return jsonify({'success': False, 'message': 'Update message is required'}), 400
        
        # Ensure tender_updates table exists (no foreign key on tender_id to avoid constraint issues)
        try:
            execute_query(
                """
                CREATE TABLE IF NOT EXISTS tender_updates (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    tender_id INT NOT NULL,
                    employee_id INT NULL,
                    update_message TEXT NULL,
                    status VARCHAR(50) NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (employee_id) REFERENCES users(id) ON DELETE SET NULL
                )
                """
            )
            # Remove foreign key on tender_id if it exists (causes issues with table switching)
            try:
                constraints = execute_query("""
                    SELECT CONSTRAINT_NAME 
                    FROM information_schema.KEY_COLUMN_USAGE 
                    WHERE TABLE_NAME = 'tender_updates' 
                    AND TABLE_SCHEMA = DATABASE()
                    AND COLUMN_NAME = 'tender_id'
                    AND REFERENCED_TABLE_NAME IS NOT NULL
                """, fetch_all=True)
                if constraints:
                    for constraint in constraints:
                        constraint_name = constraint.get('CONSTRAINT_NAME')
                        if constraint_name:
                            try:
                                execute_query(f"ALTER TABLE tender_updates DROP FOREIGN KEY {constraint_name}")
                            except Exception:
                                pass
            except Exception:
                pass
        except Exception:
            pass
        
        # Verify employee is assigned to this tender (in project_coordinator_id or project_team_ids)
        tender_query = "SELECT project_coordinator_id, project_team_ids FROM created_tenders WHERE id = %s"
        tender_data = execute_query(tender_query, (tender_id,), fetch_one=True)
        # Safety: some drivers or wrappers may still return a list; normalize it
        if isinstance(tender_data, list):
            tender_data = tender_data[0] if len(tender_data) > 0 else None
        
        if not tender_data:
            return jsonify({'success': False, 'message': 'Tender not found'}), 404
        
        # Check if employee is assigned
        coordinator_id = tender_data.get('project_coordinator_id')
        team_ids_json = tender_data.get('project_team_ids')
        team_ids = []
        
        if team_ids_json:
            try:
                import json
                team_ids = json.loads(team_ids_json) if isinstance(team_ids_json, str) else team_ids_json
                if not isinstance(team_ids, list):
                    team_ids = []
            except:
                team_ids = []
        
        is_assigned = (
            (coordinator_id and int(coordinator_id) == int(employee_id)) or
            (int(employee_id) in [int(tid) for tid in team_ids if tid])
        )
        
        if not is_assigned:
            return jsonify({'success': False, 'message': 'You are not assigned to this tender'}), 403
        
        # Insert update record
        insert_query = """
        INSERT INTO tender_updates (tender_id, employee_id, update_message, status)
        VALUES (%s, %s, %s, %s)
        """
        execute_query(insert_query, (tender_id, employee_id, update_message, status))
        
        # Optionally update tender status if provided
        if status:
            update_tender_query = "UPDATE created_tenders SET status = %s WHERE id = %s"
            execute_query(update_tender_query, (status, tender_id))
        
        # Broadcast database change for real-time updates
        try:
            socketio.emit('database_change', {
                'table': 'tenders',
                'action': 'update',
                'id': tender_id
            }, room='admin')
        except:
            pass
        
        return jsonify({
            'success': True,
            'message': 'Tender update submitted successfully'
        })
        
    except Exception as e:
        print(f"Error updating tender from employee: {e}")
        return jsonify({'success': False, 'message': 'Failed to submit update'}), 500

@app.route('/api/admin/tenders', methods=['POST'])
def create_tender():
    """Create a new tender"""
    try:
        data = request.get_json()
        print(f"📝 Creating tender with data: {json.dumps(data, indent=2)}")
        
        # Minimal validation with title fallback from tender_name
        if not data.get('tender_number'):
            return jsonify({'success': False, 'message': 'tender_number is required'}), 400
        effective_title = data.get('title') or data.get('tender_name')
        if not effective_title:
            return jsonify({'success': False, 'message': 'tender_name or title is required'}), 400
        
        # Default status 'new' for creation
        status_value = data.get('status') or 'new'
        
        # No uniqueness check needed - created_tenders allows duplicates
        
        # Helper functions to convert empty strings to None for optional fields
        def clean_value(val):
            if val == '' or val is None:
                return None
            return val
        
        def clean_json_value(val):
            if not val or (isinstance(val, list) and len(val) == 0):
                return None
            return json.dumps(val) if val else None
        
        def clean_budget_amount(val):
            """Convert budget_amount to float or None"""
            if not val or val == '':
                return None
            try:
                return float(val) if isinstance(val, str) else val
            except (ValueError, TypeError):
                return None
        
        # Validate foreign key references before inserting
        # IMPORTANT: Always start with None - only set if we can VERIFY the user exists
        created_by_id = None
        project_coord_id = clean_value(data.get('project_coordinator_id')) if data.get('project_coordinator_id') else None
        
        # Strict helper to verify user exists - returns user_id only if verified, None otherwise
        def verify_user_exists(user_id):
            if not user_id:
                return None
            try:
                result = execute_query("SELECT id FROM users WHERE id = %s", (int(user_id),), fetch_one=True)
                # Explicit check: result must be a dict with an 'id' key that matches
                if result and isinstance(result, dict) and result.get('id') == int(user_id):
                    return int(user_id)
                return None
            except Exception as e:
                print(f"⚠️ Error verifying user {user_id}: {e}")
                return None
        
        # Try session user first
        session_user_id = session.get('user_id')
        if session_user_id:
            verified = verify_user_exists(session_user_id)
            if verified:
                created_by_id = verified
                print(f"✅ Using verified session user_id: {created_by_id}")
            else:
                print(f"⚠️ Session user_id {session_user_id} verification failed, trying fallback...")
                created_by_id = None
        
        # Fallback: try user ID 1 ONLY if we still don't have a verified user
        if created_by_id is None:
            verified = verify_user_exists(1)
            if verified:
                created_by_id = verified
                print(f"✅ Using verified fallback user_id: {created_by_id}")
            else:
                print(f"⚠️ Fallback user_id 1 verification failed, using NULL for created_by")
                created_by_id = None
        
        # Validate project_coordinator_id
        if project_coord_id:
            verified = verify_user_exists(project_coord_id)
            if verified:
                project_coord_id = verified
                print(f"✅ Verified project_coordinator_id: {project_coord_id}")
            else:
                print(f"⚠️ project_coordinator_id {project_coord_id} verification failed, setting to NULL")
                project_coord_id = None
        
        # CRITICAL FINAL CHECK: Force None if verification fails
        if created_by_id is not None:
            final_verify = verify_user_exists(created_by_id)
            if final_verify is None:
                print(f"🚨 CRITICAL: created_by_id {created_by_id} failed final verification! Forcing to NULL")
                created_by_id = None
        
        print(f"✅✅✅ FINAL VALUES: created_by_id={created_by_id} (type={type(created_by_id).__name__}), project_coord_id={project_coord_id}")
        
        query = """
        INSERT INTO created_tenders (
            tender_number, title, tender_name, short_name, description, tender_type, category,
            organization_name, budget_amount, currency, published_date, rfp_date,
            submission_deadline, rfq_date, opening_date, query_submission_date, appendix_ab_submission_date, status, query_text, contact_person,
            contact_email, contact_phone, location, eligibility_criteria, documents_required,
            important_documents, project_coordinator_id, project_team_ids, created_by
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        )
        """
        
        params = (
            data.get('tender_number'),                                    # 1
            effective_title,                                              # 2
            clean_value(data.get('tender_name', data.get('title'))),      # 3
            clean_value(data.get('short_name')),                          # 4
            clean_value(data.get('description')),                         # 5
            data.get('tender_type'),                                      # 6
            clean_value(data.get('category')),                            # 7
            clean_value(data.get('organization_name')),                   # 8
            clean_budget_amount(data.get('budget_amount')),               # 9
            data.get('currency', 'INR'),                                  # 10
            clean_value(data.get('published_date')),                      # 11
            clean_value(data.get('rfp_date')),                            # 12
            clean_value(data.get('submission_deadline')),                 # 13
            clean_value(data.get('rfq_date', data.get('submission_deadline'))), # 14
            clean_value(data.get('opening_date')),                        # 15
            clean_value(data.get('query_submission_date')),               # 16
            clean_value(data.get('appendix_ab_submission_date')),         # 17
            status_value,                                                 # 18
            clean_value(data.get('query', data.get('query_text'))),       # 19
            clean_value(data.get('contact_person')),                      # 20
            clean_value(data.get('contact_email')),                       # 21
            clean_value(data.get('contact_phone')),                       # 22
            clean_value(data.get('location')),                            # 23
            clean_value(data.get('eligibility_criteria')),                # 24
            clean_value(data.get('documents_required')),                  # 25
            clean_json_value(data.get('important_documents')),            # 26
            project_coord_id,                                             # 27 (validated)
            clean_json_value(data.get('project_team_ids')),               # 28
            created_by_id                                                 # 29 (validated or None)
        )
        
        # Verify parameter count matches placeholders (29 expected)
        expected_params = 29
        actual_params = len(params)
        if actual_params != expected_params:
            print(f"❌ Parameter count mismatch! Expected {expected_params}, got {actual_params}")
            return jsonify({
                'success': False, 
                'message': f'Internal error: Parameter count mismatch ({actual_params} vs {expected_params})'
            }), 500
        
        # Execute INSERT query
        print(f"🔍 Executing INSERT with {len(params)} parameters")
        print(f"📋 First few params: tender_number={params[0]}, title={params[1]}, type={params[5]}")
        print(f"🔑 Foreign keys in params: project_coordinator_id={params[24]}, created_by={params[26]}")
        result = execute_query(query, params)
        
        # Check if INSERT failed (execute_query returns None on error)
        if result is None:
            print("❌ INSERT query returned None - check error logs above")
            return jsonify({'success': False, 'message': 'Failed to insert tender into database. Check server logs for details.'}), 500
        
        # Fetch the created tender - get the most recent one with this tender_number (since duplicates allowed)
        created_tender = execute_query(
            """
            SELECT 
                t.*, 
                u.name as created_by_name,
                DATEDIFF(t.submission_deadline, CURDATE()) as days_left
            FROM created_tenders t
            LEFT JOIN users u ON t.created_by = u.id
            WHERE t.tender_number = %s
            ORDER BY t.created_at DESC
            LIMIT 1
            """,
            (data.get('tender_number'),),
            fetch_one=True
        )

        # Broadcast database change (id may be null if not fetched, but the admin can reload list)
        try:
            socketio.emit('database_change', {
                'table': 'tenders',
                'action': 'create',
                'id': created_tender.get('id') if created_tender else None
            }, room='admin')
        except:
            pass
        
        return jsonify({
            'success': True, 
            'message': 'Tender created successfully',
            'tender': created_tender if created_tender else None
        })
        
    except Exception as e:
        error_msg = str(e)
        print(f"Error creating tender: {error_msg}")
        # Provide more specific error messages
        if 'Duplicate entry' in error_msg or 'UNIQUE' in error_msg:
            return jsonify({'success': False, 'message': 'Tender number must be unique'}), 400
        elif 'Table' in error_msg and "doesn't exist" in error_msg:
            return jsonify({'success': False, 'message': 'Database table not initialized. Please refresh the page and try again.'}), 500
        else:
            return jsonify({'success': False, 'message': f'Failed to create tender: {error_msg}'}), 500

@app.route('/api/admin/tenders/<int:tender_id>', methods=['PUT'])
def update_tender(tender_id):
    """Update a tender"""
    try:
        data = request.get_json()
        
        query = """
        UPDATE created_tenders 
        SET title = %s, tender_name = %s, short_name = %s, description = %s, tender_type = %s, category = %s,
            organization_name = %s, budget_amount = %s, currency = %s,
            published_date = %s, rfp_date = %s, submission_deadline = %s, rfq_date = %s, opening_date = %s,
            query_submission_date = %s, appendix_ab_submission_date = %s,
            status = %s, query_text = %s, contact_person = %s, contact_email = %s,
            contact_phone = %s, location = %s, eligibility_criteria = %s,
            documents_required = %s, important_documents = %s,
            project_coordinator_id = %s, project_team_ids = %s, updated_by = %s
        WHERE id = %s
        """
        
        params = (
            (data.get('title') or data.get('tender_name')),
            data.get('tender_name', data.get('title')),
            data.get('short_name', ''),
            data.get('description', ''),
            data.get('tender_type'),
            data.get('category', ''),
            data.get('organization_name', ''),
            data.get('budget_amount'),
            data.get('currency', 'INR'),
            data.get('published_date'),
            data.get('rfp_date'),
            data.get('submission_deadline'),
            data.get('rfq_date', data.get('submission_deadline')),
            data.get('opening_date'),
            data.get('query_submission_date'),
            data.get('appendix_ab_submission_date'),
            data.get('status'),
            data.get('query', data.get('query_text', '')),
            data.get('contact_person', ''),
            data.get('contact_email', ''),
            data.get('contact_phone', ''),
            data.get('location', ''),
            data.get('eligibility_criteria', ''),
            data.get('documents_required', ''),
            json.dumps(data.get('important_documents', [])),
            data.get('project_coordinator_id'),
            json.dumps(data.get('project_team_ids', [])),
            1,  # updated_by - admin ID
            tender_id
        )
        
        execute_query(query, params)
        return jsonify({'success': True, 'message': 'Tender updated successfully'})
        
    except Exception as e:
        print(f"Error updating tender: {e}")
        return jsonify({'success': False, 'message': 'Failed to update tender'}), 500

@app.route('/api/admin/tenders/<int:tender_id>', methods=['DELETE'])
def delete_tender(tender_id):
    """Delete a tender"""
    try:
        query = "DELETE FROM created_tenders WHERE id = %s"
        execute_query(query, (tender_id,))
        return jsonify({'success': True, 'message': 'Tender deleted successfully'})
        
    except Exception as e:
        print(f"Error deleting tender: {e}")
        return jsonify({'success': False, 'message': 'Failed to delete tender'}), 500

@app.route('/api/admin/tenders/<int:tender_id>/extensions', methods=['GET', 'POST', 'OPTIONS'])
def manage_tender_extensions(tender_id):
    """Create and list tender date extensions. If a date field is extended, update the tender record too."""
    try:
        # Ensure log table exists
        try:
            execute_query(
                """
                CREATE TABLE IF NOT EXISTS tender_extensions (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    tender_id INT NOT NULL,
                    date_field VARCHAR(50) NOT NULL,
                    old_date DATE NULL,
                    new_date DATE NULL,
                    reason TEXT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (tender_id) REFERENCES created_tenders(id) ON DELETE CASCADE
                )
                """
            )
        except Exception:
            pass

        # CORS preflight
        if request.method == 'OPTIONS':
            return make_response(('', 204))

        if request.method == 'GET':
            rows = execute_query(
                "SELECT id, date_field, old_date, new_date, reason, created_at FROM tender_extensions WHERE tender_id = %s ORDER BY created_at DESC",
                (tender_id,), fetch_all=True
            ) or []
            return jsonify(rows)

        data = request.get_json() or {}
        date_field = (data.get('date_field') or '').strip()
        new_date = data.get('new_date')
        reason = data.get('reason', '')

        allowed_fields = {
            'published_date', 'query_submission_date', 'appendix_ab_submission_date',
            'rfp_date', 'submission_deadline', 'rfq_date', 'opening_date'
        }
        if date_field not in allowed_fields:
            return jsonify({'success': False, 'message': 'Invalid date field'}), 400
        if not new_date:
            return jsonify({'success': False, 'message': 'new_date is required'}), 400

        # Fetch old value
        old = execute_query(f"SELECT {date_field} AS d FROM created_tenders WHERE id = %s", (tender_id,), fetch_one=True)
        old_date = old.get('d') if old else None

        # Update tender date
        execute_query(f"UPDATE created_tenders SET {date_field} = %s WHERE id = %s", (new_date, tender_id))

        # Insert log
        execute_query(
            "INSERT INTO tender_extensions (tender_id, date_field, old_date, new_date, reason) VALUES (%s, %s, %s, %s, %s)",
            (tender_id, date_field, old_date, new_date, reason)
        )

        # Broadcast change for admin views
        try:
            socketio.emit('database_change', {'table': 'tenders', 'action': 'update', 'id': tender_id}, room='admin')
        except:
            pass

        return jsonify({'success': True, 'message': 'Extension logged and date updated'})
    except Exception as e:
        print(f"Extensions endpoint error: {e}")
        return jsonify({'success': False, 'message': 'Failed to process extension'}), 500

@app.route('/api/admin/tenders/<int:tender_id>/assign-vendor', methods=['POST', 'OPTIONS'])
def assign_vendor_to_tender(tender_id):
    """Assign a vendor to a tender (stores vendor id on created_tenders)."""
    try:
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            return make_response(('', 204))
        data = request.get_json() or {}
        vendor_id = data.get('vendor_id')
        if not vendor_id:
            return jsonify({'success': False, 'message': 'vendor_id is required'}), 400

        # Verify vendor exists
        vendor = execute_query("SELECT id FROM vendors WHERE id = %s", (vendor_id,), fetch_one=True)
        if not vendor:
            return jsonify({'success': False, 'message': 'Vendor not found'}), 404

        # Column expected to exist; avoid noisy ALTER attempts in this endpoint

        # Perform assignment
        update_query = "UPDATE created_tenders SET assigned_vendor_id = %s, updated_at = NOW() WHERE id = %s"
        execute_query(update_query, (vendor_id, tender_id))

        # Optional: return vendor details for UI
        details = execute_query("SELECT id, assigned_vendor_id FROM created_tenders WHERE id = %s", (tender_id,), fetch_one=True)
        return jsonify({'success': True, 'message': 'Vendor assigned successfully', 'tender': details})
    except Exception as e:
        print(f"Error assigning vendor to tender: {e}")
        return jsonify({'success': False, 'message': 'Failed to assign vendor'}), 500

# Start the background email scheduler
email_scheduler_thread = threading.Thread(target=check_and_send_scheduled_emails, daemon=True)
email_scheduler_thread.start()
# print("🚀 Email scheduler started - checking every 30 seconds")

if __name__ == '__main__':
    print("Starting Flask server with WebSocket support...")
    socketio.run(app, debug=False, host='0.0.0.0', port=8000)
